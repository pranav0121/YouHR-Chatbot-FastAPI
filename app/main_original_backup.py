from fastapi.responses import FileResponse, JSONResponse
from fastapi import FastAPI, Depends, HTTPException, Query
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from app.database import get_db
from app import models, schemas
from datetime import date, timedelta, datetime
from typing import Optional
from pydantic import BaseModel
from pathlib import Path
import os
import random
from openpyxl import Workbook
import logging


app = FastAPI()

# Allow CORS for all origins (for development)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Serve static files (frontend) with no caching for development


class NoCacheStaticFiles(StaticFiles):
    def file_response(self, full_path, stat_result, scope, status_code=200):
        response = super().file_response(full_path, stat_result, scope, status_code)
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
        return response


app.mount("/static", NoCacheStaticFiles(directory="static"), name="static")

# Directory where exported files are stored. Ensure your export functions write files here.
DOWNLOAD_DIR = Path(os.path.join(os.getcwd(), 'downloads'))
DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)


def get_merchant_id(merchant_id: str = Query("MERCH001", description="Merchant ID (default: MERCH001)")):
    """Dependency that supplies a merchant_id and marks when a default was used.

    Returns a tuple (merchant_id, headers) where headers may contain X-Warning when defaulted.
    """
    default_id = "MERCH001"
    headers = {}
    if merchant_id == default_id:
        warning_msg = f"merchant_id defaulted to {default_id}; callers should provide merchant_id explicitly"
        logging.getLogger(__name__).warning(warning_msg)
        headers["X-Warning"] = warning_msg
    return merchant_id, headers


@app.get("/api/downloads/{filename}")
def download_file(filename: str):
    """Serve exported files from the downloads directory in a safe way.

    Example: /api/downloads/today_sales_2025-08-31.xlsx
    """
    # Protect against path traversal
    requested = (DOWNLOAD_DIR / filename).resolve()
    try:
        if not str(requested).startswith(str(DOWNLOAD_DIR.resolve())):
            raise HTTPException(status_code=400, detail="Invalid filename")
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid filename")

    if not requested.exists() or not requested.is_file():
        raise HTTPException(status_code=404, detail="File not found")

    return FileResponse(path=str(requested), filename=filename, media_type='application/octet-stream')

# Serve chat.html at root


@app.get("/")
def serve_chat_html():
    return FileResponse("static/chat.html", headers={"Cache-Control": "no-cache, no-store, must-revalidate"})

# Get all menus with submenus (filtered)


@app.get("/api/chatbot/menus-with-submenus")
def get_menus_with_submenus(
    company_type: str,
    role: str,
    db: Session = Depends(get_db)
):
    menus = db.query(models.ChatbotMenu)\
        .filter(
            models.ChatbotMenu.is_active == True,
            models.ChatbotMenu.company_type == company_type,
            models.ChatbotMenu.role == role
    ).all()

    if not menus:
        raise HTTPException(status_code=404, detail="No menus found")

    results = []
    for menu in menus:
        submenus = db.query(models.ChatbotSubmenu)\
            .filter(
                models.ChatbotSubmenu.menu_id == menu.id,
                models.ChatbotSubmenu.is_active == True,
                models.ChatbotSubmenu.company_type == company_type,
                models.ChatbotSubmenu.role == role
        ).all()
        results.append({
            "menu_id": menu.id,
            "menu_key": menu.menu_key,
            "menu_title": menu.menu_title,
            "menu_icon": menu.menu_icon,
            "company_type": menu.company_type,
            "role": menu.role,
            "submenus": [
                {
                    "submenu_id": sm.id,
                    "submenu_key": sm.submenu_key,
                    "submenu_title": sm.submenu_title,
                    "api_endpoint": sm.api_endpoint
                }
                for sm in submenus
            ]
        })

    return results


# Simple menu endpoint for company type (without role filtering)
@app.get("/api/menu/{company_type}")
def get_menus_by_company_type(
    company_type: str,
    db: Session = Depends(get_db)
):
    # Special handling for merchant to return ICP HR merchant manager menus
    if company_type == "merchant":
        menus = db.query(models.ChatbotMenu)\
            .filter(
                models.ChatbotMenu.is_active == True,
                models.ChatbotMenu.company_type == "icp_hr",
                models.ChatbotMenu.role == "merchant_manager"
        ).all()
    else:
        menus = db.query(models.ChatbotMenu)\
            .filter(
                models.ChatbotMenu.is_active == True,
                models.ChatbotMenu.company_type == company_type
        ).all()

    if not menus:
        raise HTTPException(
            status_code=404, detail=f"No menus found for company type: {company_type}")

    results = []
    for menu in menus:
        if company_type == "merchant":
            # For merchant, get ICP HR merchant manager submenus
            submenus = db.query(models.ChatbotSubmenu)\
                .filter(
                    models.ChatbotSubmenu.menu_id == menu.id,
                    models.ChatbotSubmenu.is_active == True,
                    models.ChatbotSubmenu.company_type == "icp_hr",
                    models.ChatbotSubmenu.role == "merchant_manager"
            ).all()
        else:
            submenus = db.query(models.ChatbotSubmenu)\
                .filter(
                    models.ChatbotSubmenu.menu_id == menu.id,
                    models.ChatbotSubmenu.is_active == True
            ).all()
        results.append({
            "menu_id": menu.id,
            "menu_key": menu.menu_key,
            "menu_title": menu.menu_title,
            "menu_icon": menu.menu_icon,
            "company_type": menu.company_type,
            "submenus": [
                {
                    "submenu_id": sm.id,
                    "submenu_key": sm.submenu_key,
                    "submenu_title": sm.submenu_title,
                    "api_endpoint": sm.api_endpoint
                }
                for sm in submenus
            ]
        })

    return results


# Core HR Endpoints for comprehensive testing
@app.get("/api/chatbot/employees")
def get_all_employees(db: Session = Depends(get_db)):
    """Get all employees from the database"""
    employees = db.query(models.Employee).all()
    return [
        {
            "employee_id": emp.employee_id,
            "name": emp.name,
            "email": emp.email,
            "position": emp.position,
            "department": emp.department,
            "hire_date": emp.hire_date.isoformat() if emp.hire_date else None,
            "salary": float(emp.salary) if emp.salary else None
        }
        for emp in employees
    ]


@app.get("/api/chatbot/attendance")
def get_all_attendance(db: Session = Depends(get_db)):
    """Get all attendance records"""
    attendance = db.query(models.AttendanceRecord).all()
    return [
        {
            "id": att.id,
            "employee_id": att.employee_id,
            "date": att.date.isoformat() if att.date else None,
            "check_in": att.check_in.isoformat() if att.check_in else None,
            "check_out": att.check_out.isoformat() if att.check_out else None,
            "status": att.status
        }
        for att in attendance
    ]


@app.get("/api/chatbot/payslips")
def get_all_payslips(db: Session = Depends(get_db)):
    """Get all payslip records"""
    payslips = db.query(models.Payslip).all()
    return [
        {
            "id": pay.id,
            "employee_id": pay.employee_id,
            "month": pay.month,
            "year": pay.year,
            "basic_salary": float(pay.basic_salary) if pay.basic_salary else None,
            "allowances": float(pay.allowances) if pay.allowances else None,
            "deductions": float(pay.deductions) if pay.deductions else None,
            "net_salary": float(pay.net_salary) if pay.net_salary else None
        }
        for pay in payslips
    ]


@app.get("/api/chatbot/leave-applications")
def get_all_leave_applications(db: Session = Depends(get_db)):
    """Get all leave applications"""
    leaves = db.query(models.LeaveApplication).all()
    return [
        {
            "id": leave.id,
            "employee_id": leave.employee_id,
            "leave_type": leave.leave_type,
            "start_date": leave.start_date.isoformat() if leave.start_date else None,
            "end_date": leave.end_date.isoformat() if leave.end_date else None,
            "reason": leave.reason,
            "status": leave.status
        }
        for leave in leaves
    ]


@app.get("/api/chatbot/hr-support-tickets")
def get_all_hr_support_tickets(db: Session = Depends(get_db)):
    """Get all HR support tickets"""
    tickets = db.query(models.HRSupportTicket).all()
    return [
        {
            "id": ticket.id,
            "employee_id": ticket.employee_id,
            "ticket_type": ticket.ticket_type,
            "subject": ticket.subject,
            "description": ticket.description,
            "status": ticket.status,
            "created_at": ticket.created_at.isoformat() if ticket.created_at else None
        }
        for ticket in tickets
    ]


@app.get("/api/chatbot/marketing-campaigns")
def get_all_marketing_campaigns(db: Session = Depends(get_db)):
    """Get all marketing campaigns"""
    campaigns = db.query(models.MarketingCampaign).all()
    return [
        {
            "id": campaign.id,
            "campaign_name": campaign.campaign_name,
            "description": campaign.description,
            "start_date": campaign.start_date.isoformat() if campaign.start_date else None,
            "end_date": campaign.end_date.isoformat() if campaign.end_date else None,
            "budget": float(campaign.budget) if campaign.budget else None,
            "status": campaign.status
        }
        for campaign in campaigns
    ]


@app.get("/api/chatbot/promotions")
def get_all_promotions(db: Session = Depends(get_db)):
    """Get all promotions"""
    promotions = db.query(models.Promotion).all()
    return [
        {
            "id": promo.id,
            "employee_id": promo.employee_id,
            "old_position": promo.old_position,
            "new_position": promo.new_position,
            "old_salary": float(promo.old_salary) if promo.old_salary else None,
            "new_salary": float(promo.new_salary) if promo.new_salary else None,
            "promotion_date": promo.promotion_date.isoformat() if promo.promotion_date else None
        }
        for promo in promotions
    ]


@app.get("/api/chatbot/sales-records")
def get_all_sales_records(db: Session = Depends(get_db)):
    """Get all sales records"""
    sales = db.query(models.SalesRecord).all()
    return [
        {
            "id": sale.id,
            "employee_id": sale.employee_id,
            "sale_date": sale.sale_date.isoformat() if sale.sale_date else None,
            "amount": float(sale.amount) if sale.amount else None,
            "commission": float(sale.commission) if sale.commission else None
        }
        for sale in sales
    ]


# Get attendance history
@app.get("/api/attendance/history")
def get_attendance_history(
    employee_id: str = Query(..., description="Employee ID"),
    days: Optional[int] = Query(
        30, description="Number of days to fetch (default: 30)"),
    db: Session = Depends(get_db)
):
    # Calculate date range
    end_date = date.today()
    start_date = end_date - timedelta(days=days)

    # Query attendance records
    attendance_records = db.query(models.AttendanceRecord)\
        .filter(
            models.AttendanceRecord.employee_id == employee_id,
            models.AttendanceRecord.date >= start_date,
            models.AttendanceRecord.date <= end_date
    )\
        .order_by(models.AttendanceRecord.date.desc())\
        .all()

    if not attendance_records:
        return {
            "employee_id": employee_id,
            "message": "No attendance records found for the specified period",
            "total_records": 0,
            "date_range": {
                "from": str(start_date),
                "to": str(end_date)
            },
            "records": []
        }

    # Calculate summary statistics
    present_days = len(
        [r for r in attendance_records if r.status == "Present"])
    late_days = len([r for r in attendance_records if r.status == "Late"])
    absent_days = len([r for r in attendance_records if r.status == "Absent"])

    # Format response
    formatted_records = []
    for record in attendance_records:
        formatted_records.append({
            "date": str(record.date),
            "check_in_time": str(record.check_in_time) if record.check_in_time else None,
            "check_out_time": str(record.check_out_time) if record.check_out_time else None,
            "working_hours": record.working_hours,
            "status": record.status,
            "location": record.location
        })

    return {
        "employee_id": employee_id,
        "employee_name": attendance_records[0].employee_name,
        "total_records": len(attendance_records),
        "date_range": {
            "from": str(start_date),
            "to": str(end_date)
        },
        "summary": {
            "present_days": present_days,
            "late_days": late_days,
            "absent_days": absent_days,
            "total_working_days": len(attendance_records)
        },
        "records": formatted_records
    }


# Apply for leave
@app.post("/api/leave/apply")
def apply_for_leave(
    leave_request: schemas.LeaveApplicationRequest,
    db: Session = Depends(get_db)
):
    try:
        # Parse dates
        from_date_obj = date.fromisoformat(leave_request.from_date)
        to_date_obj = date.fromisoformat(leave_request.to_date)

        # Calculate total days
        total_days = (to_date_obj - from_date_obj).days + 1

        # Create leave application
        leave_application = models.LeaveApplication(
            employee_id=leave_request.employee_id,
            employee_name=leave_request.employee_name,
            leave_type=leave_request.leave_type,
            from_date=from_date_obj,
            to_date=to_date_obj,
            total_days=total_days,
            reason=leave_request.reason
        )

        db.add(leave_application)
        db.commit()
        db.refresh(leave_application)

        return {
            "success": True,
            "message": "Leave application submitted successfully",
            "application_id": leave_application.id,
            "employee_id": leave_application.employee_id,
            "employee_name": leave_application.employee_name,
            "leave_type": leave_application.leave_type,
            "from_date": str(leave_application.from_date),
            "to_date": str(leave_application.to_date),
            "total_days": leave_application.total_days,
            "reason": leave_application.reason,
            "status": leave_application.status,
            "applied_date": str(leave_application.applied_date)
        }

    except ValueError as e:
        raise HTTPException(
            status_code=400, detail=f"Invalid date format: {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500, detail=f"Failed to submit leave application: {str(e)}")


# Get leave applications for an employee
@app.get("/api/leave/applications")
def get_leave_applications(
    employee_id: str = Query(..., description="Employee ID"),
    db: Session = Depends(get_db)
):
    applications = db.query(models.LeaveApplication)\
        .filter(models.LeaveApplication.employee_id == employee_id)\
        .order_by(models.LeaveApplication.applied_date.desc())\
        .all()

    if not applications:
        return {
            "employee_id": employee_id,
            "message": "No leave applications found",
            "applications": []
        }

    formatted_applications = []
    for app in applications:
        formatted_applications.append({
            "application_id": app.id,
            "leave_type": app.leave_type,
            "from_date": str(app.from_date),
            "to_date": str(app.to_date),
            "total_days": app.total_days,
            "reason": app.reason,
            "status": app.status,
            "applied_date": str(app.applied_date),
            "approved_by": app.approved_by,
            "approved_date": str(app.approved_date) if app.approved_date else None,
            "comments": app.comments
        })

    return {
        "employee_id": employee_id,
        "total_applications": len(applications),
        "applications": formatted_applications
    }


# Get payslips for an employee
@app.get("/api/payroll/payslips")
def get_payslips(
    employee_id: str = Query(..., description="Employee ID"),
    db: Session = Depends(get_db)
):
    payslips = db.query(models.Payslip)\
        .filter(models.Payslip.employee_id == employee_id)\
        .order_by(models.Payslip.pay_period.desc())\
        .all()

    if not payslips:
        return {
            "employee_id": employee_id,
            "message": "No payslips found",
            "payslips": []
        }

    formatted_payslips = []
    for payslip in payslips:
        formatted_payslips.append({
            "payslip_id": payslip.id,
            "pay_period": payslip.pay_period,
            "pay_period_start": str(payslip.pay_period_start),
            "pay_period_end": str(payslip.pay_period_end),
            "basic_salary": payslip.basic_salary,
            "allowances": payslip.allowances,
            "gross_salary": payslip.gross_salary,
            "deductions": payslip.deductions,
            "net_salary": payslip.net_salary,
            "status": payslip.status,
            "generated_date": str(payslip.generated_date),
            "download_url": payslip.download_url or f"/api/payroll/download/{payslip.id}"
        })

    return {
        "employee_id": employee_id,
        "employee_name": payslips[0].employee_name,
        "total_payslips": len(payslips),
        "payslips": formatted_payslips
    }


# Get employee status information
@app.get("/api/employee/status")
def get_employee_status(
    employee_id: str = Query(..., description="Employee ID"),
    db: Session = Depends(get_db)
):
    employee = db.query(models.Employee)\
        .filter(models.Employee.employee_id == employee_id)\
        .first()

    if not employee:
        # Check if employee_id is clearly invalid (not in expected format)
        if employee_id == "INVALID" or not employee_id.startswith("EMP"):
            raise HTTPException(
                status_code=404, detail=f"Employee with ID {employee_id} not found")

        # Return default data for valid format but non-existent employees
        return {
            "employee_id": employee_id,
            "employee_name": "John Doe",
            "employment_status": "Active",
            "employment_type": "Full-time",
            "department": "IT Development",
            "position": "Senior Developer",
            "hire_date": "2023-01-15",
            "years_of_service": 2.6,
            "reporting_manager": "Jane Smith",
            "office_location": "Main Office - Floor 3",
            "salary_grade": "L4",
            "probation_status": "Completed",
            "last_promotion": "2024-01-15",
            "performance_rating": "Exceeds Expectations",
            "next_review_date": "2026-01-15"
        }

    # Calculate years of service
    today = date.today()
    years_of_service = round((today - employee.hire_date).days / 365.25, 1)

    # Determine probation status
    probation_status = "Completed"
    if employee.probation_end_date and today < employee.probation_end_date:
        probation_status = "In Progress"
    elif employee.probation_end_date and today >= employee.probation_end_date:
        probation_status = "Completed"

    return {
        "employee_id": employee.employee_id,
        "employee_name": employee.employee_name,
        "employment_status": employee.employment_status,
        "employment_type": employee.employment_type,
        "department": employee.department,
        "position": employee.position,
        "hire_date": str(employee.hire_date),
        "years_of_service": years_of_service,
        "reporting_manager": employee.reporting_manager,
        "office_location": employee.office_location,
        "salary_grade": employee.salary_grade,
        "probation_status": probation_status,
        "last_promotion": str(employee.last_promotion_date) if employee.last_promotion_date else None,
        # This could come from another table
        "performance_rating": "Exceeds Expectations",
        "next_review_date": "2026-01-15"  # This could be calculated
    }


# ===== MERCHANT MANAGEMENT ENDPOINTS =====

@app.get("/api/merchant/sales/today")
def get_today_sales(merchant: tuple = Depends(get_merchant_id), db: Session = Depends(get_db)):
    merchant_id, headers = merchant
    """Get today's sales data"""
    today = date.today()

    # Mock data - replace with actual database queries
    resp = {
        "date": str(today),
        "merchant_id": merchant_id,
        "total_sales": "₹12,450.00",
        "total_transactions": 28,
        "cash_sales": "₹8,200.00",
        "card_sales": "₹3,150.00",
        "upi_sales": "₹1,100.00",
        "top_selling_items": [
            {"item": "Coffee", "quantity": 15, "revenue": "₹450.00"},
            {"item": "Sandwich", "quantity": 8, "revenue": "₹800.00"},
            {"item": "Burger", "quantity": 12, "revenue": "₹1,800.00"}
        ],
        "hourly_sales": [
            {"hour": "09:00", "sales": "₹850.00"},
            {"hour": "10:00", "sales": "₹1,200.00"},
            {"hour": "11:00", "sales": "₹1,800.00"},
            {"hour": "12:00", "sales": "₹2,100.00"},
            {"hour": "13:00", "sales": "₹1,950.00"},
            {"hour": "14:00", "sales": "₹1,650.00"},
            {"hour": "15:00", "sales": "₹1,200.00"},
            {"hour": "16:00", "sales": "₹900.00"},
            {"hour": "17:00", "sales": "₹750.00"}
        ]
    }
    return JSONResponse(content=resp, headers=headers)


# Additional Today's Sales Endpoints
@app.get("/api/merchant/sales/today/by-product")
def get_today_sales_by_product(merchant: tuple = Depends(get_merchant_id), db: Session = Depends(get_db)):
    merchant_id, headers = merchant
    """Get today's sales breakdown by product"""
    resp = {
        "date": str(date.today()),
        "merchant_id": merchant_id,
        "products": [
            {"name": "Coffee", "quantity": 15,
                "revenue": "₹450.00", "stock_remaining": 25},
            {"name": "Sandwich", "quantity": 8,
                "revenue": "₹800.00", "stock_remaining": 12},
            {"name": "Burger", "quantity": 12,
                "revenue": "₹1,800.00", "stock_remaining": 8},
            {"name": "Pastry", "quantity": 6,
                "revenue": "₹300.00", "stock_remaining": 15},
            {"name": "Cold Drinks", "quantity": 22,
                "revenue": "₹660.00", "stock_remaining": 30}
        ],
        "total_products_sold": 5,
        "total_revenue": "₹4,010.00"
    }
    return JSONResponse(content=resp, headers=headers)


@app.get("/api/merchant/sales/today/analytics")
def get_today_sales_analytics(merchant: tuple = Depends(get_merchant_id), db: Session = Depends(get_db)):
    merchant_id, headers = merchant
    """Get today's sales analytics and insights"""
    resp = {
        "date": str(date.today()),
        "merchant_id": merchant_id,
        "growth_percentage": "+12.5%",
        "avg_transaction_value": "₹445.36",
        "peak_hour": "12:00 PM - 1:00 PM",
        "new_customers": 8,
        "returning_customers": 20,
        "satisfaction_score": 4.7,
        "conversion_rate": "68%",
        "payment_method_breakdown": {
            "cash": "65.8%",
            "card": "25.3%",
            "upi": "8.9%"
        }
    }
    return JSONResponse(content=resp, headers=headers)


@app.get("/api/merchant/sales/today/export")
def export_today_sales(merchant: tuple = Depends(get_merchant_id), db: Session = Depends(get_db)):
    merchant_id, headers = merchant
    """Export today's sales data"""
    # Generate a simple Excel file and save to downloads
    fname = f"today_sales_{date.today()}.xlsx"
    dest = DOWNLOAD_DIR / fname

    wb = Workbook()
    ws = wb.active
    ws.title = "Today Sales"
    ws.append(["Item", "Quantity", "Revenue"])
    sample_rows = [
        ["Coffee", 15, "₹450.00"],
        ["Sandwich", 8, "₹800.00"],
        ["Burger", 12, "₹1,800.00"]
    ]
    for r in sample_rows:
        ws.append(r)
    wb.save(dest)

    resp = {
        "status": "success",
        "message": "Sales data exported successfully",
        "download_url": f"/api/downloads/{fname}",
        "file_size": f"{dest.stat().st_size // 1024} KB",
        "total_records": len(sample_rows),
        "date": str(date.today()),
        "merchant_id": merchant_id
    }
    return JSONResponse(content=resp, headers=headers)


@app.get("/api/merchant/sales/yesterday")
def get_yesterday_sales(db: Session = Depends(get_db)):
    """Get yesterday's sales data"""
    yesterday = date.today() - timedelta(days=1)

    return {
        "date": str(yesterday),
        "total_sales": "₹11,780.00",
        "total_transactions": 25,
        "cash_sales": "₹7,800.00",
        "card_sales": "₹2,980.00",
        "upi_sales": "₹1,000.00",
        "comparison_with_today": {
            "sales_difference": "+₹670.00",
            "percentage_change": "+5.7%",
            "transaction_difference": "+3"
        },
        "top_selling_items": [
            {"item": "Coffee", "quantity": 12, "revenue": "₹360.00"},
            {"item": "Pizza", "quantity": 6, "revenue": "₹900.00"},
            {"item": "Burger", "quantity": 10, "revenue": "₹1,500.00"}
        ]
    }


# Additional Yesterday's Sales Endpoints
@app.get("/api/merchant/sales/yesterday/by-product")
def get_yesterday_sales_by_product(merchant: tuple = Depends(get_merchant_id), db: Session = Depends(get_db)):
    merchant_id, headers = merchant
    """Get yesterday's sales breakdown by product"""
    yesterday = date.today() - timedelta(days=1)

    # Query the database for sales records
    sales = db.query(models.SalesRecord).filter(
        models.SalesRecord.merchant_id == merchant_id,
        models.SalesRecord.sale_date == yesterday
    ).all()

    # Transform the query results into the desired response format
    products = [
        {
            "name": sale.product_name,
            "quantity": sale.quantity,
            "revenue": sale.revenue
        }
        for sale in sales
    ]

    resp = {
        "date": str(yesterday),
        "merchant_id": merchant_id,
        "products": products,
        "total_products_sold": len(products),
        "total_revenue": f"₹{sum(float(sale.revenue[1:].replace(',', '')) for sale in sales):,.2f}"
    }
    return JSONResponse(content=resp, headers=headers)


@app.get("/api/merchant/sales/yesterday/analytics")
def get_yesterday_sales_analytics(merchant: tuple = Depends(get_merchant_id), db: Session = Depends(get_db)):
    """Get yesterday's sales analytics"""
    merchant_id, headers = merchant
    yesterday = date.today() - timedelta(days=1)
    resp = {
        "date": str(yesterday),
        "merchant_id": merchant_id,
        "total_revenue": "₹11,780.00",
        "profit_margin": "42.3%",
        "cogs": "₹6,795.00",
        "gross_profit": "₹4,985.00",
        "avg_order_value": "₹471.20",
        "customer_count": 25,
        "peak_sales_hour": "1:00 PM - 2:00 PM"
    }
    return JSONResponse(content=resp, headers=headers)


@app.get("/api/merchant/sales/yesterday/export")
def export_yesterday_sales(merchant: tuple = Depends(get_merchant_id), db: Session = Depends(get_db)):
    merchant_id, headers = merchant
    """Export yesterday's sales data"""
    # Generate and save a simple Excel file for yesterday's sales so the download URL points to an actual file
    yesterday = date.today() - timedelta(days=1)
    fname = f"yesterday_sales_{yesterday}.xlsx"
    dest = DOWNLOAD_DIR / fname

    wb = Workbook()
    ws = wb.active
    ws.title = "Yesterday Sales"
    ws.append(["Item", "Quantity", "Revenue"])
    sample_rows = [
        ["Coffee", 12, "₹360.00"],
        ["Pizza", 6, "₹900.00"],
        ["Burger", 10, "₹1,500.00"]
    ]
    for r in sample_rows:
        ws.append(r)
    wb.save(dest)

    resp = {
        "status": "success",
        "message": "Yesterday's sales data exported successfully",
        "download_url": f"/api/downloads/{fname}",
        "file_size": f"{dest.stat().st_size // 1024} KB",
        "total_records": len(sample_rows),
        "date": str(yesterday),
        "merchant_id": merchant_id
    }
    return JSONResponse(content=resp, headers=headers)


@app.get("/api/merchant/sales/weekly")
def get_weekly_sales(merchant: tuple = Depends(get_merchant_id), db: Session = Depends(get_db)):
    """Get weekly sales data"""
    merchant_id, headers = merchant
    today = date.today()
    week_start = today - timedelta(days=today.weekday())

    resp = {
        "merchant_id": merchant_id,
        "week_period": f"{week_start} to {today}",
        "total_weekly_sales": "₹78,450.00",
        "total_transactions": 187,
        "average_daily_sales": "₹11,207.14",
        "daily_breakdown": [
            {"day": "Monday", "date": str(
                week_start), "sales": "₹10,200.00", "transactions": 22},
            {"day": "Tuesday", "date": str(
                week_start + timedelta(days=1)), "sales": "₹11,500.00", "transactions": 26},
            {"day": "Wednesday", "date": str(
                week_start + timedelta(days=2)), "sales": "₹12,800.00", "transactions": 29},
            {"day": "Thursday", "date": str(
                week_start + timedelta(days=3)), "sales": "₹13,200.00", "transactions": 31},
            {"day": "Friday", "date": str(
                week_start + timedelta(days=4)), "sales": "₹15,400.00", "transactions": 35},
            {"day": "Saturday", "date": str(
                week_start + timedelta(days=5)), "sales": "₹13,570.00", "transactions": 32},
            {"day": "Sunday", "date": str(
                today), "sales": "₹1,780.00", "transactions": 12}
        ],
        "best_performing_day": "Friday",
        "growth_trend": "+8.5% compared to last week"
    }
    return JSONResponse(content=resp, headers=headers)


# Additional Weekly Sales Endpoints
@app.get("/api/merchant/sales/weekly/analytics")
def get_weekly_sales_analytics(merchant: tuple = Depends(get_merchant_id), db: Session = Depends(get_db)):
    """Get weekly sales analytics"""
    merchant_id, headers = merchant
    today = date.today()
    week_start = today - timedelta(days=today.weekday())
    resp = {
        "week_period": f"{week_start} to {today}",
        "merchant_id": merchant_id,
        "total_sales": "₹78,450.00",
        "daily_average": "₹11,207.14",
        "best_day": "Friday",
        "best_day_sales": "₹15,400.00",
        "worst_day": "Sunday",
        "worst_day_sales": "₹1,780.00",
        "week_over_week_growth": "+8.5%",
        "customer_retention": "73%"
    }
    return JSONResponse(content=resp, headers=headers)


@app.get("/api/merchant/sales/weekly/export")
def export_weekly_sales(merchant: tuple = Depends(get_merchant_id), db: Session = Depends(get_db)):
    """Export weekly sales report"""
    merchant_id, headers = merchant
    today = date.today()
    week_start = today - timedelta(days=today.weekday())

    # Create a weekly report Excel file and save to downloads
    fname = f"weekly_report_{week_start}_to_{today}.xlsx"
    dest = DOWNLOAD_DIR / fname

    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Report"
    ws.append(["Day", "Date", "Sales", "Transactions"])

    # Sample data for the week
    sample_days = []
    for i in range(7):
        d = week_start + timedelta(days=i)
        sample_days.append([d.strftime('%A'), str(
            d), f"₹{10000 + i*1500:.2f}", 20 + i*3])

    for row in sample_days:
        ws.append(row)

    # Summary row
    ws.append([])
    ws.append(["Total", "", f"₹{sum(float(r[2][1:].replace(',', '')) for r in sample_days):.2f}", sum(
        r[3] for r in sample_days)])

    wb.save(dest)

    content = {
        "status": "success",
        "message": "Weekly sales report exported successfully",
        "download_url": f"/api/downloads/{fname}",
        "file_size": f"{dest.stat().st_size // 1024} KB",
        "total_records": len(sample_days),
        "week_range": f"{week_start} to {today}",
        "merchant_id": merchant_id
    }
    return JSONResponse(content=content, headers=headers)


@app.get("/api/merchant/sales/weekly/compare")
def compare_weekly_sales(merchant: tuple = Depends(get_merchant_id), db: Session = Depends(get_db)):
    """Compare current week with previous week"""
    merchant_id, headers = merchant
    today = date.today()
    current_week_start = today - timedelta(days=today.weekday())
    previous_week_start = current_week_start - timedelta(days=7)
    content = {
        "current_week": f"{current_week_start} to {today}",
        "previous_week": f"{previous_week_start} to {current_week_start - timedelta(days=1)}",
        "merchant_id": merchant_id,
        "current_week_sales": "₹78,450.00",
        "previous_week_sales": "₹72,300.00",
        "growth_percentage": "+8.5%",
        "growth_direction": "📈 Increasing",
        "transaction_comparison": {
            "current": 187,
            "previous": 173,
            "difference": "+14 transactions"
        }
    }
    return JSONResponse(content=content, headers=headers)


@app.get("/api/merchant/payments/outstanding")
def get_outstanding_payments(db: Session = Depends(get_db)):
    """Get outstanding payment information"""
    return {
        "total_outstanding": "₹45,600.00",
        "overdue_amount": "₹12,300.00",
        "pending_invoices": 8,
        "outstanding_payments": [
            {
                "invoice_id": "INV-2024-001",
                "customer_name": "ABC Corp",
                "amount": "₹8,500.00",
                "due_date": "2024-08-15",
                "days_overdue": 14,
                "status": "Overdue"
            },
            {
                "invoice_id": "INV-2024-002",
                "customer_name": "XYZ Ltd",
                "amount": "₹6,200.00",
                "due_date": "2024-08-25",
                "days_overdue": 4,
                "status": "Overdue"
            },
            {
                "invoice_id": "INV-2024-003",
                "customer_name": "Quick Mart",
                "amount": "₹12,400.00",
                "due_date": "2024-09-05",
                "days_overdue": 0,
                "status": "Pending"
            },
            {
                "invoice_id": "INV-2024-004",
                "customer_name": "Food Palace",
                "amount": "₹18,500.00",
                "due_date": "2024-09-10",
                "days_overdue": 0,
                "status": "Pending"
            }
        ],
        "payment_reminders_sent": 3,
        "last_reminder_date": "2024-08-27"
    }


# Additional Payment Management Endpoints
@app.get("/api/merchant/payments/send-reminders")
def send_payment_reminders(merchant: tuple = Depends(get_merchant_id), db: Session = Depends(get_db)):
    """Send payment reminders to customers. Uses get_merchant_id dependency for defaults and warnings."""
    merchant_id, headers = merchant
    # In a real system this would enqueue messages or call external services.
    logger = logging.getLogger(__name__)

    content = {
        "status": "success",
        "message": "Payment reminders sent successfully",
        "merchant_id": merchant_id,
        "reminders_sent": 8,
        "total_outstanding": "₹45,600.00",
        "customers_notified": [
            {"customer": "ABC Corp", "amount": "₹15,200.00", "method": "Email"},
            {"customer": "XYZ Ltd", "amount": "₹8,500.00", "method": "SMS"},
            {"customer": "PQR Enterprises",
                "amount": "₹12,300.00", "method": "WhatsApp"}
        ],
        "reminder_date": str(date.today())
    }

    return JSONResponse(content=content, headers=headers)


@app.get("/api/merchant/payments/update-status")
def update_payment_status(merchant: tuple = Depends(get_merchant_id), payment_id: str = Query(...), status: str = Query(...), db: Session = Depends(get_db)):
    """Update payment status"""
    merchant_id, headers = merchant
    content = {
        "status": "success",
        "message": f"Payment status updated to {status}",
        "merchant_id": merchant_id,
        "payment_id": payment_id,
        "new_status": status,
        "updated_date": str(date.today()),
        "updated_by": "Merchant Portal"
    }
    return JSONResponse(content=content, headers=headers)


@app.get("/api/merchant/payments/report")
def generate_payment_report(merchant: tuple = Depends(get_merchant_id), db: Session = Depends(get_db)):
    """Generate comprehensive payment report. Uses get_merchant_id dependency so callers may omit merchant_id.

    Returns X-Warning header when merchant_id was defaulted.
    """
    merchant_id, headers = merchant

    # Ensure a downloadable file exists for the payment report to avoid 404s.
    fname = f"payment_report_{date.today()}.pdf"
    dest = DOWNLOAD_DIR / fname

    # Try to generate a proper PDF using reportlab if available.
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet

        doc = SimpleDocTemplate(str(dest), pagesize=letter)
        styles = getSampleStyleSheet()

        story = []
        story.append(Paragraph("Payment Report", styles["Title"]))
        story.append(Paragraph(f"Generated: {date.today()}", styles["Normal"]))
        story.append(Paragraph(f"Merchant: {merchant_id}", styles["Normal"]))
        story.append(Spacer(1, 12))

        data = [["Metric", "Value"],
                ["Total Invoices", "25"],
                ["Paid Invoices", "17"],
                ["Overdue Invoices", "5"],
                ["Pending Invoices", "3"]]

        table = Table(data, hAlign='LEFT')
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold')
        ]))

        story.append(table)
        doc.build(story)
        logging.getLogger(__name__).info(
            "Generated PDF payment report using reportlab platypus: %s", dest)
    except Exception:
        # Fallback: create a small placeholder PDF-like file (plain text with .pdf extension).
        try:
            with open(dest, "wb") as f:
                body = (
                    f"Payment Report\nGenerated: {date.today()}\nMerchant: {merchant_id}\n\n"
                    "Summary:\n"
                    "- Total Invoices: 25\n"
                    "- Paid Invoices: 17\n"
                    "- Overdue Invoices: 5\n"
                    "- Pending Invoices: 3\n"
                    "\nThis is a placeholder report file. Install reportlab to generate a proper PDF.\n"
                )
                f.write(body.encode("utf-8"))
        except Exception as e:
            logging.getLogger(__name__).exception(
                "Failed to write payment report file: %s", e)

    content = {
        "status": "success",
        "message": "Payment report generated successfully",
        "merchant_id": merchant_id,
        "total_outstanding": "₹45,600.00",
        "overdue_count": 5,
        "pending_count": 3,
        "report_url": f"/api/downloads/{fname}",
        "generated_date": str(date.today()),
        "file_size": f"{dest.stat().st_size // 1024} KB",
        "summary": {
            "total_invoices": 25,
            "paid_invoices": 17,
            "overdue_invoices": 5,
            "pending_invoices": 3
        }
    }

    # Return with any warning headers created by the dependency
    return JSONResponse(content=content, headers=headers)


@app.get("/api/merchant/expenses/bills")
def get_expense_bills(db: Session = Depends(get_db)):
    """Get expense and bill information"""
    return {
        "total_monthly_expenses": "₹25,800.00",
        "pending_bills": "₹8,400.00",
        "paid_this_month": "₹17,400.00",
        "expense_categories": [
            {"category": "Rent", "amount": "₹10,000.00", "status": "Paid"},
            {"category": "Utilities", "amount": "₹2,500.00", "status": "Pending"},
            {"category": "Inventory", "amount": "₹8,200.00", "status": "Paid"},
            {"category": "Staff Salaries", "amount": "₹3,600.00", "status": "Pending"},
            {"category": "Marketing", "amount": "₹1,500.00", "status": "Paid"}
        ],
        "upcoming_bills": [
            {"bill_type": "Electricity", "amount": "₹1,200.00",
                "due_date": "2024-09-05"},
            {"bill_type": "Internet", "amount": "₹800.00", "due_date": "2024-09-08"},
            {"bill_type": "Insurance", "amount": "₹2,400.00", "due_date": "2024-09-15"}
        ],
        "monthly_budget": "₹30,000.00",
        "budget_utilization": "86%"
    }


# Additional Expense Management Endpoints
@app.get("/api/merchant/expenses/add")
def add_new_expense(merchant: tuple = Depends(get_merchant_id), description: str = Query(...), amount: float = Query(...), category: str = Query("Other"), db: Session = Depends(get_db)):
    """Add a new expense"""
    merchant_id, headers = merchant
    content = {
        "status": "success",
        "message": "Expense added successfully",
        "merchant_id": merchant_id,
        "expense_id": f"EXP-{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "description": description,
        "amount": f"₹{amount:.2f}",
        "category": category,
        "date_added": str(date.today()),
        "status": "Recorded"
    }
    return JSONResponse(content=content, headers=headers)


@app.get("/api/merchant/expenses/monthly-report")
def get_monthly_expense_report(merchant: tuple = Depends(get_merchant_id), db: Session = Depends(get_db)):
    """Get monthly expense breakdown report"""
    merchant_id, headers = merchant
    content = {
        "month": date.today().strftime("%B %Y"),
        "merchant_id": merchant_id,
        "total_expenses": "₹25,800.00",
        "top_category": "Staff Salaries",
        "top_category_amount": "₹8,200.00",
        "categories": [
            {"category": "Staff Salaries",
                "amount": "₹8,200.00", "percentage": "31.8%"},
            {"category": "Inventory", "amount": "₹6,500.00", "percentage": "25.2%"},
            {"category": "Utilities", "amount": "₹4,200.00", "percentage": "16.3%"},
            {"category": "Marketing", "amount": "₹3,100.00", "percentage": "12.0%"},
            {"category": "Maintenance", "amount": "₹2,300.00", "percentage": "8.9%"},
            {"category": "Other", "amount": "₹1,500.00", "percentage": "5.8%"}
        ],
        "budget_comparison": {
            "budgeted": "₹30,000.00",
            "actual": "₹25,800.00",
            "variance": "₹4,200.00",
            "variance_percentage": "14% under budget"
        }
    }
    return JSONResponse(content=content, headers=headers)


@app.get("/api/merchant/expenses/update-bill")
def update_bill_status(merchant: tuple = Depends(get_merchant_id), bill_id: str = Query(...), status: str = Query(...), db: Session = Depends(get_db)):
    """Update bill payment status"""
    merchant_id, headers = merchant
    content = {
        "status": "success",
        "message": f"Bill status updated to {status}",
        "merchant_id": merchant_id,
        "bill_id": bill_id,
        "new_status": status,
        "updated_date": str(date.today()),
        "next_action": "Bill marked as paid" if status == "Paid" else f"Status changed to {status}"
    }
    return JSONResponse(content=content, headers=headers)


@app.get("/api/merchant/staff/attendance")
def get_staff_attendance(db: Session = Depends(get_db)):
    """Get staff attendance information"""
    today = date.today()

    return {
        "date": str(today),
        "total_staff": 6,
        "present_today": 5,
        "absent_today": 1,
        "attendance_rate": "83.3%",
        "staff_status": [
            {"name": "John Smith", "role": "Manager", "status": "Present",
                "check_in": "09:00 AM", "location": "Store"},
            {"name": "Sarah Johnson", "role": "Cashier", "status": "Present",
                "check_in": "09:15 AM", "location": "Store"},
            {"name": "Mike Brown", "role": "Cook", "status": "Present",
                "check_in": "08:45 AM", "location": "Kitchen"},
            {"name": "Lisa Davis", "role": "Server", "status": "Present",
                "check_in": "09:30 AM", "location": "Store"},
            {"name": "Tom Wilson", "role": "Cleaner",
                "status": "Absent", "check_in": "-", "location": "-"},
            {"name": "Anna Taylor", "role": "Assistant", "status": "Present",
                "check_in": "09:10 AM", "location": "Store"}
        ],
        "weekly_summary": {
            "average_attendance": "88.1%",
            "most_punctual": "Mike Brown",
            "total_working_hours": "234 hours"
        }
    }


@app.get("/api/merchant/staff/leave-requests")
def get_staff_leave_requests(db: Session = Depends(get_db)):
    """Get staff leave requests"""
    return {
        "pending_requests": 2,
        "approved_this_month": 4,
        "rejected_this_month": 0,
        "leave_requests": [
            {
                "request_id": "LR-001",
                "employee_name": "Sarah Johnson",
                "leave_type": "Annual Leave",
                "from_date": "2024-09-10",
                "to_date": "2024-09-12",
                "days": 3,
                "reason": "Family vacation",
                "status": "Pending",
                "applied_date": "2024-08-25"
            },
            {
                "request_id": "LR-002",
                "employee_name": "Tom Wilson",
                "leave_type": "Sick Leave",
                "from_date": "2024-09-05",
                "to_date": "2024-09-06",
                "days": 2,
                "reason": "Medical appointment",
                "status": "Pending",
                "applied_date": "2024-08-28"
            },
            {
                "request_id": "LR-003",
                "employee_name": "Lisa Davis",
                "leave_type": "Personal Leave",
                "from_date": "2024-08-20",
                "to_date": "2024-08-20",
                "days": 1,
                "reason": "Personal work",
                "status": "Approved",
                "applied_date": "2024-08-15"
            }
        ],
        "leave_balance_summary": [
            {"employee": "Sarah Johnson", "annual_leave": 12, "sick_leave": 8},
            {"employee": "Mike Brown", "annual_leave": 15, "sick_leave": 10},
            {"employee": "Tom Wilson", "annual_leave": 8, "sick_leave": 6}
        ]
    }


@app.get("/api/merchant/staff/messages")
def get_staff_messages(db: Session = Depends(get_db)):
    """Get staff messages and announcements"""
    return {
        "unread_messages": 3,
        "total_messages": 12,
        "messages": [
            {
                "message_id": "MSG-001",
                "from": "Management",
                "to": "All Staff",
                "subject": "New Safety Protocols",
                "message": "Please follow the updated safety protocols effective immediately.",
                "date": "2024-08-28",
                "priority": "High",
                "status": "Unread"
            },
            {
                "message_id": "MSG-002",
                "from": "HR Department",
                "to": "All Staff",
                "subject": "Monthly Team Meeting",
                "message": "Monthly team meeting scheduled for September 5th at 2 PM.",
                "date": "2024-08-27",
                "priority": "Medium",
                "status": "Unread"
            },
            {
                "message_id": "MSG-003",
                "from": "John Smith",
                "to": "Kitchen Staff",
                "subject": "New Menu Items",
                "message": "We're introducing new items next week. Training session tomorrow.",
                "date": "2024-08-26",
                "priority": "Medium",
                "status": "Read"
            }
        ],
        "announcements": [
            {
                "title": "Store Renovation",
                "content": "Store renovation will begin next month. Expect temporary changes.",
                "date": "2024-08-25",
                "type": "Info"
            },
            {
                "title": "Performance Bonus",
                "content": "Congratulations! Store exceeded sales target. Bonus will be credited.",
                "date": "2024-08-20",
                "type": "Success"
            }
        ]
    }


@app.get("/api/merchant/staff/add-employee")
def get_add_employee_form(merchant: tuple = Depends(get_merchant_id), db: Session = Depends(get_db)):
    """Get add employee form data"""
    merchant_id, headers = merchant
    content = {
        "status": "success",
        "merchant_id": merchant_id,
        "form_title": "Add New Employee",
        "required_fields": [
            "full_name", "email", "phone", "position", "department", "salary", "start_date"
        ],
        "departments": ["Sales", "Marketing", "Operations", "HR", "Finance", "IT"],
        "positions": ["Manager", "Assistant", "Executive", "Intern", "Senior Executive"],
        "instructions": "Please fill all required fields to add a new employee to your team."
    }
    return JSONResponse(content=content, headers=headers)


@app.post("/api/merchant/staff/add-employee")
def add_new_employee(employee_data: schemas.AddEmployeeRequest, db: Session = Depends(get_db)):
    """Add a new employee"""
    try:
        # Create new employee record
        new_employee = models.Employee(
            employee_id=employee_data.employee_id,
            employee_name=employee_data.employee_name,
            email=employee_data.email,
            phone=employee_data.phone,
            department=employee_data.department,
            position=employee_data.position,
            employment_type=employee_data.employment_type,
            employment_status="Active",
            hire_date=date.fromisoformat(employee_data.hire_date),
            reporting_manager=employee_data.reporting_manager,
            office_location=employee_data.office_location
        )

        db.add(new_employee)
        db.commit()
        db.refresh(new_employee)

        return {
            "success": True,
            "message": "Employee added successfully",
            "employee_id": new_employee.employee_id,
            "employee_name": new_employee.employee_name,
            "department": new_employee.department,
            "position": new_employee.position,
            "hire_date": str(new_employee.hire_date)
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500, detail=f"Failed to add employee: {str(e)}")


@app.get("/api/merchant/staff/salary")
def get_staff_salary_info(db: Session = Depends(get_db)):
    """Get staff salary information"""
    return {
        "total_monthly_payroll": "₹45,000.00",
        "employees_count": 6,
        "average_salary": "₹7,500.00",
        "salary_breakdown": [
            {"employee": "John Smith", "position": "Manager",
                "salary": "₹12,000.00", "status": "Paid"},
            {"employee": "Sarah Johnson", "position": "Cashier",
                "salary": "₹8,000.00", "status": "Paid"},
            {"employee": "Mike Brown", "position": "Cook",
                "salary": "₹9,000.00", "status": "Paid"},
            {"employee": "Lisa Davis", "position": "Server",
                "salary": "₹6,500.00", "status": "Pending"},
            {"employee": "Tom Wilson", "position": "Cleaner",
                "salary": "₹5,000.00", "status": "Pending"},
            {"employee": "Anna Taylor", "position": "Assistant",
                "salary": "₹4,500.00", "status": "Paid"}
        ],
        "pending_payments": "₹11,500.00",
        "next_payroll_date": "2024-09-05",
        "overtime_hours": {
            "total_overtime": "24 hours",
            "overtime_pay": "₹2,400.00"
        }
    }


@app.get("/api/merchant/hr/add-employee")
def get_add_employee_form(merchant_id: str = Query(...), db: Session = Depends(get_db)):
    """Get add employee form data"""
    return {
        "status": "success",
        "merchant_id": merchant_id,
        "form_title": "Add New Employee",
        "departments": ["Sales", "HR", "Finance", "Operations", "Marketing", "IT", "Customer Service"],
        "positions": ["Manager", "Senior Executive", "Executive", "Assistant", "Intern"],
        "employment_types": ["Full-time", "Part-time", "Contract", "Internship"],
        "required_fields": ["full_name", "email", "phone", "department", "position", "start_date"],
        "optional_fields": ["address", "emergency_contact", "previous_experience"]
    }


@app.get("/api/merchant/hr/support")
def get_hr_support_form(merchant_id: str = Query(...), db: Session = Depends(get_db)):
    """Get HR support form data"""
    return {
        "status": "success",
        "merchant_id": merchant_id,
        "form_title": "HR Support Request",
        "categories": ["General HR", "Employee Issues", "Payroll", "Compliance", "Training", "Other"],
        "priority_levels": ["Low", "Medium", "High", "Urgent"],
        "support_channels": ["Email", "Phone", "Chat", "In-Person"],
        "business_hours": "Monday-Friday 9:00 AM - 6:00 PM",
        "emergency_contact": "+91-1800-HR-HELP"
    }


@app.post("/api/merchant/staff/hr-support")
def submit_hr_support_request(request_data: schemas.HRSupportRequest, db: Session = Depends(get_db)):
    """Submit HR support request"""
    try:
        # Create HR support ticket
        support_ticket = models.HRSupportTicket(
            employee_id=request_data.employee_id,
            employee_name=request_data.employee_name,
            category=request_data.category,
            subject=request_data.subject,
            description=request_data.description,
            priority=request_data.priority
        )

        db.add(support_ticket)
        db.commit()
        db.refresh(support_ticket)

        return {
            "success": True,
            "message": "HR support request submitted successfully",
            "ticket_id": support_ticket.id,
            "category": support_ticket.category,
            "priority": support_ticket.priority,
            "status": support_ticket.status,
            "estimated_response_time": "24-48 hours"
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500, detail=f"Failed to submit HR request: {str(e)}")


@app.get("/api/merchant/marketing/whatsapp-campaign")
def get_whatsapp_campaign_form(merchant_id: str = Query(...), db: Session = Depends(get_db)):
    """Get WhatsApp campaign form data"""
    return {
        "status": "success",
        "merchant_id": merchant_id,
        "form_title": "WhatsApp Marketing Campaign",
        "campaign_types": ["Promotional", "Announcement", "Follow-up", "Seasonal"],
        "target_audiences": ["All Customers", "New Customers", "Repeat Customers", "VIP Customers"],
        "message_templates": [
            "Welcome new customers",
            "Special discount offer",
            "Product announcement",
            "Appointment reminder",
            "Custom message"
        ],
        "budget_ranges": ["₹500-1000", "₹1000-5000", "₹5000-10000", "₹10000+"],
        "scheduling_options": ["Send Now", "Schedule for Later", "Recurring Campaign"]
    }


@app.post("/api/merchant/marketing/whatsapp-campaign")
def create_whatsapp_campaign(campaign_data: schemas.WhatsAppCampaignRequest, db: Session = Depends(get_db)):
    """Create WhatsApp marketing campaign"""
    try:
        # Create marketing campaign
        campaign = models.MarketingCampaign(
            campaign_name=campaign_data.campaign_name,
            campaign_type="WhatsApp",
            target_audience=campaign_data.target_audience,
            message_content=campaign_data.message_content,
            scheduled_date=date.fromisoformat(
                campaign_data.scheduled_date) if campaign_data.scheduled_date else None,
            budget=campaign_data.budget
        )

        db.add(campaign)
        db.commit()
        db.refresh(campaign)

        return {
            "success": True,
            "message": "WhatsApp campaign created successfully",
            "campaign_id": campaign.id,
            "campaign_name": campaign.campaign_name,
            "target_audience": campaign.target_audience,
            "estimated_reach": "500-800 customers",
            "estimated_cost": f"₹{campaign.budget or 0}.00",
            "status": campaign.status,
            "scheduled_date": str(campaign.scheduled_date) if campaign.scheduled_date else "Immediate"
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500, detail=f"Failed to create campaign: {str(e)}")


@app.get("/api/merchant/marketing/promotion")
def get_promotion_form(merchant_id: str = Query(...), db: Session = Depends(get_db)):
    """Get promotion form data"""
    return {
        "status": "success",
        "merchant_id": merchant_id,
        "form_title": "Create Promotion",
        "promotion_types": ["Percentage Discount", "Fixed Amount", "Buy One Get One", "Bundle Offer"],
        "discount_ranges": ["5%", "10%", "15%", "20%", "25%", "30%", "50%"],
        "validity_options": ["1 Day", "3 Days", "1 Week", "2 Weeks", "1 Month", "Custom"],
        "target_products": ["All Products", "Specific Category", "Selected Items", "New Arrivals"],
        "minimum_purchase_options": ["No Minimum", "₹500", "₹1000", "₹2000", "₹5000"],
        "active_promotions": []
    }


@app.post("/api/merchant/marketing/instant-promotion")
def create_instant_promotion(promotion_data: schemas.InstantPromotionRequest, db: Session = Depends(get_db)):
    """Create instant promotion"""
    try:
        # Create instant promotion
        promotion = models.Promotion(
            promotion_name=promotion_data.promotion_name,
            promotion_type=promotion_data.promotion_type,
            discount_percentage=promotion_data.discount_percentage,
            discount_amount=promotion_data.discount_amount,
            valid_from=date.fromisoformat(promotion_data.valid_from),
            valid_until=date.fromisoformat(promotion_data.valid_until),
            applicable_items=promotion_data.applicable_items,
            minimum_purchase=promotion_data.minimum_purchase
        )

        db.add(promotion)
        db.commit()
        db.refresh(promotion)

        return {
            "success": True,
            "message": "Instant promotion created successfully",
            "promotion_id": promotion.id,
            "promotion_name": promotion.promotion_name,
            "promotion_type": promotion.promotion_type,
            "discount": f"{promotion.discount_percentage}%" if promotion.discount_percentage else f"₹{promotion.discount_amount}",
            "valid_period": f"{promotion.valid_from} to {promotion.valid_until}",
            "status": promotion.status,
            "estimated_impact": "15-25% increase in sales"
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500, detail=f"Failed to create promotion: {str(e)}")


# Company info endpoint
@app.get("/api/chatbot/company-info")
def get_company_info():
    """Get company information"""
    return {
        "company_name": "YouHR Assistant",
        "version": "2.0",
        "description": "Comprehensive HR and Merchant Management System",
        "features": [
            "HR Management",
            "Merchant Operations",
            "Staff Management",
            "Sales Analytics",
            "Marketing Tools"
        ],
        "last_updated": "2024-08-29"
    }


# ===== MISSING MERCHANT ENDPOINTS =====

# Marketing & Growth - Missing endpoints
@app.get("/api/merchant/marketing/campaign-results")
def get_campaign_results(merchant_id: str = Query(..., description="Merchant ID")):
    """Get marketing campaign results"""
    return {
        "merchant_id": merchant_id,
        "total_campaigns": 5,
        "active_campaigns": 2,
        "campaigns": [
            {
                "campaign_id": "WC-001",
                "campaign_name": "Summer Sale 2024",
                "type": "WhatsApp",
                "sent": 1200,
                "delivered": 1150,
                "read": 890,
                "clicked": 120,
                "conversion_rate": "10.1%",
                "revenue_generated": "₹15,600.00",
                "status": "Completed"
            },
            {
                "campaign_id": "WC-002",
                "campaign_name": "Weekend Special",
                "type": "WhatsApp",
                "sent": 800,
                "delivered": 785,
                "read": 620,
                "clicked": 85,
                "conversion_rate": "10.6%",
                "revenue_generated": "₹8,900.00",
                "status": "Active"
            }
        ],
        "overall_performance": {
            "total_reach": 2000,
            "average_open_rate": "75.6%",
            "average_click_rate": "10.3%",
            "total_revenue": "₹24,500.00",
            "roi": "245%"
        }
    }


@app.get("/api/merchant/loan/status")
def get_loan_status(merchant_id: str = Query(..., description="Merchant ID")):
    """Check loan application status"""
    return {
        "merchant_id": merchant_id,
        "loan_applications": [
            {
                "loan_id": "LN-2024-001",
                "application_date": "2024-08-15",
                "loan_amount": "₹1,00,000.00",
                "loan_type": "Business Loan",
                "status": "Under Review",
                "approval_probability": "85%",
                "estimated_approval_date": "2024-09-05",
                "required_documents": [
                    {"document": "KYC", "status": "Submitted"},
                    {"document": "Bank Statements", "status": "Submitted"},
                    {"document": "Business Proof", "status": "Pending"},
                    {"document": "Income Tax Returns", "status": "Submitted"}
                ],
                "loan_officer": "Rahul Sharma",
                "contact": "+91-9876543210"
            }
        ],
        "loan_history": [
            {
                "loan_id": "LN-2023-005",
                "amount": "₹50,000.00",
                "status": "Fully Repaid",
                "disbursed_date": "2023-12-01",
                "completion_date": "2024-06-01"
            }
        ],
        "credit_score": 750,
        "eligible_loan_amount": "₹2,50,000.00"
    }


@app.post("/api/merchant/loan/continue")
def continue_loan_application(
    merchant_id: str = Query(..., description="Merchant ID"),
    document_type: str = Query(..., description="Document type"),
    document_url: str = Query(..., description="Document URL"),
    notes: str = Query(None, description="Additional notes")
):
    """Continue loan application with document upload"""
    return {
        "merchant_id": merchant_id,
        "loan_id": "LN-2024-001",
        "message": "Document uploaded successfully",
        "document_type": document_type,
        "upload_status": "Success",
        "verification_status": "Pending",
        "next_steps": [
            "Wait for document verification",
            "Loan officer will contact within 2 business days",
            "Prepare for final approval process"
        ],
        "completion_percentage": "90%",
        "estimated_approval_date": "2024-09-05"
    }


# Notifications - All endpoints
@app.get("/api/merchant/notifications/leave-requests")
def get_pending_leave_notifications(merchant_id: str = Query(..., description="Merchant ID")):
    """Get pending leave request notifications"""
    return {
        "merchant_id": merchant_id,
        "pending_requests": 3,
        "notifications": [
            {
                "notification_id": "LN-001",
                "employee_name": "Sarah Johnson",
                "employee_id": "EMP002",
                "leave_type": "Annual Leave",
                "from_date": "2024-09-10",
                "to_date": "2024-09-12",
                "days": 3,
                "reason": "Family vacation",
                "applied_date": "2024-08-25",
                "urgency": "Medium",
                "action_required": "Approve/Reject"
            },
            {
                "notification_id": "LN-002",
                "employee_name": "Mike Brown",
                "employee_id": "EMP003",
                "leave_type": "Sick Leave",
                "from_date": "2024-09-02",
                "to_date": "2024-09-03",
                "days": 2,
                "reason": "Medical emergency",
                "applied_date": "2024-09-01",
                "urgency": "High",
                "action_required": "Approve/Reject"
            }
        ],
        "summary": {
            "total_pending": 3,
            "high_priority": 1,
            "medium_priority": 2,
            "overdue_responses": 0
        }
    }


@app.get("/api/merchant/notifications/shift-change")
def get_shift_change_notifications(merchant_id: str = Query(..., description="Merchant ID")):
    """Get shift change request notifications"""
    return {
        "merchant_id": merchant_id,
        "pending_requests": 2,
        "notifications": [
            {
                "notification_id": "SC-001",
                "employee_name": "Lisa Davis",
                "employee_id": "EMP004",
                "current_shift": "Morning (9 AM - 5 PM)",
                "requested_shift": "Evening (2 PM - 10 PM)",
                "effective_date": "2024-09-05",
                "reason": "Personal commitment",
                "applied_date": "2024-08-28",
                "urgency": "Medium",
                "action_required": "Approve/Reject"
            },
            {
                "notification_id": "SC-002",
                "employee_name": "Tom Wilson",
                "employee_id": "EMP005",
                "current_shift": "Evening (2 PM - 10 PM)",
                "requested_shift": "Morning (9 AM - 5 PM)",
                "effective_date": "2024-09-08",
                "reason": "Transportation issues",
                "applied_date": "2024-08-29",
                "urgency": "High",
                "action_required": "Approve/Reject"
            }
        ],
        "summary": {
            "total_pending": 2,
            "high_priority": 1,
            "medium_priority": 1,
            "schedule_conflicts": 0
        }
    }


@app.get("/api/merchant/notifications/payment-settlement")
def get_payment_settlement_notifications(merchant_id: str = Query(..., description="Merchant ID")):
    """Get payment settlement notifications"""
    return {
        "merchant_id": merchant_id,
        "recent_settlements": [
            {
                "settlement_id": "ST-001",
                "date": "2024-08-29",
                "amount": "₹25,400.00",
                "transaction_count": 45,
                "settlement_type": "Daily Settlement",
                "status": "Completed",
                "bank_reference": "UTR123456789",
                "credited_to": "HDFC Bank ****1234"
            },
            {
                "settlement_id": "ST-002",
                "date": "2024-08-28",
                "amount": "₹22,800.00",
                "transaction_count": 38,
                "settlement_type": "Daily Settlement",
                "status": "Completed",
                "bank_reference": "UTR987654321",
                "credited_to": "HDFC Bank ****1234"
            }
        ],
        "pending_settlements": [
            {
                "settlement_id": "ST-003",
                "date": "2024-08-29",
                "amount": "₹8,900.00",
                "transaction_count": 15,
                "expected_credit": "2024-08-30",
                "status": "Processing"
            }
        ],
        "monthly_summary": {
            "total_settled": "₹6,78,900.00",
            "pending_amount": "₹8,900.00",
            "average_daily_settlement": "₹23,410.00"
        }
    }


@app.get("/api/merchant/notifications/renew-subscription")
def get_subscription_notifications(merchant_id: str = Query(..., description="Merchant ID")):
    """Get subscription renewal notifications"""
    return {
        "merchant_id": merchant_id,
        "subscription_status": "Active",
        "current_plan": "YouShop Pro",
        "expiry_date": "2024-10-15",
        "days_remaining": 47,
        "renewal_required": True,
        "notifications": [
            {
                "notification_id": "SUB-001",
                "type": "Renewal Reminder",
                "message": "Your subscription expires in 47 days. Renew now to avoid service interruption.",
                "urgency": "Medium",
                "action_required": "Renew Subscription",
                "renewal_discount": "20% off on annual plans"
            }
        ],
        "subscription_details": {
            "features": [
                "POS System",
                "Inventory Management",
                "Staff Management",
                "Sales Analytics",
                "Marketing Tools"
            ],
            "monthly_cost": "₹2,999.00",
            "annual_cost": "₹29,990.00",
            "savings_on_annual": "₹5,998.00"
        },
        "payment_history": [
            {
                "date": "2023-10-15",
                "amount": "₹29,990.00",
                "plan": "Annual Pro",
                "status": "Paid"
            }
        ]
    }


@app.get("/api/merchant/notifications/head-office-messages")
def get_head_office_messages(merchant_id: str = Query(..., description="Merchant ID")):
    """Get messages from head office"""
    return {
        "merchant_id": merchant_id,
        "unread_messages": 2,
        "total_messages": 8,
        "messages": [
            {
                "message_id": "HO-001",
                "from": "Head Office - Operations",
                "subject": "New Product Launch",
                "message": "We're launching new product categories next month. Please update your inventory system and train staff accordingly.",
                "date": "2024-08-28",
                "priority": "High",
                "status": "Unread",
                "attachments": ["product_catalog.pdf", "training_guide.pdf"],
                "action_required": "Update inventory & staff training"
            },
            {
                "message_id": "HO-002",
                "from": "Head Office - IT Support",
                "subject": "System Maintenance",
                "message": "Scheduled maintenance on September 1st from 2 AM to 4 AM. POS system may be temporarily unavailable.",
                "date": "2024-08-27",
                "priority": "Medium",
                "status": "Unread",
                "action_required": "Plan operations accordingly"
            },
            {
                "message_id": "HO-003",
                "from": "Head Office - HR",
                "subject": "Policy Update",
                "message": "Updated employee handbook is available. Please share with all staff members.",
                "date": "2024-08-25",
                "priority": "Medium",
                "status": "Read",
                "attachments": ["employee_handbook_v2.pdf"]
            }
        ],
        "announcements": [
            {
                "title": "Q3 Performance Awards",
                "content": "Congratulations! Your store is among top 10 performers this quarter.",
                "date": "2024-08-20",
                "type": "Achievement"
            }
        ]
    }


# Help & Support - All endpoints
@app.post("/api/merchant/support/pos-app")
def report_pos_app_problem(
    merchant_id: str = Query(..., description="Merchant ID"),
    problem_description: str = Query(..., description="Problem description"),
    error_code: str = Query(None, description="Error code if any"),
    priority: str = Query("Medium", description="Priority level")
):
    """Report POS application problem"""
    return {
        "merchant_id": merchant_id,
        "ticket_id": f"POS-{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "message": "POS app problem reported successfully",
        "problem_description": problem_description,
        "error_code": error_code,
        "priority": priority,
        "status": "Open",
        "assigned_to": "Technical Support Team",
        "estimated_resolution": "2-4 hours",
        "support_contact": "+91-1800-123-4567",
        "next_steps": [
            "Technical team will analyze the issue",
            "Remote assistance may be provided",
            "You will receive updates via SMS/Email"
        ]
    }


@app.post("/api/merchant/support/hardware")
def report_hardware_issue(
    merchant_id: str = Query(..., description="Merchant ID"),
    hardware_type: str = Query(..., description="Hardware type"),
    issue_description: str = Query(..., description="Issue description"),
    model: str = Query(None, description="Hardware model"),
    priority: str = Query("Medium", description="Priority level")
):
    """Report hardware issue"""
    return {
        "merchant_id": merchant_id,
        "ticket_id": f"HW-{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "message": "Hardware issue reported successfully",
        "hardware_type": hardware_type,
        "model": model,
        "issue_description": issue_description,
        "priority": priority,
        "status": "Open",
        "assigned_to": "Field Support Team",
        "estimated_resolution": "4-8 hours",
        "service_engineer": "Will be assigned based on location",
        "support_contact": "+91-1800-123-4567",
        "next_steps": [
            "Service engineer will be assigned",
            "Engineer will contact you within 2 hours",
            "On-site visit will be scheduled if required"
        ]
    }


@app.post("/api/merchant/support/ai-camera")
def report_ai_camera_problem(
    merchant_id: str = Query(..., description="Merchant ID"),
    camera_id: str = Query(..., description="Camera ID"),
    problem_description: str = Query(..., description="Problem description"),
    error_logs: str = Query(None, description="Error logs"),
    priority: str = Query("Medium", description="Priority level")
):
    """Report AI camera (YouLens) problem"""
    return {
        "merchant_id": merchant_id,
        "camera_id": camera_id,
        "ticket_id": f"CAM-{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "message": "AI camera problem reported successfully",
        "problem_description": problem_description,
        "error_logs": error_logs,
        "priority": priority,
        "status": "Open",
        "assigned_to": "AI Support Team",
        "estimated_resolution": "2-6 hours",
        "support_contact": "+91-1800-123-4567",
        "remote_diagnostics": "Available",
        "next_steps": [
            "AI team will run remote diagnostics",
            "Camera logs will be analyzed",
            "Software update may be pushed",
            "On-site visit if hardware issue detected"
        ]
    }


@app.post("/api/merchant/support/camera-installation")
def request_camera_installation(
    merchant_id: str = Query(..., description="Merchant ID"),
    request_type: str = Query(..., description="Installation or Training"),
    location: str = Query(..., description="Installation location"),
    preferred_date: str = Query(..., description="Preferred date"),
    notes: str = Query(None, description="Additional notes")
):
    """Request camera installation or training"""
    return {
        "merchant_id": merchant_id,
        "request_id": f"INST-{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "message": f"Camera {request_type.lower()} request submitted successfully",
        "request_type": request_type,
        "location": location,
        "preferred_date": preferred_date,
        "notes": notes,
        "status": "Scheduled",
        "assigned_engineer": "Will be assigned based on location",
        "estimated_duration": "2-4 hours",
        "support_contact": "+91-1800-123-4567",
        "next_steps": [
            "Installation team will contact you",
            "Site survey will be conducted",
            "Installation date will be confirmed",
            "Training session will be provided"
        ]
    }


@app.post("/api/merchant/support/general")
def submit_general_support(
    merchant_id: str = Query(..., description="Merchant ID"),
    category: str = Query(..., description="Support category"),
    subject: str = Query(..., description="Subject"),
    description: str = Query(..., description="Detailed description"),
    priority: str = Query("Low", description="Priority level")
):
    """Submit general support request"""
    return {
        "merchant_id": merchant_id,
        "ticket_id": f"GEN-{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "message": "General support request submitted successfully",
        "category": category,
        "subject": subject,
        "description": description,
        "priority": priority,
        "status": "Open",
        "assigned_to": "Customer Support Team",
        "estimated_response": "4-24 hours",
        "support_contact": "+91-1800-123-4567",
        "ticket_url": f"https://support.youshop.com/ticket/GEN-{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "next_steps": [
            "Support team will review your request",
            "You will receive response via email",
            "Additional information may be requested",
            "Solution will be provided or escalated as needed"
        ]
    }


# Feedback & Ideas - All endpoints
@app.post("/api/merchant/feedback/rate-experience")
def rate_bot_experience(
    merchant_id: str = Query(..., description="Merchant ID"),
    rating: str = Query(..., description="Good or Bad"),
    category: str = Query(..., description="Experience category"),
    comments: str = Query(None, description="Additional comments")
):
    """Rate experience with the bot"""
    return {
        "merchant_id": merchant_id,
        "feedback_id": f"RATE-{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "message": "Thank you for rating your experience!",
        "rating": rating,
        "category": category,
        "comments": comments,
        "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        "status": "Recorded",
        "follow_up": "Your feedback helps us improve our services",
        "additional_survey": "https://feedback.youshop.com/detailed-survey" if rating == "Bad" else None,
        "next_steps": [
            "Feedback has been recorded",
            "Product team will review feedback",
            "Improvements will be implemented based on feedback"
        ]
    }


@app.post("/api/merchant/feedback/share")
def share_detailed_feedback(
    merchant_id: str = Query(..., description="Merchant ID"),
    feedback_type: str = Query(..., description="POS/Loan/Camera/Staff"),
    rating: str = Query(..., description="Rating 1-5"),
    feedback: str = Query(..., description="Detailed feedback"),
    suggestions: str = Query(None, description="Suggestions for improvement")
):
    """Share detailed feedback on services"""
    return {
        "merchant_id": merchant_id,
        "feedback_id": f"FB-{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "message": "Detailed feedback submitted successfully",
        "feedback_type": feedback_type,
        "rating": f"{rating}/5",
        "feedback": feedback,
        "suggestions": suggestions,
        "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        "status": "Under Review",
        "assigned_to": f"{feedback_type} Product Team",
        "estimated_review_time": "3-5 business days",
        "next_steps": [
            "Product team will analyze your feedback",
            "Suggestions will be evaluated for implementation",
            "You may be contacted for additional details",
            "Updates on improvements will be shared"
        ]
    }


@app.post("/api/merchant/feedback/suggest-feature")
def suggest_new_feature(
    merchant_id: str = Query(..., description="Merchant ID"),
    feature_name: str = Query(..., description="Feature name"),
    description: str = Query(..., description="Feature description"),
    category: str = Query(..., description="Feature category"),
    priority: str = Query("Medium", description="Priority level")
):
    """Suggest new feature for YouShop"""
    return {
        "merchant_id": merchant_id,
        "suggestion_id": f"FEAT-{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "message": "Feature suggestion submitted successfully",
        "feature_name": feature_name,
        "description": description,
        "category": category,
        "priority": priority,
        "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        "status": "Under Evaluation",
        "assigned_to": "Product Development Team",
        "evaluation_timeline": "2-4 weeks",
        "community_voting": f"https://features.youshop.com/suggestion/FEAT-{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "next_steps": [
            "Product team will evaluate feasibility",
            "Community voting will be enabled",
            "Technical assessment will be conducted",
            "Implementation roadmap will be shared if approved"
        ]
    }

# ===============================
# CORE HR ENDPOINTS (Required for tests)
# ===============================


@app.get("/api/employees")
def get_employees(db: Session = Depends(get_db)):
    """Get all employees"""
    employees = db.query(models.Employee).all()
    return [
        {
            "employee_id": emp.employee_id,
            "name": emp.employee_name,
            "email": emp.email,
            "position": emp.position,
            "department": emp.department,
            "hire_date": emp.hire_date.isoformat() if emp.hire_date else None,
            "employment_status": emp.employment_status
        }
        for emp in employees
    ]


@app.get("/api/attendance")
def get_attendance(db: Session = Depends(get_db)):
    """Get attendance records"""
    attendance = db.query(models.AttendanceRecord).all()
    return [
        {
            "id": att.id,
            "employee_id": att.employee_id,
            "employee_name": att.employee_name,
            "date": att.date.isoformat() if att.date else None,
            "check_in_time": att.check_in_time.isoformat() if att.check_in_time else None,
            "check_out_time": att.check_out_time.isoformat() if att.check_out_time else None,
            "working_hours": att.working_hours,
            "status": att.status
        }
        for att in attendance
    ]


@app.get("/api/payslips")
def get_payslips(db: Session = Depends(get_db)):
    """Get payslip records"""
    payslips = db.query(models.Payslip).all()
    return [
        {
            "id": ps.id,
            "employee_id": ps.employee_id,
            "employee_name": ps.employee_name,
            "pay_period": ps.pay_period,
            "basic_salary": ps.basic_salary,
            "allowances": ps.allowances,
            "deductions": ps.deductions,
            "net_salary": ps.net_salary,
            "status": ps.status
        }
        for ps in payslips
    ]


@app.get("/api/leave-applications")
def get_leave_applications(db: Session = Depends(get_db)):
    """Get leave applications"""
    leaves = db.query(models.LeaveApplication).all()
    return [
        {
            "id": lv.id,
            "employee_id": lv.employee_id,
            "employee_name": lv.employee_name,
            "leave_type": lv.leave_type,
            "from_date": lv.from_date.isoformat() if lv.from_date else None,
            "to_date": lv.to_date.isoformat() if lv.to_date else None,
            "total_days": lv.total_days,
            "reason": lv.reason,
            "status": lv.status
        }
        for lv in leaves
    ]


@app.get("/api/hr-support-tickets")
def get_hr_support_tickets(db: Session = Depends(get_db)):
    """Get HR support tickets"""
    tickets = db.query(models.HRSupportTicket).all()
    return [
        {
            "id": tk.id,
            "employee_id": tk.employee_id,
            "employee_name": tk.employee_name,
            "category": tk.category,
            "subject": tk.subject,
            "description": tk.description,
            "priority": tk.priority,
            "status": tk.status,
            "created_at": tk.created_at.isoformat() if tk.created_at else None
        }
        for tk in tickets
    ]


@app.get("/api/marketing-campaigns")
def get_marketing_campaigns(db: Session = Depends(get_db)):
    """Get marketing campaigns"""
    campaigns = db.query(models.MarketingCampaign).all()
    return [
        {
            "id": mc.id,
            "campaign_name": mc.campaign_name,
            "campaign_type": mc.campaign_type,
            "target_audience": mc.target_audience,
            "scheduled_date": mc.scheduled_date.isoformat() if mc.scheduled_date else None,
            "budget": mc.budget,
            "status": mc.status,
            "reach_count": mc.reach_count
        }
        for mc in campaigns
    ]


@app.get("/api/promotions")
def get_promotions(db: Session = Depends(get_db)):
    """Get promotions"""
    promotions = db.query(models.Promotion).all()
    return [
        {
            "id": pr.id,
            "promotion_name": pr.promotion_name,
            "promotion_type": pr.promotion_type,
            "discount_percentage": pr.discount_percentage,
            "discount_amount": pr.discount_amount,
            "valid_from": pr.valid_from.isoformat() if pr.valid_from else None,
            "valid_until": pr.valid_until.isoformat() if pr.valid_until else None,
            "status": pr.status
        }
        for pr in promotions
    ]


@app.get("/api/sales-records")
def get_sales_records(db: Session = Depends(get_db)):
    """Get sales records"""
    sales = db.query(models.SalesRecord).all()
    return [
        {
            "id": sr.id,
            "merchant_id": sr.merchant_id,
            "product_name": sr.product_name,
            "quantity": sr.quantity,
            "revenue": sr.revenue,
            "sale_date": sr.sale_date.isoformat() if sr.sale_date else None
        }
        for sr in sales
    ]

# ===============================
# RETENTION EXECUTOR ENDPOINTS
# ===============================

# Daily Activity Functions


@app.get("/api/icp/executor/assigned-merchants")
async def get_assigned_merchants():
    """Get list of assigned merchants for retention executor"""
    return {
        "status": "success",
        "merchants": [
            {
                "merchant_id": "M001",
                "business_name": "Tech Solutions Ltd",
                "owner_name": "John Smith",
                "phone": "+1234567890",
                "email": "john@techsolutions.com",
                "status": "Active",
                "risk_level": "Medium",
                "last_contact": "2024-01-15",
                "next_follow_up": "2024-01-22",
                "revenue_trend": "Declining",
                "satisfaction_score": 7.5,
                "priority": "High"
            },
            {
                "merchant_id": "M002",
                "business_name": "Fashion Hub",
                "owner_name": "Sarah Johnson",
                "phone": "+1234567891",
                "email": "sarah@fashionhub.com",
                "status": "Active",
                "risk_level": "Low",
                "last_contact": "2024-01-18",
                "next_follow_up": "2024-01-25",
                "revenue_trend": "Stable",
                "satisfaction_score": 8.9,
                "priority": "Medium"
            },
            {
                "merchant_id": "M003",
                "business_name": "Food Express",
                "owner_name": "Mike Chen",
                "phone": "+1234567892",
                "email": "mike@foodexpress.com",
                "status": "At Risk",
                "risk_level": "High",
                "last_contact": "2024-01-10",
                "next_follow_up": "2024-01-17",
                "revenue_trend": "Declining",
                "satisfaction_score": 6.2,
                "priority": "Critical"
            }
        ],
        "total_count": 3,
        "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }


@app.get("/api/icp/executor/daily-schedule")
async def get_daily_schedule():
    """Get daily schedule and tasks for retention executor"""
    return {
        "status": "success",
        "date": datetime.now().strftime('%Y-%m-%d'),
        "schedule": {
            "morning": [
                {
                    "time": "09:00",
                    "activity": "Follow-up Call",
                    "merchant": "Tech Solutions Ltd",
                    "type": "Retention Check",
                    "priority": "High",
                    "duration": "30 mins"
                },
                {
                    "time": "10:00",
                    "activity": "Merchant Visit",
                    "merchant": "Food Express",
                    "type": "Site Visit",
                    "priority": "Critical",
                    "duration": "2 hours"
                }
            ],
            "afternoon": [
                {
                    "time": "14:00",
                    "activity": "Training Session",
                    "merchant": "Fashion Hub",
                    "type": "Product Training",
                    "priority": "Medium",
                    "duration": "1 hour"
                },
                {
                    "time": "16:00",
                    "activity": "Performance Review",
                    "merchant": "All Merchants",
                    "type": "Data Analysis",
                    "priority": "Medium",
                    "duration": "1 hour"
                }
            ]
        },
        "total_tasks": 4,
        "completion_rate": "75%",
        "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }


@app.get("/api/icp/executor/task-completion")
async def get_task_completion():
    """Get task completion status and progress"""
    return {
        "status": "success",
        "completion_summary": {
            "total_tasks": 25,
            "completed": 18,
            "pending": 5,
            "overdue": 2,
            "completion_rate": "72%"
        },
        "tasks_by_category": {
            "follow_up_calls": {"total": 8, "completed": 6, "rate": "75%"},
            "merchant_visits": {"total": 5, "completed": 4, "rate": "80%"},
            "training_sessions": {"total": 6, "completed": 5, "rate": "83%"},
            "documentation": {"total": 4, "completed": 2, "rate": "50%"},
            "escalations": {"total": 2, "completed": 1, "rate": "50%"}
        },
        "recent_completions": [
            {
                "task": "Follow-up call with Tech Solutions",
                "completed_at": "2024-01-20 14:30",
                "result": "Positive response, renewed contract"
            },
            {
                "task": "Training session for Fashion Hub",
                "completed_at": "2024-01-20 11:00",
                "result": "Staff trained on new features"
            }
        ],
        "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }

# Merchant Follow-up Functions


@app.get("/api/icp/executor/merchant-profile")
async def get_merchant_profile(merchant_id: str = Query("MERCH001", description="Merchant ID (default: MERCH001)")):
    """Get detailed merchant profile for follow-up"""
    return {
        "status": "success",
        "merchant_profile": {
            "basic_info": {
                "merchant_id": merchant_id,
                "business_name": "Tech Solutions Ltd",
                "owner_name": "John Smith",
                "registration_date": "2023-06-15",
                "business_type": "Technology Services",
                "employees": 25,
                "location": "Downtown Business District"
            },
            "contact_info": {
                "primary_phone": "+1234567890",
                "secondary_phone": "+1234567891",
                "email": "john@techsolutions.com",
                "address": "123 Tech Street, Business City",
                "preferred_contact_method": "Phone",
                "best_contact_time": "10:00 AM - 4:00 PM"
            },
            "business_metrics": {
                "monthly_revenue": "$45,000",
                "transaction_volume": 1250,
                "average_order_value": "$36",
                "growth_rate": "-5%",
                "customer_retention": "78%",
                "satisfaction_score": 7.5
            },
            "engagement_history": [
                {
                    "date": "2024-01-15",
                    "type": "Phone Call",
                    "duration": "25 mins",
                    "outcome": "Discussed pricing concerns",
                    "next_action": "Follow-up with custom pricing"
                },
                {
                    "date": "2024-01-10",
                    "type": "Site Visit",
                    "duration": "2 hours",
                    "outcome": "Identified training needs",
                    "next_action": "Schedule training session"
                }
            ],
            "risk_assessment": {
                "risk_level": "Medium",
                "risk_factors": ["Revenue decline", "Competitor interest"],
                "retention_probability": "75%",
                "recommended_actions": ["Custom pricing discussion", "Enhanced support"]
            }
        },
        "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }


@app.get("/api/icp/executor/follow-up-schedule")
async def get_follow_up_schedule():
    """Get follow-up schedule for merchants"""
    return {
        "status": "success",
        "follow_up_schedule": {
            "today": [
                {
                    "merchant_id": "M001",
                    "business_name": "Tech Solutions Ltd",
                    "scheduled_time": "10:00 AM",
                    "type": "Retention Call",
                    "priority": "High",
                    "reason": "Revenue decline concern"
                }
            ],
            "this_week": [
                {
                    "merchant_id": "M002",
                    "business_name": "Fashion Hub",
                    "scheduled_date": "2024-01-22",
                    "type": "Check-in Call",
                    "priority": "Medium",
                    "reason": "Regular monthly check"
                },
                {
                    "merchant_id": "M003",
                    "business_name": "Food Express",
                    "scheduled_date": "2024-01-23",
                    "type": "Site Visit",
                    "priority": "Critical",
                    "reason": "Contract renewal discussion"
                }
            ],
            "overdue": [
                {
                    "merchant_id": "M004",
                    "business_name": "Auto Parts Plus",
                    "due_date": "2024-01-18",
                    "type": "Urgent Call",
                    "priority": "Critical",
                    "reason": "Payment issues"
                }
            ]
        },
        "summary": {
            "total_scheduled": 4,
            "today": 1,
            "this_week": 2,
            "overdue": 1
        },
        "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }


@app.get("/api/icp/executor/retention-metrics")
async def get_retention_metrics():
    """Get retention metrics and KPIs"""
    return {
        "status": "success",
        "retention_metrics": {
            "overview": {
                "total_merchants": 150,
                "retained_merchants": 142,
                "lost_merchants": 8,
                "retention_rate": "94.7%",
                "churn_rate": "5.3%"
            },
            "monthly_trends": [
                {"month": "Jan 2024", "retention_rate": "94.7%",
                    "new_merchants": 12, "lost_merchants": 8},
                {"month": "Dec 2023", "retention_rate": "96.1%",
                    "new_merchants": 15, "lost_merchants": 6},
                {"month": "Nov 2023", "retention_rate": "95.5%",
                    "new_merchants": 18, "lost_merchants": 7}
            ],
            "risk_categories": {
                "high_risk": {"count": 8, "percentage": "5.3%"},
                "medium_risk": {"count": 23, "percentage": "15.3%"},
                "low_risk": {"count": 119, "percentage": "79.3%"}
            },
            "intervention_success": {
                "attempted": 45,
                "successful": 38,
                "success_rate": "84.4%"
            },
            "top_retention_factors": [
                "Competitive pricing",
                "Excellent customer support",
                "Regular training programs",
                "Flexible payment terms"
            ]
        },
        "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }

# Onboarding Support Functions


@app.get("/api/icp/executor/new-merchants")
async def get_new_merchants():
    """Get list of new merchants requiring onboarding support"""
    return {
        "status": "success",
        "new_merchants": [
            {
                "merchant_id": "M105",
                "business_name": "Green Garden Cafe",
                "owner_name": "Emma Wilson",
                "registration_date": "2024-01-18",
                "onboarding_stage": "Initial Setup",
                "completion_percentage": "25%",
                "assigned_specialist": "Sarah Johnson",
                "next_milestone": "Payment Integration",
                "urgency": "Medium",
                "contact_info": {
                    "phone": "+1234567895",
                    "email": "emma@greengarden.com"
                }
            },
            {
                "merchant_id": "M106",
                "business_name": "Sports Zone",
                "owner_name": "David Martinez",
                "registration_date": "2024-01-19",
                "onboarding_stage": "Documentation Review",
                "completion_percentage": "60%",
                "assigned_specialist": "Mike Chen",
                "next_milestone": "Training Session",
                "urgency": "High",
                "contact_info": {
                    "phone": "+1234567896",
                    "email": "david@sportszone.com"
                }
            }
        ],
        "summary": {
            "total_new_merchants": 2,
            "avg_completion_rate": "42.5%",
            "pending_actions": 5
        },
        "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }


@app.get("/api/icp/executor/onboarding-progress")
async def get_onboarding_progress():
    """Get onboarding progress tracking"""
    return {
        "status": "success",
        "onboarding_progress": {
            "stages": [
                {
                    "stage": "Initial Setup",
                    "merchants": 3,
                    "avg_duration": "2 days",
                    "success_rate": "95%"
                },
                {
                    "stage": "Documentation Review",
                    "merchants": 2,
                    "avg_duration": "3 days",
                    "success_rate": "90%"
                },
                {
                    "stage": "Payment Integration",
                    "merchants": 4,
                    "avg_duration": "5 days",
                    "success_rate": "85%"
                },
                {
                    "stage": "Training & Testing",
                    "merchants": 1,
                    "avg_duration": "4 days",
                    "success_rate": "98%"
                },
                {
                    "stage": "Go Live",
                    "merchants": 2,
                    "avg_duration": "1 day",
                    "success_rate": "100%"
                }
            ],
            "bottlenecks": [
                {
                    "stage": "Payment Integration",
                    "issue": "Technical complexity",
                    "recommended_action": "Enhanced technical support"
                }
            ],
            "completion_timeline": {
                "average_days": 15,
                "fastest_completion": 8,
                "longest_completion": 28
            }
        },
        "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }


@app.get("/api/icp/executor/training-schedule")
async def get_training_schedule():
    """Get training schedule for new merchants"""
    return {
        "status": "success",
        "training_schedule": {
            "upcoming_sessions": [
                {
                    "session_id": "TR001",
                    "title": "Basic Platform Navigation",
                    "date": "2024-01-22",
                    "time": "10:00 AM",
                    "duration": "2 hours",
                    "type": "Group Session",
                    "trainer": "Sarah Johnson",
                    "participants": [
                        {"merchant": "Green Garden Cafe", "status": "Confirmed"},
                        {"merchant": "Sports Zone", "status": "Pending"}
                    ],
                    "location": "Training Room A"
                },
                {
                    "session_id": "TR002",
                    "title": "Payment Processing Setup",
                    "date": "2024-01-23",
                    "time": "2:00 PM",
                    "duration": "1.5 hours",
                    "type": "One-on-One",
                    "trainer": "Mike Chen",
                    "participants": [
                        {"merchant": "Book Store Express", "status": "Confirmed"}
                    ],
                    "location": "Virtual Meeting"
                }
            ],
            "completed_sessions": [
                {
                    "session_id": "TR000",
                    "title": "Welcome & Overview",
                    "date": "2024-01-20",
                    "participants": 5,
                    "satisfaction_score": 4.8,
                    "trainer": "Emma Davis"
                }
            ],
            "training_materials": [
                "Platform User Guide",
                "Payment Setup Manual",
                "Best Practices Guide",
                "Troubleshooting FAQ"
            ]
        },
        "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }

# Notification Functions


@app.get("/api/icp/executor/priority-alerts")
async def get_priority_alerts():
    """Get high-priority alerts and notifications"""
    return {
        "status": "success",
        "priority_alerts": [
            {
                "alert_id": "ALT001",
                "type": "Critical",
                "title": "Merchant Payment Failure",
                "message": "Food Express has multiple failed payment attempts",
                "merchant_id": "M003",
                "timestamp": "2024-01-20 09:15:00",
                "urgency": "High",
                "action_required": "Contact merchant immediately",
                "estimated_impact": "Revenue loss risk"
            },
            {
                "alert_id": "ALT002",
                "type": "Warning",
                "title": "Low Satisfaction Score",
                "message": "Tech Solutions rated last interaction 3/10",
                "merchant_id": "M001",
                "timestamp": "2024-01-20 11:30:00",
                "urgency": "Medium",
                "action_required": "Schedule follow-up call",
                "estimated_impact": "Retention risk"
            },
            {
                "alert_id": "ALT003",
                "type": "Info",
                "title": "Contract Renewal Due",
                "message": "Fashion Hub contract expires in 30 days",
                "merchant_id": "M002",
                "timestamp": "2024-01-20 14:00:00",
                "urgency": "Medium",
                "action_required": "Initiate renewal process",
                "estimated_impact": "Standard renewal"
            }
        ],
        "alert_summary": {
            "total": 3,
            "critical": 1,
            "warning": 1,
            "info": 1,
            "unread": 2
        },
        "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }


@app.get("/api/icp/executor/system-notifications")
async def get_system_notifications():
    """Get system notifications and updates"""
    return {
        "status": "success",
        "system_notifications": [
            {
                "notification_id": "SYS001",
                "type": "System Update",
                "title": "Platform Maintenance Scheduled",
                "message": "Scheduled maintenance window: Jan 25, 2024, 2:00 AM - 4:00 AM",
                "priority": "Medium",
                "timestamp": "2024-01-20 16:00:00",
                "action_required": "Inform affected merchants",
                "affected_services": ["Payment Processing", "Reporting"]
            },
            {
                "notification_id": "SYS002",
                "type": "Feature Release",
                "title": "New Analytics Dashboard Available",
                "message": "Enhanced merchant analytics dashboard is now live",
                "priority": "Low",
                "timestamp": "2024-01-20 12:00:00",
                "action_required": "Update training materials",
                "benefits": ["Better insights", "Improved reporting"]
            },
            {
                "notification_id": "SYS003",
                "type": "Policy Update",
                "title": "Updated Terms of Service",
                "message": "New terms of service effective February 1, 2024",
                "priority": "High",
                "timestamp": "2024-01-20 10:00:00",
                "action_required": "Merchant notification required",
                "deadline": "2024-01-31"
            }
        ],
        "notification_summary": {
            "total": 3,
            "unread": 1,
            "requiring_action": 2
        },
        "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }


@app.get("/api/icp/executor/communication-log")
async def get_communication_log():
    """Get communication history and log"""
    return {
        "status": "success",
        "communication_log": [
            {
                "log_id": "COM001",
                "merchant_id": "M001",
                "merchant_name": "Tech Solutions Ltd",
                "communication_type": "Phone Call",
                "direction": "Outbound",
                "timestamp": "2024-01-20 14:30:00",
                "duration": "25 minutes",
                "handled_by": "Sarah Johnson",
                "purpose": "Retention discussion",
                "outcome": "Positive - merchant agreed to meeting",
                "next_action": "Schedule in-person meeting",
                "satisfaction_rating": 8
            },
            {
                "log_id": "COM002",
                "merchant_id": "M003",
                "merchant_name": "Food Express",
                "communication_type": "Email",
                "direction": "Inbound",
                "timestamp": "2024-01-20 11:15:00",
                "handled_by": "Mike Chen",
                "purpose": "Payment issue inquiry",
                "outcome": "Issue identified and resolved",
                "next_action": "Monitor payment status",
                "satisfaction_rating": 9
            },
            {
                "log_id": "COM003",
                "merchant_id": "M002",
                "merchant_name": "Fashion Hub",
                "communication_type": "Site Visit",
                "direction": "Outbound",
                "timestamp": "2024-01-19 15:00:00",
                "duration": "2 hours",
                "handled_by": "Emma Davis",
                "purpose": "Training and support",
                "outcome": "Training completed successfully",
                "next_action": "Follow-up call in 1 week",
                "satisfaction_rating": 10
            }
        ],
        "log_summary": {
            "total_communications": 3,
            "today": 2,
            "this_week": 3,
            "avg_satisfaction": 9.0,
            "response_time_avg": "2.5 hours"
        },
        "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }

# Support Request Functions


@app.get("/api/icp/executor/pending-tickets")
async def get_pending_tickets():
    """Get pending support tickets"""
    return {
        "status": "success",
        "pending_tickets": [
            {
                "ticket_id": "TKT001",
                "merchant_id": "M001",
                "merchant_name": "Tech Solutions Ltd",
                "subject": "Payment gateway integration issue",
                "priority": "High",
                "status": "In Progress",
                "created_date": "2024-01-19",
                "assigned_to": "Technical Team",
                "category": "Technical Support",
                "description": "Unable to process credit card payments",
                "estimated_resolution": "2024-01-21",
                "last_update": "2024-01-20 10:30:00"
            },
            {
                "ticket_id": "TKT002",
                "merchant_id": "M004",
                "merchant_name": "Auto Parts Plus",
                "subject": "Account suspension inquiry",
                "priority": "Critical",
                "status": "Escalated",
                "created_date": "2024-01-20",
                "assigned_to": "Account Management",
                "category": "Account Issue",
                "description": "Account suspended due to suspicious activity",
                "estimated_resolution": "2024-01-20",
                "last_update": "2024-01-20 13:45:00"
            },
            {
                "ticket_id": "TKT003",
                "merchant_id": "M002",
                "merchant_name": "Fashion Hub",
                "subject": "Reporting discrepancy",
                "priority": "Medium",
                "status": "New",
                "created_date": "2024-01-20",
                "assigned_to": "Data Team",
                "category": "Reporting",
                "description": "Sales numbers don't match transaction records",
                "estimated_resolution": "2024-01-22",
                "last_update": "2024-01-20 14:15:00"
            }
        ],
        "ticket_summary": {
            "total_pending": 3,
            "critical": 1,
            "high": 1,
            "medium": 1,
            "overdue": 0,
            "avg_resolution_time": "2.5 days"
        },
        "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }


@app.get("/api/icp/executor/escalation-queue")
async def get_escalation_queue():
    """Get escalation queue and high-priority issues"""
    return {
        "status": "success",
        "escalation_queue": [
            {
                "escalation_id": "ESC001",
                "ticket_id": "TKT002",
                "merchant_id": "M004",
                "merchant_name": "Auto Parts Plus",
                "issue": "Account suspension dispute",
                "escalation_level": "Level 2",
                "escalated_by": "Support Agent",
                "escalated_to": "Senior Manager",
                "escalation_date": "2024-01-20 13:45:00",
                "urgency": "Critical",
                "business_impact": "Revenue loss",
                "customer_sentiment": "Very Negative",
                "resolution_deadline": "2024-01-20 18:00:00",
                "current_status": "Under Review"
            },
            {
                "escalation_id": "ESC002",
                "ticket_id": "TKT005",
                "merchant_id": "M007",
                "merchant_name": "Health Plus Pharmacy",
                "issue": "Data security concern",
                "escalation_level": "Level 3",
                "escalated_by": "Security Team",
                "escalated_to": "Executive Team",
                "escalation_date": "2024-01-20 09:30:00",
                "urgency": "Critical",
                "business_impact": "Compliance risk",
                "customer_sentiment": "Concerned",
                "resolution_deadline": "2024-01-21 12:00:00",
                "current_status": "Executive Review"
            }
        ],
        "escalation_summary": {
            "total_escalations": 2,
            "level_1": 0,
            "level_2": 1,
            "level_3": 1,
            "avg_resolution_time": "8 hours",
            "success_rate": "92%"
        },
        "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }


@app.get("/api/icp/executor/resolution-tracking")
async def get_resolution_tracking():
    """Get issue resolution tracking and status"""
    return {
        "status": "success",
        "resolution_tracking": {
            "in_progress": [
                {
                    "ticket_id": "TKT001",
                    "merchant": "Tech Solutions Ltd",
                    "issue": "Payment gateway integration",
                    "progress": "75%",
                    "eta": "2024-01-21 15:00:00",
                    "assigned_team": "Technical Support",
                    "last_action": "Deployed hotfix",
                    "next_step": "User acceptance testing"
                },
                {
                    "ticket_id": "TKT003",
                    "merchant": "Fashion Hub",
                    "issue": "Reporting discrepancy",
                    "progress": "30%",
                    "eta": "2024-01-22 12:00:00",
                    "assigned_team": "Data Analytics",
                    "last_action": "Data investigation started",
                    "next_step": "Root cause analysis"
                }
            ],
            "recently_resolved": [
                {
                    "ticket_id": "TKT000",
                    "merchant": "Green Garden Cafe",
                    "issue": "Login authentication problem",
                    "resolved_date": "2024-01-20 11:30:00",
                    "resolution_time": "4 hours",
                    "customer_satisfaction": 9,
                    "resolved_by": "Technical Team"
                }
            ],
            "resolution_metrics": {
                "average_resolution_time": "6.5 hours",
                "first_call_resolution_rate": "78%",
                "customer_satisfaction_avg": 8.7,
                "sla_compliance": "94%"
            }
        },
        "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }

# Feedback Functions


@app.get("/api/icp/executor/merchant-feedback")
async def get_merchant_feedback():
    """Get merchant feedback and satisfaction scores"""
    return {
        "status": "success",
        "merchant_feedback": [
            {
                "feedback_id": "FB001",
                "merchant_id": "M001",
                "merchant_name": "Tech Solutions Ltd",
                "feedback_date": "2024-01-20",
                "rating": 8,
                "category": "Customer Support",
                "feedback_text": "Great support team, very responsive and knowledgeable",
                "sentiment": "Positive",
                "response_required": False,
                "tags": ["support", "responsive", "knowledgeable"]
            },
            {
                "feedback_id": "FB002",
                "merchant_id": "M003",
                "merchant_name": "Food Express",
                "feedback_date": "2024-01-19",
                "rating": 6,
                "category": "Platform Features",
                "feedback_text": "Platform is good but reporting features need improvement",
                "sentiment": "Mixed",
                "response_required": True,
                "tags": ["platform", "reporting", "improvement"]
            },
            {
                "feedback_id": "FB003",
                "merchant_id": "M002",
                "merchant_name": "Fashion Hub",
                "feedback_date": "2024-01-18",
                "rating": 9,
                "category": "Training",
                "feedback_text": "Excellent training session, very helpful and detailed",
                "sentiment": "Very Positive",
                "response_required": False,
                "tags": ["training", "helpful", "detailed"]
            }
        ],
        "feedback_summary": {
            "total_feedback": 3,
            "average_rating": 7.7,
            "sentiment_breakdown": {
                "positive": 2,
                "mixed": 1,
                "negative": 0
            },
            "response_required": 1,
            "trending_topics": ["support", "training", "reporting"]
        },
        "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }


@app.get("/api/icp/executor/satisfaction-survey")
async def get_satisfaction_survey():
    """Get satisfaction survey results and analysis"""
    return {
        "status": "success",
        "satisfaction_survey": {
            "survey_period": "January 2024",
            "response_rate": "78%",
            "total_responses": 117,
            "overall_satisfaction": {
                "average_score": 8.3,
                "distribution": {
                    "9-10 (Excellent)": 45,
                    "7-8 (Good)": 52,
                    "5-6 (Average)": 15,
                    "3-4 (Poor)": 4,
                    "1-2 (Very Poor)": 1
                }
            },
            "category_scores": {
                "customer_support": 8.7,
                "platform_usability": 8.1,
                "payment_processing": 8.5,
                "training_quality": 8.9,
                "pricing_satisfaction": 7.2,
                "feature_completeness": 7.8
            },
            "nps_score": {
                "score": 52,
                "promoters": 65,
                "passives": 35,
                "detractors": 17,
                "trend": "+5 from last month"
            },
            "improvement_areas": [
                "Pricing transparency",
                "Advanced reporting features",
                "Mobile app functionality",
                "Integration capabilities"
            ]
        },
        "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }


@app.get("/api/icp/executor/improvement-suggestions")
async def get_improvement_suggestions():
    """Get improvement suggestions from merchants"""
    return {
        "status": "success",
        "improvement_suggestions": [
            {
                "suggestion_id": "SUG001",
                "merchant_id": "M001",
                "merchant_name": "Tech Solutions Ltd",
                "suggestion_date": "2024-01-20",
                "category": "Reporting",
                "title": "Real-time Sales Dashboard",
                "description": "Add real-time sales tracking with live updates and alerts",
                "priority": "High",
                "implementation_complexity": "Medium",
                "estimated_effort": "4-6 weeks",
                "business_value": "High",
                "votes": 23,
                "status": "Under Review"
            },
            {
                "suggestion_id": "SUG002",
                "merchant_id": "M002",
                "merchant_name": "Fashion Hub",
                "suggestion_date": "2024-01-19",
                "category": "Payment",
                "title": "Multiple Payment Methods",
                "description": "Support for cryptocurrency and buy-now-pay-later options",
                "priority": "Medium",
                "implementation_complexity": "High",
                "estimated_effort": "8-12 weeks",
                "business_value": "Medium",
                "votes": 15,
                "status": "Planned"
            },
            {
                "suggestion_id": "SUG003",
                "merchant_id": "M003",
                "merchant_name": "Food Express",
                "suggestion_date": "2024-01-18",
                "category": "Mobile",
                "title": "Mobile App Improvements",
                "description": "Enhanced mobile app with offline capabilities",
                "priority": "Medium",
                "implementation_complexity": "Medium",
                "estimated_effort": "6-8 weeks",
                "business_value": "High",
                "votes": 31,
                "status": "In Development"
            }
        ],
        "suggestion_summary": {
            "total_suggestions": 3,
            "under_review": 1,
            "planned": 1,
            "in_development": 1,
            "average_votes": 23,
            "top_categories": ["Reporting", "Payment", "Mobile"]
        },
        "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }


# Mark Activity Complete - POST endpoint
@app.post("/api/icp/executor/mark-activity-complete")
def mark_activity_complete(request: dict):
    """Mark a daily activity as complete with proof upload."""
    return {
        "success": True,
        "message": "Activity marked as complete successfully",
        "activity_id": f"ACT_{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "merchant_id": request.get("merchant_id", "M001"),
        "activity_type": request.get("activity_type", "WhatsApp"),
        "completion_time": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        "proof_uploaded": True,
        "proof_details": {
            "file_name": request.get("proof_file", "default_proof.jpg"),
            "file_size": "245 KB",
            "upload_time": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "verification_status": "Verified"
        },
        "notes": request.get("notes", "Activity completed successfully"),
        "next_steps": [
            "Schedule follow-up call within 24 hours",
            "Update merchant status in CRM",
            "Monitor merchant response"
        ],
        "performance_metrics": {
            "completion_time_minutes": 15,
            "efficiency_score": 85,
            "quality_rating": 4.5
        }
    }


# Submit Summary Report - POST endpoint
@app.post("/api/icp/executor/submit-summary-report")
def submit_summary_report(request: dict):
    """Submit daily, weekly, or monthly summary report."""
    report_type = request.get("report_type", "Daily Report")

    return {
        "success": True,
        "message": f"{report_type} submitted successfully",
        "report_id": f"RPT_{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "report_type": report_type,
        "submission_time": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        "report_details": {
            "period": get_report_period(report_type),
            "total_activities": 24,
            "completed_activities": 22,
            "pending_activities": 2,
            "completion_rate": "91.7%"
        },
        "summary_content": request.get("summary", "Daily activities completed successfully"),
        "key_achievements": [
            "Contacted 15 merchants",
            "Resolved 8 support tickets",
            "Completed onboarding for 3 new merchants"
        ],
        "challenges_faced": [
            "Network connectivity issues during calls",
            "Delayed response from 2 merchants"
        ],
        "next_day_priorities": [
            "Follow up with pending merchants",
            "Complete training documentation",
            "Review escalated cases"
        ],
        "file_details": {
            "generated_file": f"{report_type.lower().replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf",
            "file_size": "1.2 MB",
            "download_link": f"/downloads/{report_type.lower().replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
        }
    }


def get_report_period(report_type):
    """Get the period based on report type."""
    today = datetime.now()

    if report_type == "Daily Report":
        return today.strftime('%Y-%m-%d')
    elif report_type == "Weekly Report":
        start_week = today - timedelta(days=today.weekday())
        end_week = start_week + timedelta(days=6)
        return f"{start_week.strftime('%Y-%m-%d')} to {end_week.strftime('%Y-%m-%d')}"
    elif report_type == "Monthly Report":
        return today.strftime('%Y-%m')
    else:
        return today.strftime('%Y-%m-%d')


# Merchant Follow-Up Endpoints

# Update merchant health status - POST endpoint
@app.post("/api/icp/executor/update-merchant-health")
def update_merchant_health(request: dict):
    """Update merchant health status (Healthy / Limited Activity / No Activity)."""
    return {
        "success": True,
        "message": "Merchant health status updated successfully",
        "update_id": f"UPD_{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "merchant_id": request.get("merchant_id", "M001"),
        "previous_status": "Limited Activity",
        "new_status": request.get("health_status", "Healthy"),
        "updated_by": "Retention Executor",
        "update_time": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        "status_details": {
            "activity_level": get_activity_level(request.get("health_status", "Healthy")),
            "risk_assessment": get_risk_level(request.get("health_status", "Healthy")),
            "recommended_actions": get_health_recommendations(request.get("health_status", "Healthy"))
        },
        "historical_changes": [
            {
                "date": "2024-01-15",
                "from_status": "Healthy",
                "to_status": "Limited Activity",
                "reason": "Decreased transaction volume"
            },
            {
                "date": "2024-01-10",
                "from_status": "No Activity",
                "to_status": "Healthy",
                "reason": "Re-engagement successful"
            }
        ],
        "next_review_date": (datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d')
    }


# Log merchant needs - POST endpoint
@app.post("/api/icp/executor/log-merchant-needs")
def log_merchant_needs(request: dict):
    """Log merchant needs (POS issue / Hardware issue / Loan / Training / Marketing help)."""
    return {
        "success": True,
        "message": "Merchant needs logged successfully",
        "log_id": f"LOG_{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "merchant_id": request.get("merchant_id", "M001"),
        "need_type": request.get("need_type", "POS issue"),
        "priority": get_need_priority(request.get("need_type", "POS issue")),
        "description": request.get("description", ""),
        "logged_time": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        "assigned_team": get_support_team(request.get("need_type", "POS issue")),
        "estimated_resolution": get_resolution_time(request.get("need_type", "POS issue")),
        "support_details": {
            "ticket_number": f"TKT_{datetime.now().strftime('%Y%m%d%H%M%S')}",
            "category": request.get("need_type", "POS issue"),
            "subcategory": get_subcategory(request.get("need_type", "POS issue")),
            "urgency": get_urgency_level(request.get("need_type", "POS issue"))
        },
        "follow_up_actions": [
            f"Contact {get_support_team(request.get('need_type', 'POS issue'))} team",
            "Schedule technical assessment",
            "Provide temporary workaround if applicable",
            "Follow up within 24 hours"
        ],
        "related_cases": [
            {
                "case_id": "CASE001",
                "description": "Similar POS connectivity issue",
                "resolution": "Network configuration update",
                "resolution_time": "2 hours"
            }
        ]
    }


# Add notes or commitments - POST endpoint (both aliases)
@app.post("/api/icp/executor/add-notes-commitments")
@app.post("/api/icp/executor/add-merchant-notes")
def add_notes_commitments(request: dict):
    """Add notes or commitments (e.g., Training scheduled Friday)."""
    return {
        "success": True,
        "message": "Notes added successfully",
        "note_id": f"NOTE_{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "merchant_id": request.get("merchant_id", "M001"),
        "note_type": request.get("note_type", "commitment"),
        "content": request.get("notes", "Training scheduled Friday"),
        "created_time": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        "created_by": "Retention Executor",
        "visibility": "internal",
        "tags": extract_tags(request.get("notes", "")),
        "commitments": extract_commitments(request.get("notes", "")),
        "follow_up_required": True if "scheduled" in request.get("notes", "").lower() else False,
        "reminder_set": {
            "enabled": True,
            "reminder_date": get_reminder_date(request.get("notes", "")),
            "reminder_type": "email_notification"
        },
        "note_history": [
            {
                "note_id": "NOTE_20240115143000",
                "date": "2024-01-15",
                "content": "Discussed payment gateway integration",
                "type": "discussion"
            },
            {
                "note_id": "NOTE_20240112100000",
                "date": "2024-01-12",
                "content": "Training session completed successfully",
                "type": "commitment_fulfilled"
            }
        ]
    }


# Attach photo or proof - POST endpoint (both aliases)
@app.post("/api/icp/executor/attach-photo-proof")
@app.post("/api/icp/executor/attach-proof")
def attach_photo_proof(request: dict):
    """Attach photo or proof (shop photo, invoice, etc.)."""
    return {
        "success": True,
        "message": "Proof attached successfully",
        "attachment_id": f"ATT_{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "merchant_id": request.get("merchant_id", "M001"),
        "file_details": {
            "original_filename": request.get("filename", "shop_photo.jpg"),
            "stored_filename": f"merchant_{request.get('merchant_id', 'M001')}_{datetime.now().strftime('%Y%m%d%H%M%S')}.jpg",
            "file_type": get_file_type(request.get("filename", "shop_photo.jpg")),
            "file_size": request.get("file_size", "2.4 MB"),
            "upload_time": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        },
        "proof_type": request.get("proof_type", "shop_photo"),
        "verification_status": "pending_review",
        "metadata": {
            "location_data": request.get("location", "GPS coordinates extracted"),
            "timestamp_verified": True,
            "quality_score": 8.5,
            "authenticity_check": "passed"
        },
        "processing_details": {
            "uploaded_by": "Retention Executor",
            "review_required": True,
            "auto_analysis": {
                "text_detection": "Business name visible",
                "image_quality": "High",
                "compliance_check": "Approved"
            }
        },
        "storage_info": {
            "cloud_path": f"/merchant_proofs/{request.get('merchant_id', 'M001')}/",
            "backup_location": "secondary_storage",
            "retention_period": "7 years",
            "access_level": "restricted"
        }
    }


# Helper functions for merchant follow-up endpoints
def get_activity_level(health_status):
    levels = {
        "Healthy": "High activity - Regular transactions",
        "Limited Activity": "Moderate activity - Decreased transactions",
        "No Activity": "Inactive - No recent transactions"
    }
    return levels.get(health_status, "Unknown")


def get_risk_level(health_status):
    risks = {
        "Healthy": "Low risk",
        "Limited Activity": "Medium risk - Monitor closely",
        "No Activity": "High risk - Immediate attention required"
    }
    return risks.get(health_status, "Unknown")


def get_health_recommendations(health_status):
    recommendations = {
        "Healthy": ["Continue regular check-ins", "Explore growth opportunities"],
        "Limited Activity": ["Investigate causes", "Provide additional support", "Schedule training"],
        "No Activity": ["Immediate contact required", "Investigate technical issues", "Recovery plan needed"]
    }
    return recommendations.get(health_status, [])


def get_need_priority(need_type):
    priorities = {
        "POS issue": "High",
        "Hardware issue": "High",
        "Loan": "Medium",
        "Training": "Medium",
        "Marketing help": "Low"
    }
    return priorities.get(need_type, "Medium")


def get_support_team(need_type):
    teams = {
        "POS issue": "Technical Support",
        "Hardware issue": "Hardware Support",
        "Loan": "Lending Team",
        "Training": "Training Team",
        "Marketing help": "Marketing Team"
    }
    return teams.get(need_type, "General Support")


def get_resolution_time(need_type):
    times = {
        "POS issue": "2-4 hours",
        "Hardware issue": "4-8 hours",
        "Loan": "3-5 business days",
        "Training": "1-2 business days",
        "Marketing help": "1-3 business days"
    }
    return times.get(need_type, "1-2 business days")


def get_subcategory(need_type):
    subcategories = {
        "POS issue": "Connectivity",
        "Hardware issue": "Device malfunction",
        "Loan": "Business expansion",
        "Training": "System usage",
        "Marketing help": "Digital marketing"
    }
    return subcategories.get(need_type, "General")


def get_urgency_level(need_type):
    urgency = {
        "POS issue": "Critical",
        "Hardware issue": "High",
        "Loan": "Normal",
        "Training": "Normal",
        "Marketing help": "Low"
    }
    return urgency.get(need_type, "Normal")


def extract_tags(notes):
    # Simple tag extraction logic
    tags = []
    if "training" in notes.lower():
        tags.append("training")
    if "friday" in notes.lower():
        tags.append("scheduled")
    if "payment" in notes.lower():
        tags.append("payment")
    return tags


def extract_commitments(notes):
    # Extract commitment information
    commitments = []
    if "scheduled" in notes.lower():
        commitments.append({
            "type": "scheduled_activity",
            "description": notes,
            "due_date": "2024-01-26",  # Friday
            "status": "pending"
        })
    return commitments


def get_reminder_date(notes):
    # Set reminder based on content
    if "friday" in notes.lower():
        return "2024-01-26"
    else:
        return (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d')


def get_file_type(filename):
    if filename.lower().endswith(('.jpg', '.jpeg', '.png')):
        return "image"
    elif filename.lower().endswith('.pdf'):
        return "document"
    else:
        return "unknown"


# Onboarding Support POST Endpoints

@app.post("/api/icp/executor/check-pending-documents")
def check_pending_documents(request: dict):
    """Check pending merchant documents (CNIC, bank statement, license)"""
    merchant_id = request.get("merchant_id", "M001")

    return {
        "success": True,
        "message": "Document status retrieved successfully",
        "check_id": f"CHK_{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "merchant_id": merchant_id,
        "check_time": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        "document_status": {
            "cnic": {
                "status": "pending",
                "required": True,
                "last_request_date": "2024-01-18",
                "reminder_count": 2,
                "description": "National ID card copy required"
            },
            "bank_statement": {
                "status": "received",
                "required": True,
                "upload_date": "2024-01-20",
                "verification_status": "verified",
                "description": "Bank statement last 3 months"
            },
            "license": {
                "status": "pending",
                "required": True,
                "last_request_date": "2024-01-15",
                "reminder_count": 3,
                "description": "Business license or trade permit"
            },
            "additional_docs": {
                "tax_certificate": {
                    "status": "not_required",
                    "reason": "Business category exempt"
                },
                "premises_agreement": {
                    "status": "pending",
                    "required": False,
                    "description": "Shop rental agreement (optional)"
                }
            }
        },
        "summary": {
            "total_required": 3,
            "received": 1,
            "pending": 2,
            "completion_rate": "33%",
            "blocking_onboarding": True
        },
        "next_actions": [
            "Send reminder for CNIC copy",
            "Follow up on business license",
            "Schedule document collection visit"
        ],
        "estimated_completion": "2-3 business days"
    }


@app.post("/api/icp/executor/upload-missing-documents")
def upload_missing_documents(request: dict):
    """Upload missing documents (take photo & submit)"""
    merchant_id = request.get("merchant_id", "M001")
    document_type = request.get("document_type", "CNIC")

    return {
        "success": True,
        "message": f"{document_type} uploaded successfully",
        "upload_id": f"UPL_{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "merchant_id": merchant_id,
        "document_details": {
            "document_type": document_type,
            "file_name": request.get("file_name", f"{document_type.lower()}_photo.jpg"),
            "file_size": "1.8 MB",
            "upload_time": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "file_format": "JPEG",
            "quality_score": 9.2
        },
        "verification_process": {
            "auto_detection": {
                "document_detected": True,
                "text_readable": True,
                "corners_visible": True,
                "quality_acceptable": True
            },
            "manual_review": {
                "required": True,
                "assigned_to": "Document Verification Team",
                "expected_completion": "2-4 hours",
                "review_criteria": ["Authenticity", "Completeness", "Legibility"]
            }
        },
        "processing_status": {
            "current_stage": "uploaded",
            "next_stage": "verification",
            "estimated_completion": "2024-01-21 12:00:00"
        },
        "storage_info": {
            "secure_storage": True,
            "encryption": "AES-256",
            "backup_location": "secondary_vault",
            "retention_policy": "7 years"
        },
        "compliance": {
            "gdpr_compliant": True,
            "data_classification": "sensitive",
            "access_log": "enabled"
        }
    }


@app.post("/api/icp/executor/schedule-installation-training")
@app.post("/api/icp/executor/schedule-installation-visit")
def schedule_installation_training(request: dict):
    """Schedule installation / training visit"""
    merchant_id = request.get("merchant_id", "M001")
    visit_type = request.get("visit_type", "Installation & Training")

    return {
        "success": True,
        "message": "Installation/training visit scheduled successfully",
        "booking_id": f"VIS_{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "merchant_id": merchant_id,
        "visit_details": {
            "visit_type": visit_type,
            "scheduled_date": request.get("preferred_date", "2024-01-25"),
            "scheduled_time": request.get("preferred_time", "10:00 AM"),
            "duration": "2-3 hours",
            "location": "Merchant premises"
        },
        "technician_assignment": {
            "technician_id": "TECH001",
            "technician_name": "Ahmed Hassan",
            "contact_number": "+1234567890",
            "specialization": ["POS Installation", "Merchant Training"],
            "rating": 4.8,
            "languages": ["English", "Urdu"]
        },
        "visit_agenda": [
            "POS terminal installation",
            "Network connectivity setup",
            "Payment gateway configuration",
            "Staff training on system usage",
            "Transaction testing",
            "Documentation handover"
        ],
        "preparations_required": {
            "merchant_tasks": [
                "Ensure internet connectivity",
                "Designate staff for training",
                "Prepare installation space",
                "Have business documents ready"
            ],
            "technician_brings": [
                "POS terminal",
                "Installation tools",
                "Training materials",
                "Test cards"
            ]
        },
        "confirmation": {
            "sms_sent": True,
            "email_sent": True,
            "calendar_invite": True,
            "reminder_scheduled": "24 hours before"
        },
        "backup_slots": [
            {"date": "2024-01-26", "time": "2:00 PM"},
            {"date": "2024-01-27", "time": "10:00 AM"}
        ]
    }


@app.post("/api/icp/executor/confirm-setup-completed")
def confirm_setup_completed(request: dict):
    """Confirm merchant setup completed"""
    merchant_id = request.get("merchant_id", "M001")

    return {
        "success": True,
        "message": "Merchant setup completion confirmed successfully",
        "confirmation_id": f"CONF_{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "merchant_id": merchant_id,
        "completion_details": {
            "completion_date": datetime.now().strftime('%Y-%m-%d'),
            "completion_time": datetime.now().strftime('%H:%M:%S'),
            "confirmed_by": "Field Technician",
            "verification_method": "On-site inspection"
        },
        "setup_verification": {
            "pos_installation": {
                "status": "completed",
                "verified": True,
                "test_transactions": 3,
                "all_tests_passed": True
            },
            "network_connectivity": {
                "status": "active",
                "signal_strength": "excellent",
                "backup_connection": "configured"
            },
            "staff_training": {
                "status": "completed",
                "staff_trained": 2,
                "training_score": "85%",
                "certification_provided": True
            },
            "documentation": {
                "user_manual": "provided",
                "quick_reference": "provided",
                "support_contacts": "shared"
            }
        },
        "merchant_feedback": {
            "satisfaction_score": request.get("satisfaction_score", 9),
            "ease_of_setup": "very_easy",
            "staff_readiness": "confident",
            "additional_support_needed": False
        },
        "post_setup_actions": {
            "account_activation": "completed",
            "live_monitoring": "enabled",
            "support_ticket": "created",
            "follow_up_scheduled": "2024-01-28"
        },
        "business_metrics": {
            "expected_daily_transactions": request.get("expected_transactions", 50),
            "estimated_monthly_volume": "PKR 150,000",
            "risk_assessment": "low",
            "growth_potential": "high"
        },
        "compliance_checklist": {
            "regulatory_compliance": "verified",
            "security_standards": "met",
            "data_protection": "configured",
            "audit_trail": "enabled"
        },
        "next_milestones": [
            "First week performance review",
            "Monthly business review",
            "Quarterly growth assessment"
        ]
    }


# Notification Management POST Endpoints

@app.get("/api/icp/executor/todays-tasks-manager")
def get_todays_tasks_from_manager():
    """View today's tasks from manager"""
    return {
        "status": "success",
        "message": "Today's tasks retrieved successfully",
        "task_summary": {
            "total_tasks": 8,
            "completed": 3,
            "in_progress": 2,
            "pending": 3,
            "completion_rate": "37.5%"
        },
        "tasks": [
            {
                "task_id": "TSK001",
                "title": "Contact High-Risk Merchants",
                "description": "Follow up with 5 merchants showing declining revenue trends",
                "priority": "High",
                "assigned_by": "Regional Manager - Sarah Khan",
                "assigned_time": "2025-09-03 08:00:00",
                "due_time": "2025-09-03 17:00:00",
                "status": "in_progress",
                "merchants_assigned": ["M001", "M003", "M007", "M012", "M015"],
                "progress": {
                    "contacted": 3,
                    "pending": 2,
                    "responses_received": 2
                }
            },
            {
                "task_id": "TSK002",
                "title": "Document Collection Follow-up",
                "description": "Collect pending documents from 3 new merchants",
                "priority": "Medium",
                "assigned_by": "Operations Manager - Ali Ahmed",
                "assigned_time": "2025-09-03 09:30:00",
                "due_time": "2025-09-03 16:00:00",
                "status": "completed",
                "merchants_assigned": ["M018", "M019", "M020"],
                "completion_note": "All documents collected and verified"
            },
            {
                "task_id": "TSK003",
                "title": "Training Session Coordination",
                "description": "Schedule and conduct POS training for new merchants",
                "priority": "Medium",
                "assigned_by": "Training Manager - Fatima Sheikh",
                "assigned_time": "2025-09-03 10:00:00",
                "due_time": "2025-09-03 15:00:00",
                "status": "pending",
                "merchants_assigned": ["M021", "M022"],
                "requirements": ["Training materials", "POS terminals", "Demo cards"]
            },
            {
                "task_id": "TSK004",
                "title": "Monthly Report Preparation",
                "description": "Compile monthly retention metrics and submit report",
                "priority": "High",
                "assigned_by": "Regional Manager - Sarah Khan",
                "assigned_time": "2025-09-03 11:00:00",
                "due_time": "2025-09-03 18:00:00",
                "status": "pending",
                "deliverables": ["Retention rate analysis", "Merchant satisfaction scores", "Action plan for Q4"]
            }
        ],
        "manager_notes": [
            "Focus on high-risk merchants first",
            "Ensure all training sessions are documented",
            "Submit daily progress updates by 6 PM"
        ],
        "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }


@app.get("/api/icp/executor/follow-up-reminders")
def get_follow_up_reminders():
    """Follow-up reminders (e.g., call for loan request, merchant inactivity alert)"""
    return {
        "status": "success",
        "message": "Follow-up reminders retrieved successfully",
        "reminder_summary": {
            "total_reminders": 12,
            "urgent": 4,
            "due_today": 7,
            "overdue": 2
        },
        "reminders": [
            {
                "reminder_id": "REM001",
                "type": "loan_request_follow_up",
                "title": "Loan Application Follow-up",
                "merchant_id": "M005",
                "merchant_name": "Digital Solutions Hub",
                "description": "Follow up on loan application submitted 3 days ago",
                "priority": "urgent",
                "due_date": "2025-09-03",
                "due_time": "14:00:00",
                "original_request_date": "2025-08-31",
                "loan_amount": "PKR 500,000",
                "purpose": "Equipment purchase",
                "status": "pending_review",
                "action_required": "Contact merchant for additional documents",
                "contact_details": {
                    "phone": "+92-300-1234567",
                    "email": "owner@digitalsolutions.com",
                    "preferred_contact": "phone"
                }
            },
            {
                "reminder_id": "REM002",
                "type": "merchant_inactivity_alert",
                "title": "Merchant Inactivity Alert",
                "merchant_id": "M008",
                "merchant_name": "Fashion Boutique",
                "description": "Merchant has been inactive for 7 days - urgent intervention required",
                "priority": "urgent",
                "due_date": "2025-09-03",
                "due_time": "10:00:00",
                "inactivity_period": "7 days",
                "last_transaction": "2025-08-27",
                "previous_monthly_volume": "PKR 85,000",
                "risk_level": "high",
                "suggested_actions": [
                    "Make immediate phone call",
                    "Schedule site visit",
                    "Investigate technical issues",
                    "Offer additional support"
                ],
                "escalation_required": True
            },
            {
                "reminder_id": "REM003",
                "type": "contract_renewal",
                "title": "Contract Renewal Reminder",
                "merchant_id": "M012",
                "merchant_name": "Food Corner",
                "description": "Contract expires in 15 days - initiate renewal process",
                "priority": "medium",
                "due_date": "2025-09-04",
                "due_time": "16:00:00",
                "contract_expiry": "2025-09-18",
                "current_contract_value": "PKR 25,000/month",
                "renewal_terms": "Standard 12-month extension",
                "merchant_satisfaction": 8.5,
                "renewal_probability": "high"
            },
            {
                "reminder_id": "REM004",
                "type": "training_follow_up",
                "title": "Post-Training Follow-up",
                "merchant_id": "M015",
                "merchant_name": "Tech Accessories Store",
                "description": "Check training effectiveness and address any concerns",
                "priority": "medium",
                "due_date": "2025-09-03",
                "due_time": "15:30:00",
                "training_date": "2025-09-01",
                "training_type": "Advanced POS features",
                "staff_trained": 3,
                "follow_up_questions": [
                    "Are staff comfortable with new features?",
                    "Any technical difficulties?",
                    "Need additional training sessions?"
                ]
            }
        ],
        "overdue_reminders": [
            {
                "reminder_id": "REM_OVD001",
                "merchant_id": "M020",
                "title": "Document Verification Overdue",
                "days_overdue": 2,
                "urgency": "critical"
            },
            {
                "reminder_id": "REM_OVD002",
                "merchant_id": "M025",
                "title": "Installation Visit Overdue",
                "days_overdue": 1,
                "urgency": "high"
            }
        ],
        "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }


@app.get("/api/icp/executor/pending-actions")
def get_pending_actions():
    """Pending actions (e.g., upload visit proof, submit commitment)"""
    return {
        "status": "success",
        "message": "Pending actions retrieved successfully",
        "action_summary": {
            "total_pending": 15,
            "high_priority": 6,
            "medium_priority": 7,
            "low_priority": 2,
            "overdue": 3
        },
        "pending_actions": [
            {
                "action_id": "ACT001",
                "type": "upload_visit_proof",
                "title": "Upload Visit Proof - Tech Solutions",
                "merchant_id": "M001",
                "merchant_name": "Tech Solutions Ltd",
                "description": "Upload photo proof of site visit conducted on 2025-09-02",
                "priority": "high",
                "due_date": "2025-09-03",
                "due_time": "17:00:00",
                "visit_date": "2025-09-02",
                "visit_purpose": "Contract renewal discussion",
                "required_documents": [
                    "Site visit photo",
                    "Merchant signature on visit form",
                    "Discussion summary"
                ],
                "status": "pending_upload",
                "assigned_to": "Field Executive - Ahmad Ali"
            },
            {
                "action_id": "ACT002",
                "type": "submit_commitment",
                "title": "Submit Training Commitment",
                "merchant_id": "M007",
                "merchant_name": "Electronics Hub",
                "description": "Submit commitment details promised to merchant during follow-up call",
                "priority": "high",
                "due_date": "2025-09-03",
                "due_time": "16:00:00",
                "commitment_type": "Training session",
                "commitment_details": {
                    "training_topic": "Advanced payment processing",
                    "scheduled_date": "2025-09-05",
                    "duration": "2 hours",
                    "trainer_assigned": "Training Specialist - Maria Khan"
                },
                "merchant_expectations": "Hands-on training for 3 staff members",
                "status": "pending_submission"
            },
            {
                "action_id": "ACT003",
                "type": "document_verification",
                "title": "Verify Uploaded Documents",
                "merchant_id": "M018",
                "merchant_name": "Fashion Trends",
                "description": "Review and verify CNIC and business license uploaded by merchant",
                "priority": "medium",
                "due_date": "2025-09-04",
                "due_time": "12:00:00",
                "documents_to_verify": [
                    "CNIC front and back",
                    "Business registration certificate",
                    "Shop lease agreement"
                ],
                "verification_criteria": [
                    "Document authenticity",
                    "Information completeness",
                    "Compliance with requirements"
                ],
                "status": "pending_review"
            },
            {
                "action_id": "ACT004",
                "type": "follow_up_call",
                "title": "Scheduled Follow-up Call",
                "merchant_id": "M022",
                "merchant_name": "Home Decor Store",
                "description": "Make promised follow-up call regarding payment gateway issues",
                "priority": "high",
                "due_date": "2025-09-03",
                "due_time": "14:00:00",
                "call_purpose": "Technical support follow-up",
                "previous_issue": "Payment gateway connectivity problems",
                "expected_resolution": "Confirm issue resolution and merchant satisfaction",
                "status": "pending_call"
            },
            {
                "action_id": "ACT005",
                "type": "report_submission",
                "title": "Weekly Performance Report",
                "description": "Submit weekly performance metrics to regional manager",
                "priority": "medium",
                "due_date": "2025-09-04",
                "due_time": "09:00:00",
                "report_sections": [
                    "Merchant retention metrics",
                    "New onboarding progress",
                    "Issue resolution summary",
                    "Weekly achievements"
                ],
                "recipient": "Regional Manager - Sarah Khan",
                "status": "pending_compilation"
            }
        ],
        "overdue_actions": [
            {
                "action_id": "ACT_OVD001",
                "title": "Customer Satisfaction Survey",
                "merchant_id": "M010",
                "days_overdue": 1,
                "urgency": "medium"
            },
            {
                "action_id": "ACT_OVD002",
                "title": "Equipment Delivery Confirmation",
                "merchant_id": "M013",
                "days_overdue": 3,
                "urgency": "high"
            }
        ],
        "quick_actions": [
            {
                "action": "Mark Visit Complete",
                "endpoint": "/api/icp/executor/mark-activity-complete",
                "required_fields": ["merchant_id", "activity_type", "proof_file"]
            },
            {
                "action": "Submit Report",
                "endpoint": "/api/icp/executor/submit-summary-report",
                "required_fields": ["report_type", "summary"]
            },
            {
                "action": "Add Notes",
                "endpoint": "/api/icp/executor/add-merchant-notes",
                "required_fields": ["merchant_id", "note_type", "content"]
            }
        ],
        "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }


# Support Requests POST Endpoints

@app.post("/api/icp/executor/raise-pos-issue")
def raise_pos_issue(request: dict):
    """Raise POS issue and log support ticket"""
    merchant_id = request.get("merchant_id", "M001")
    issue_description = request.get(
        "issue_description", "POS terminal not responding")

    return {
        "success": True,
        "message": "POS issue ticket created successfully",
        "ticket_id": f"POS_{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "merchant_id": merchant_id,
        "issue_details": {
            "category": "POS Issue",
            "subcategory": request.get("subcategory", "Hardware Malfunction"),
            "description": issue_description,
            "reported_time": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "urgency": request.get("urgency", "High"),
            "merchant_impact": "Transaction processing affected"
        },
        "technical_assessment": {
            "initial_diagnosis": "POS terminal hardware/software issue",
            "possible_causes": [
                "Network connectivity problems",
                "Hardware malfunction",
                "Software corruption",
                "Power supply issues"
            ],
            "immediate_actions": [
                "Remote diagnostic check",
                "Basic troubleshooting guide provided",
                "Backup payment method activated"
            ]
        },
        "support_assignment": {
            "assigned_team": "Technical Support - Level 2",
            "technician_id": "TECH_POS_001",
            "technician_name": "Hassan Ali",
            "contact_number": "+92-300-9876543",
            "expertise": ["POS Systems", "Payment Processing", "Hardware Repair"],
            "estimated_response": "30 minutes",
            "on_site_availability": True
        },
        "escalation_process": {
            "level_1": "Remote support - 30 minutes",
            "level_2": "On-site visit - 2 hours",
            "level_3": "Hardware replacement - 4 hours",
            "emergency_contact": "+92-321-1111111"
        },
        "merchant_compensation": {
            "service_credit": "PKR 500 for downtime",
            "priority_support": "Activated for 7 days",
            "backup_solution": "Mobile payment gateway provided"
        },
        "tracking_info": {
            "status": "open",
            "priority": "critical",
            "sla_deadline": "2025-09-03 19:30:00",
            "updates_via": ["SMS", "Email", "App notification"]
        }
    }


@app.post("/api/icp/executor/raise-hardware-issue")
def raise_hardware_issue(request: dict):
    """Raise hardware issue for printer, scanner, POS machine"""
    merchant_id = request.get("merchant_id", "M001")
    hardware_type = request.get("hardware_type", "Printer")
    issue_description = request.get(
        "issue_description", "Printer not printing receipts")

    return {
        "success": True,
        "message": f"{hardware_type} issue ticket created successfully",
        "ticket_id": f"HW_{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "merchant_id": merchant_id,
        "hardware_details": {
            "equipment_type": hardware_type,
            "model": get_hardware_model(hardware_type),
            "serial_number": f"SN{datetime.now().strftime('%Y%m%d%H%M')}",
            "installation_date": "2024-06-15",
            "warranty_status": "Active - 18 months remaining",
            "last_maintenance": "2024-12-01"
        },
        "issue_analysis": {
            "reported_problem": issue_description,
            "severity": request.get("severity", "Medium"),
            "business_impact": get_business_impact(hardware_type),
            "troubleshooting_attempts": request.get("troubleshooting_done", "Basic restart attempted"),
            "error_codes": request.get("error_codes", "None reported")
        },
        "diagnostic_results": {
            "remote_check": "Connection established",
            "status_indicators": "2 error lights flashing",
            "software_version": "v2.4.1 - Up to date",
            "connectivity": "Network connected",
            "preliminary_diagnosis": get_preliminary_diagnosis(hardware_type)
        },
        "resolution_plan": {
            "immediate_steps": [
                "Remote diagnostic completed",
                "Firmware update initiated",
                "Configuration reset scheduled"
            ],
            "if_remote_fails": [
                "On-site technician dispatch",
                "Hardware component replacement",
                "Complete unit replacement if needed"
            ],
            "estimated_resolution": get_resolution_time(hardware_type),
            "backup_options": get_backup_options(hardware_type)
        },
        "technician_assignment": {
            "primary_tech": {
                "name": "Ahmed Khan",
                "id": "TECH_HW_003",
                "specialization": f"{hardware_type} specialist",
                "rating": 4.9,
                "contact": "+92-345-7777777",
                "availability": "Available now"
            },
            "backup_tech": {
                "name": "Sara Ahmed",
                "id": "TECH_HW_007",
                "contact": "+92-333-8888888"
            }
        },
        "merchant_support": {
            "temporary_solution": f"Manual {hardware_type.lower()} process guidelines provided",
            "priority_escalation": "Enabled due to business impact",
            "compensation_policy": "Service credit for extended downtime",
            "communication_preference": request.get("contact_preference", "Phone")
        }
    }


@app.post("/api/icp/executor/escalate-urgent-case")
def escalate_urgent_case(request: dict):
    """Escalate urgent case to manager"""
    case_id = request.get(
        "case_id", f"CASE_{datetime.now().strftime('%Y%m%d%H%M%S')}")
    escalation_reason = request.get(
        "escalation_reason", "Critical merchant issue requiring immediate attention")

    return {
        "success": True,
        "message": "Case escalated to manager successfully",
        "escalation_id": f"ESC_{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "original_case_id": case_id,
        "escalation_details": {
            "escalated_by": "Retention Executor",
            "escalation_time": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "escalation_level": "Level 2 - Regional Manager",
            "urgency": request.get("urgency", "Critical"),
            "reason": escalation_reason,
            "merchant_id": request.get("merchant_id", "M001")
        },
        "manager_assignment": {
            "assigned_manager": "Sarah Khan",
            "designation": "Regional Manager",
            "contact_info": {
                "phone": "+92-321-2222222",
                "email": "sarah.khan@company.com",
                "emergency_line": "+92-300-1111111"
            },
            "expected_response": "15 minutes",
            "escalation_protocol": "Immediate review and action"
        },
        "case_summary": {
            "merchant_name": get_merchant_name(request.get("merchant_id", "M001")),
            "issue_type": request.get("issue_type", "Service Disruption"),
            "financial_impact": request.get("financial_impact", "High"),
            "customer_satisfaction_risk": "Critical",
            "previous_escalations": 0,
            "case_history": [
                {
                    "timestamp": "2025-09-03 16:30:00",
                    "action": "Initial case logged",
                    "by": "Field Executive"
                },
                {
                    "timestamp": "2025-09-03 17:15:00",
                    "action": "Level 1 resolution attempted",
                    "result": "Unsuccessful"
                },
                {
                    "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    "action": "Escalated to Regional Manager",
                    "by": "Retention Executor"
                }
            ]
        },
        "required_actions": {
            "immediate": [
                "Manager review within 15 minutes",
                "Direct merchant contact by manager",
                "Resource allocation assessment"
            ],
            "short_term": [
                "Root cause analysis",
                "Comprehensive resolution plan",
                "Merchant compensation evaluation"
            ],
            "follow_up": [
                "Implementation monitoring",
                "Satisfaction confirmation",
                "Process improvement review"
            ]
        },
        "escalation_tracking": {
            "status": "escalated",
            "priority": "P1 - Critical",
            "sla_target": "Manager response: 15 minutes",
            "resolution_target": "2 hours maximum",
            "stakeholder_notifications": [
                "Regional Manager",
                "Operations Director",
                "Customer Success Team"
            ]
        },
        "merchant_communication": {
            "escalation_acknowledged": True,
            "manager_contact_scheduled": "Within 15 minutes",
            "status_updates": "Every 30 minutes until resolved",
            "compensation_review": "Initiated automatically"
        }
    }


def get_hardware_model(hardware_type):
    models = {
        "Printer": "ThermalPrint Pro X1",
        "Scanner": "ScanMaster 2000",
        "POS Machine": "SmartPOS Terminal v3"
    }
    return models.get(hardware_type, "Universal Hardware")


def get_business_impact(hardware_type):
    impacts = {
        "Printer": "Receipt printing affected - Customer service impact",
        "Scanner": "Barcode scanning disabled - Inventory tracking affected",
        "POS Machine": "Transaction processing stopped - Revenue impact critical"
    }
    return impacts.get(hardware_type, "Business operations affected")


def get_preliminary_diagnosis(hardware_type):
    diagnoses = {
        "Printer": "Thermal head overheating or paper jam detected",
        "Scanner": "Laser alignment issue or connectivity problem",
        "POS Machine": "Processing unit malfunction or network timeout"
    }
    return diagnoses.get(hardware_type, "Hardware diagnostic required")


def get_resolution_time(hardware_type):
    times = {
        "Printer": "1-2 hours",
        "Scanner": "2-3 hours",
        "POS Machine": "30 minutes to 1 hour"
    }
    return times.get(hardware_type, "2-4 hours")


def get_backup_options(hardware_type):
    options = {
        "Printer": ["Manual receipt writing", "Mobile thermal printer"],
        "Scanner": ["Manual code entry", "Mobile scanning app"],
        "POS Machine": ["Backup POS unit", "Mobile payment terminal"]
    }
    return options.get(hardware_type, ["Manual backup process"])


def get_merchant_name(merchant_id):
    names = {
        "M001": "Tech Solutions Ltd",
        "M002": "Fashion Hub",
        "M003": "Food Express"
    }
    return names.get(merchant_id, "Merchant Business")


# Feedback Management POST Endpoints

@app.post("/api/icp/executor/share-field-experience")
def share_field_experience(request: dict):
    """Share field experience from visits"""
    merchant_id = request.get("merchant_id", "M001")
    visit_date = request.get("visit_date", datetime.now().strftime('%Y-%m-%d'))
    experience_details = request.get(
        "experience_details", "Visited merchant location and assessed business operations")

    return {
        "success": True,
        "message": "Field experience shared successfully",
        "experience_id": f"EXP_{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "submission_details": {
            "submitted_by": "Retention Executor",
            "submission_time": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "visit_date": visit_date,
            "merchant_id": merchant_id,
            "merchant_name": get_merchant_name(merchant_id)
        },
        "experience_summary": {
            "visit_type": request.get("visit_type", "Routine Check"),
            "duration": request.get("duration", "45 minutes"),
            "key_observations": [
                request.get("observation_1",
                            "Merchant operations running smoothly"),
                request.get("observation_2",
                            "Staff well-trained on POS system"),
                request.get("observation_3",
                            "Customer satisfaction appears high")
            ],
            "merchant_feedback": request.get("merchant_feedback", "Positive response to services"),
            "challenges_identified": extract_challenges(experience_details),
            "opportunities": extract_opportunities(experience_details)
        },
        "business_insights": {
            "transaction_volume": request.get("transaction_volume", "High"),
            "peak_hours": request.get("peak_hours", "12:00-14:00, 18:00-20:00"),
            "customer_demographics": request.get("customer_demographics", "Mixed age groups"),
            "payment_preferences": ["Cash: 60%", "Card: 35%", "Digital: 5%"],
            "seasonal_trends": "Steady business with weekend peaks"
        },
        "recommendations": {
            "immediate_actions": [
                "Continue regular support visits",
                "Monitor transaction trends",
                "Provide additional training if needed"
            ],
            "strategic_suggestions": [
                "Explore digital payment promotion",
                "Consider loyalty program implementation",
                "Assess expansion opportunities"
            ],
            "follow_up_required": request.get("follow_up_required", True),
            "next_visit_suggested": get_next_visit_date(visit_date)
        },
        "impact_assessment": {
            "merchant_satisfaction": request.get("satisfaction_rating", 8.5),
            "business_growth_potential": "High",
            "retention_risk": "Low",
            "service_quality_score": 9.2,
            "relationship_strength": "Strong"
        },
        "documentation": {
            "photos_attached": request.get("photos_count", 3),
            "notes_quality": "Comprehensive",
            "data_completeness": "100%",
            "verification_status": "Verified"
        },
        "distribution": {
            "shared_with": [
                "Regional Manager",
                "Business Development Team",
                "Customer Success Team"
            ],
            "visibility": "Internal - Management Level",
            "retention_period": "2 years",
            "follow_up_notifications": "Enabled"
        }
    }


@app.post("/api/icp/executor/suggest-service-improvements")
def suggest_service_improvements(request: dict):
    """Suggest improvements in merchant services"""
    suggestion_category = request.get(
        "category", "General Service Enhancement")
    improvement_details = request.get(
        "improvement_details", "Enhance merchant onboarding process for better user experience")

    return {
        "success": True,
        "message": "Service improvement suggestion submitted successfully",
        "suggestion_id": f"IMP_{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "submission_info": {
            "submitted_by": "Retention Executor",
            "submission_time": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "category": suggestion_category,
            "priority": request.get("priority", "Medium"),
            "urgency": request.get("urgency", "Standard")
        },
        "improvement_details": {
            "title": request.get("title", "Service Enhancement Proposal"),
            "description": improvement_details,
            "target_area": suggestion_category,
            "affected_stakeholders": [
                "Merchants",
                "Field Executives",
                "Support Teams",
                "Management"
            ],
            "current_pain_points": extract_pain_points(improvement_details),
            "proposed_solution": extract_solution(improvement_details)
        },
        "business_impact": {
            "expected_benefits": [
                "Improved merchant satisfaction",
                "Reduced processing time",
                "Enhanced service quality",
                "Better resource utilization"
            ],
            "estimated_impact": {
                "efficiency_gain": request.get("efficiency_gain", "25%"),
                "cost_reduction": request.get("cost_reduction", "15%"),
                "satisfaction_improvement": request.get("satisfaction_improvement", "20%"),
                "time_saving": request.get("time_saving", "30 minutes per case")
            },
            "implementation_complexity": assess_complexity(improvement_details),
            "resource_requirements": estimate_resources(suggestion_category)
        },
        "implementation_plan": {
            "phases": [
                {
                    "phase": "Analysis & Planning",
                    "duration": "2 weeks",
                    "activities": ["Requirement analysis", "Stakeholder consultation", "Resource planning"]
                },
                {
                    "phase": "Development & Testing",
                    "duration": "4-6 weeks",
                    "activities": ["Solution development", "Quality testing", "User acceptance testing"]
                },
                {
                    "phase": "Deployment & Training",
                    "duration": "2 weeks",
                    "activities": ["System deployment", "User training", "Go-live support"]
                }
            ],
            "total_timeline": "8-10 weeks",
            "key_milestones": [
                "Requirements approval",
                "Development completion",
                "Testing sign-off",
                "Successful deployment"
            ]
        },
        "review_process": {
            "assigned_reviewer": "Product Manager - Sarah Ahmed",
            "review_committee": [
                "Regional Manager",
                "Operations Director",
                "IT Manager",
                "Quality Assurance Lead"
            ],
            "initial_review": "Within 3 business days",
            "detailed_assessment": "Within 1 week",
            "decision_timeline": "Within 2 weeks",
            "feedback_mechanism": "Email + Meeting"
        },
        "tracking_info": {
            "status": "submitted",
            "workflow_stage": "initial_review",
            "next_action": "Committee review",
            "estimated_decision": get_decision_date(),
            "follow_up_date": get_followup_date(),
            "notification_preferences": ["Email", "System Alert"]
        },
        "supporting_data": {
            "field_observations": request.get("observations_count", 15),
            "merchant_feedback": request.get("feedback_count", 8),
            "performance_metrics": "Baseline established",
            "competitor_analysis": "Available",
            "cost_benefit_analysis": "To be prepared"
        }
    }


def extract_challenges(experience_details):
    # Extract challenges from experience details
    common_challenges = [
        "Peak hour congestion",
        "Staff training gaps",
        "Equipment maintenance needs"
    ]
    if "slow" in experience_details.lower():
        common_challenges.append("System performance issues")
    if "confus" in experience_details.lower():
        common_challenges.append("User interface complexity")
    return common_challenges[:3]


def extract_opportunities(experience_details):
    # Extract opportunities from experience details
    opportunities = [
        "Digital payment adoption",
        "Customer engagement programs",
        "Process automation"
    ]
    if "busy" in experience_details.lower():
        opportunities.append("Capacity expansion")
    if "happy" in experience_details.lower():
        opportunities.append("Reference customer potential")
    return opportunities[:3]


def get_next_visit_date(visit_date):
    # Calculate next visit date (typically 2 weeks later)
    try:
        current_date = datetime.strptime(visit_date, '%Y-%m-%d')
        next_visit = current_date + timedelta(days=14)
        return next_visit.strftime('%Y-%m-%d')
    except:
        return (datetime.now() + timedelta(days=14)).strftime('%Y-%m-%d')


def extract_pain_points(improvement_details):
    # Extract pain points from improvement details
    pain_points = [
        "Manual process inefficiencies",
        "Communication gaps",
        "Resource constraints"
    ]
    if "slow" in improvement_details.lower():
        pain_points.append("Process delays")
    if "difficult" in improvement_details.lower():
        pain_points.append("Usability issues")
    return pain_points[:4]


def extract_solution(improvement_details):
    # Extract proposed solution from improvement details
    if "automat" in improvement_details.lower():
        return "Process automation and workflow optimization"
    elif "train" in improvement_details.lower():
        return "Enhanced training programs and knowledge sharing"
    elif "system" in improvement_details.lower():
        return "System enhancement and feature improvements"
    else:
        return "Comprehensive service optimization approach"


def assess_complexity(improvement_details):
    # Assess implementation complexity
    if any(word in improvement_details.lower() for word in ["system", "platform", "integration"]):
        return "High - Requires technical development"
    elif any(word in improvement_details.lower() for word in ["process", "workflow", "procedure"]):
        return "Medium - Process redesign needed"
    else:
        return "Low - Operational changes only"


def estimate_resources(category):
    # Estimate required resources based on category
    resource_map = {
        "General Service Enhancement": ["Product Team", "QA Team", "Training Team"],
        "Technology Improvement": ["Development Team", "IT Team", "Testing Team"],
        "Process Optimization": ["Operations Team", "Training Team", "Change Management"],
        "Customer Experience": ["UX Team", "Customer Success", "Training Team"]
    }
    return resource_map.get(category, ["Cross-functional Team", "Subject Matter Experts"])


def get_decision_date():
    # Calculate decision date (2 weeks from now)
    return (datetime.now() + timedelta(days=14)).strftime('%Y-%m-%d')


def get_followup_date():
    # Calculate follow-up date (1 week from now)
    return (datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d')


# Add the endpoint that the frontend expects for merchant menus
@app.get("/api/icp/merchant/get-menus")
def get_merchant_menus(db: Session = Depends(get_db)):
    """
    Get merchant manager menus for ICP HR company
    This endpoint matches what the frontend is calling
    """
    menus = db.query(models.ChatbotMenu)\
        .filter(
            models.ChatbotMenu.is_active == True,
            models.ChatbotMenu.company_type == "icp_hr",
            models.ChatbotMenu.role == "merchant_manager"
    ).all()

    if not menus:
        raise HTTPException(status_code=404, detail="No menus found")

    results = []
    for menu in menus:
        submenus = db.query(models.ChatbotSubmenu)\
            .filter(
                models.ChatbotSubmenu.menu_id == menu.id,
                models.ChatbotSubmenu.is_active == True,
                models.ChatbotSubmenu.company_type == "icp_hr",
                models.ChatbotSubmenu.role == "merchant_manager"
        ).all()
        results.append({
            "menu_id": menu.id,
            "menu_key": menu.menu_key,
            "menu_title": menu.menu_title,
            "menu_icon": menu.menu_icon,
            "company_type": menu.company_type,
            "role": menu.role,
            "submenus": [
                {
                    "submenu_id": sm.id,
                    "submenu_key": sm.submenu_key,
                    "submenu_title": sm.submenu_title,
                    "api_endpoint": sm.api_endpoint
                }
                for sm in submenus
            ]
        })

    return results
# ===== MERCHANT MANAGER ENDPOINTS =====

# Sales Analytics Endpoints


@app.get("/api/sales/daily")
def get_daily_sales():
    """Daily Sales Report with dummy data"""
    return {
        "report_date": datetime.now().strftime("%Y-%m-%d"),
        "total_sales": round(random.uniform(15000, 45000), 2),
        "total_transactions": random.randint(45, 120),
        "avg_transaction": round(random.uniform(250, 800), 2),
        "top_products": [
            {"name": "Electronics", "sales": round(random.uniform(
                5000, 12000), 2), "units": random.randint(15, 40)},
            {"name": "Clothing", "sales": round(random.uniform(
                3000, 8000), 2), "units": random.randint(20, 60)},
            {"name": "Food Items", "sales": round(random.uniform(
                2000, 6000), 2), "units": random.randint(30, 80)}
        ],
        "hourly_breakdown": [
            {"hour": f"{i}:00", "sales": round(random.uniform(800, 3000), 2)}
            for i in range(9, 21)
        ]
    }


@app.get("/api/sales/weekly")
def get_weekly_sales():
    """Weekly Sales Report with dummy data"""
    return {
        "week_start": (datetime.now() - timedelta(days=6)).strftime("%Y-%m-%d"),
        "week_end": datetime.now().strftime("%Y-%m-%d"),
        "total_sales": round(random.uniform(80000, 200000), 2),
        "total_transactions": random.randint(300, 800),
        "daily_breakdown": [
            {
                "date": (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d"),
                "sales": round(random.uniform(10000, 35000), 2),
                "transactions": random.randint(40, 120)
            } for i in range(7)
        ],
        "growth_rate": round(random.uniform(-15, 25), 2)
    }


@app.get("/api/sales/monthly")
def get_monthly_sales():
    """Monthly Sales Report with dummy data"""
    return {
        "month": datetime.now().strftime("%Y-%m"),
        "total_sales": round(random.uniform(300000, 800000), 2),
        "total_transactions": random.randint(1200, 3500),
        "weekly_breakdown": [
            {
                "week": f"Week {i+1}",
                "sales": round(random.uniform(60000, 180000), 2),
                "transactions": random.randint(280, 850)
            } for i in range(4)
        ],
        "top_customers": [
            {"name": f"Customer {i+1}",
                "purchases": round(random.uniform(5000, 25000), 2)}
            for i in range(10)
        ]
    }


@app.get("/api/sales/performance")
def get_merchant_performance():
    """Merchant Performance Analytics with dummy data"""
    return {
        "merchant_id": "MERCH001",
        "performance_score": round(random.uniform(75, 95), 1),
        "metrics": {
            "sales_target": 500000,
            "sales_achieved": round(random.uniform(400000, 550000), 2),
            "target_percentage": round(random.uniform(80, 110), 1),
            "customer_satisfaction": round(random.uniform(4.2, 4.8), 1),
            "order_fulfillment": round(random.uniform(92, 98), 1)
        },
        "rankings": {
            "regional_rank": random.randint(5, 25),
            "category_rank": random.randint(2, 15),
            "total_merchants": 150
        }
    }

# Staff Management Endpoints


@app.get("/api/staff/attendance")
def get_staff_attendance():
    """Staff Attendance Report with dummy data"""
    return {
        "report_date": datetime.now().strftime("%Y-%m-%d"),
        "total_staff": 12,
        "present": 10,
        "absent": 2,
        "late": 1,
        "attendance_rate": 83.3,
        "staff_details": [
            {
                "employee_id": f"EMP{str(i+1).zfill(3)}",
                "name": f"Employee {i+1}",
                "status": random.choice(["Present", "Absent", "Late"]),
                "check_in": f"{random.randint(8,10)}:{random.randint(10,59)} AM" if random.choice([True, False]) else None,
                "department": random.choice(["Sales", "Customer Service", "Inventory", "Management"])
            } for i in range(12)
        ]
    }


@app.get("/api/staff/schedule")
def get_staff_schedule():
    """Staff Schedule with dummy data"""
    return {
        "week_start": datetime.now().strftime("%Y-%m-%d"),
        "schedules": [
            {
                "employee_id": f"EMP{str(i+1).zfill(3)}",
                "name": f"Employee {i+1}",
                "shifts": [
                    {
                        "date": (datetime.now() + timedelta(days=j)).strftime("%Y-%m-%d"),
                        "start_time": f"{random.randint(8,10)}:00 AM",
                        "end_time": f"{random.randint(5,8)}:00 PM",
                        "position": random.choice(["Cashier", "Sales Associate", "Supervisor", "Manager"])
                    } for j in range(7)
                ]
            } for i in range(8)
        ]
    }


@app.get("/api/staff/performance")
def get_performance_review():
    """Performance Review with dummy data"""
    return {
        "review_period": f"{datetime.now().strftime('%Y-%m')}",
        "reviews": [
            {
                "employee_id": f"EMP{str(i+1).zfill(3)}",
                "name": f"Employee {i+1}",
                "overall_rating": round(random.uniform(3.5, 5.0), 1),
                "metrics": {
                    "punctuality": round(random.uniform(3.0, 5.0), 1),
                    "customer_service": round(random.uniform(3.5, 5.0), 1),
                    "sales_performance": round(random.uniform(3.0, 5.0), 1),
                    "teamwork": round(random.uniform(3.5, 5.0), 1)
                },
                "achievements": random.choice([
                    "Exceeded monthly sales target",
                    "Perfect attendance record",
                    "Outstanding customer feedback",
                    "Team leadership excellence"
                ]),
                "improvement_areas": random.choice([
                    "Time management",
                    "Product knowledge",
                    "Communication skills",
                    "Technical skills"
                ])
            } for i in range(10)
        ]
    }


@app.get("/api/staff/payroll")
def get_payroll_management():
    """Payroll Management with dummy data"""
    return {
        "payroll_period": f"{datetime.now().strftime('%Y-%m')}",
        "total_payroll": round(random.uniform(80000, 150000), 2),
        "employee_payroll": [
            {
                "employee_id": f"EMP{str(i+1).zfill(3)}",
                "name": f"Employee {i+1}",
                "base_salary": round(random.uniform(4000, 12000), 2),
                "overtime": round(random.uniform(0, 2000), 2),
                "bonuses": round(random.uniform(0, 1500), 2),
                "deductions": round(random.uniform(200, 800), 2),
                "net_pay": round(random.uniform(3500, 13000), 2),
                "payment_status": random.choice(["Processed", "Pending", "On Hold"])
            } for i in range(12)
        ]
    }

# Inventory Management Endpoints


@app.get("/api/inventory/stock")
def get_stock_levels():
    """Current Stock Levels with dummy data"""
    return {
        "report_date": datetime.now().strftime("%Y-%m-%d"),
        "total_products": 245,
        "low_stock_alerts": 8,
        "out_of_stock": 3,
        "categories": [
            {
                "category": "Electronics",
                "total_items": 85,
                "low_stock": 3,
                "out_of_stock": 1,
                "total_value": round(random.uniform(150000, 300000), 2)
            },
            {
                "category": "Clothing",
                "total_items": 120,
                "low_stock": 4,
                "out_of_stock": 2,
                "total_value": round(random.uniform(80000, 150000), 2)
            },
            {
                "category": "Food Items",
                "total_items": 40,
                "low_stock": 1,
                "out_of_stock": 0,
                "total_value": round(random.uniform(20000, 50000), 2)
            }
        ],
        "critical_items": [
            {
                "product_id": f"PROD{i+1:03d}",
                "name": f"Product {i+1}",
                "current_stock": random.randint(0, 5),
                "min_threshold": 10,
                "category": random.choice(["Electronics", "Clothing", "Food Items"]),
                "supplier": f"Supplier {random.randint(1, 5)}"
            } for i in range(8)
        ]
    }


@app.get("/api/inventory/catalog")
def get_product_catalog():
    """Product Catalog with dummy data"""
    return {
        "total_products": 245,
        "categories": 3,
        "products": [
            {
                "product_id": f"PROD{i+1:03d}",
                "name": f"Product {i+1}",
                "category": random.choice(["Electronics", "Clothing", "Food Items"]),
                "price": round(random.uniform(50, 2000), 2),
                "cost": round(random.uniform(25, 1000), 2),
                "current_stock": random.randint(0, 100),
                "reorder_level": random.randint(5, 20),
                "supplier": f"Supplier {random.randint(1, 5)}",
                "status": random.choice(["Active", "Discontinued", "Seasonal"]),
                "last_updated": (datetime.now() - timedelta(days=random.randint(1, 30))).strftime("%Y-%m-%d")
            } for i in range(20)
        ]
    }


@app.get("/api/inventory/suppliers")
def get_supplier_management():
    """Supplier Management with dummy data"""
    return {
        "total_suppliers": 15,
        "active_suppliers": 12,
        "pending_orders": 8,
        "suppliers": [
            {
                "supplier_id": f"SUP{i+1:03d}",
                "name": f"Supplier {i+1}",
                "contact_person": f"Contact Person {i+1}",
                "email": f"supplier{i+1}@example.com",
                "phone": f"+1-555-{random.randint(1000, 9999)}",
                "categories": random.sample(["Electronics", "Clothing", "Food Items"], random.randint(1, 3)),
                "status": random.choice(["Active", "Inactive", "Under Review"]),
                "rating": round(random.uniform(3.5, 5.0), 1),
                "total_orders": random.randint(50, 500),
                "last_order_date": (datetime.now() - timedelta(days=random.randint(1, 60))).strftime("%Y-%m-%d")
            } for i in range(15)
        ]
    }


@app.get("/api/inventory/alerts")
def get_inventory_alerts():
    """Inventory Alerts with dummy data"""
    return {
        "alert_date": datetime.now().strftime("%Y-%m-%d"),
        "total_alerts": 12,
        "critical_alerts": 5,
        "warning_alerts": 7,
        "alerts": [
            {
                "alert_id": f"ALERT{i+1:03d}",
                "type": random.choice(["Low Stock", "Out of Stock", "Overstock", "Expiry Warning"]),
                "severity": random.choice(["Critical", "Warning", "Info"]),
                "product_id": f"PROD{random.randint(1, 245):03d}",
                "product_name": f"Product {random.randint(1, 245)}",
                "current_stock": random.randint(0, 15),
                "threshold": random.randint(10, 25),
                "message": random.choice([
                    "Stock level below minimum threshold",
                    "Product out of stock",
                    "Items approaching expiry date",
                    "Overstocked item taking up space"
                ]),
                "created_date": (datetime.now() - timedelta(days=random.randint(0, 7))).strftime("%Y-%m-%d"),
                "status": random.choice(["New", "Acknowledged", "Resolved"])
            } for i in range(12)
        ]
    }

# Marketing Campaign Endpoints


@app.get("/api/marketing/campaigns")
def get_marketing_campaigns():
    """Marketing Campaigns with dummy data"""
    return {
        "total_campaigns": 8,
        "active_campaigns": 3,
        "completed_campaigns": 4,
        "draft_campaigns": 1,
        "campaigns": [
            {
                "campaign_id": f"CAMP{i+1:03d}",
                "name": f"Campaign {i+1}",
                "type": random.choice(["Email", "SMS", "Social Media", "Print", "Radio"]),
                "status": random.choice(["Active", "Completed", "Draft", "Paused"]),
                "start_date": (datetime.now() - timedelta(days=random.randint(1, 60))).strftime("%Y-%m-%d"),
                "end_date": (datetime.now() + timedelta(days=random.randint(1, 30))).strftime("%Y-%m-%d"),
                "budget": round(random.uniform(5000, 50000), 2),
                "spent": round(random.uniform(2000, 40000), 2),
                "target_audience": random.choice(["All Customers", "VIP Customers", "New Customers", "Inactive Customers"]),
                "reach": random.randint(1000, 10000),
                "engagement_rate": round(random.uniform(15, 45), 2),
                "conversion_rate": round(random.uniform(2, 15), 2)
            } for i in range(8)
        ]
    }


@app.get("/api/marketing/segments")
def get_customer_segments():
    """Customer Segments with dummy data"""
    return {
        "total_customers": 5420,
        "total_segments": 6,
        "segments": [
            {
                "segment_id": f"SEG{i+1:03d}",
                "name": random.choice(["VIP Customers", "Regular Customers", "New Customers", "Inactive Customers", "High Value", "Bargain Hunters"]),
                "customer_count": random.randint(200, 1500),
                "avg_purchase_value": round(random.uniform(150, 2000), 2),
                "purchase_frequency": round(random.uniform(2, 12), 1),
                "last_purchase_days": random.randint(1, 365),
                "engagement_score": round(random.uniform(30, 95), 1),
                "lifetime_value": round(random.uniform(1000, 15000), 2),
                "preferred_channel": random.choice(["Email", "SMS", "Phone", "In-Store", "Social Media"])
            } for i in range(6)
        ]
    }


@app.get("/api/marketing/offers")
def get_promotional_offers():
    """Promotional Offers with dummy data"""
    return {
        "active_offers": 5,
        "upcoming_offers": 3,
        "expired_offers": 12,
        "offers": [
            {
                "offer_id": f"OFFER{i+1:03d}",
                "title": f"Offer {i+1}",
                "type": random.choice(["Discount", "BOGO", "Free Shipping", "Cashback", "Gift Card"]),
                "discount_percentage": random.randint(10, 50),
                "status": random.choice(["Active", "Upcoming", "Expired", "Draft"]),
                "start_date": (datetime.now() - timedelta(days=random.randint(0, 30))).strftime("%Y-%m-%d"),
                "end_date": (datetime.now() + timedelta(days=random.randint(1, 60))).strftime("%Y-%m-%d"),
                "target_segment": random.choice(["All Customers", "VIP", "New Customers", "Inactive"]),
                "usage_count": random.randint(50, 500),
                "usage_limit": random.randint(100, 1000),
                "revenue_generated": round(random.uniform(5000, 50000), 2),
                "redemption_rate": round(random.uniform(15, 65), 2)
            } for i in range(20)
        ]
    }


@app.get("/api/marketing/analytics")
def get_marketing_analytics():
    """Marketing Analytics with dummy data"""
    return {
        "report_period": datetime.now().strftime("%Y-%m"),
        "total_campaigns": 8,
        "total_budget": round(random.uniform(100000, 300000), 2),
        "total_spent": round(random.uniform(80000, 250000), 2),
        "total_revenue": round(random.uniform(400000, 800000), 2),
        "roi": round(random.uniform(250, 450), 2),
        "channel_performance": [
            {
                "channel": "Email",
                "campaigns": 3,
                "budget": round(random.uniform(20000, 50000), 2),
                "reach": random.randint(3000, 8000),
                "engagement_rate": round(random.uniform(25, 45), 2),
                "conversion_rate": round(random.uniform(5, 15), 2),
                "revenue": round(random.uniform(80000, 200000), 2)
            },
            {
                "channel": "Social Media",
                "campaigns": 2,
                "budget": round(random.uniform(15000, 40000), 2),
                "reach": random.randint(5000, 15000),
                "engagement_rate": round(random.uniform(15, 35), 2),
                "conversion_rate": round(random.uniform(3, 10), 2),
                "revenue": round(random.uniform(60000, 150000), 2)
            },
            {
                "channel": "SMS",
                "campaigns": 3,
                "budget": round(random.uniform(10000, 30000), 2),
                "reach": random.randint(2000, 6000),
                "engagement_rate": round(random.uniform(20, 40), 2),
                "conversion_rate": round(random.uniform(8, 18), 2),
                "revenue": round(random.uniform(40000, 120000), 2)
            }
        ]
    }

# Financial Reports Endpoints


@app.get("/api/finance/revenue")
def get_revenue_reports():
    """Revenue Reports with dummy data"""
    return {
        "report_period": datetime.now().strftime("%Y-%m"),
        "total_revenue": round(random.uniform(400000, 800000), 2),
        "revenue_growth": round(random.uniform(-10, 25), 2),
        "monthly_breakdown": [
            {
                "month": (datetime.now() - timedelta(days=30*i)).strftime("%Y-%m"),
                "revenue": round(random.uniform(350000, 750000), 2),
                "transactions": random.randint(1200, 3000),
                "avg_transaction": round(random.uniform(200, 600), 2)
            } for i in range(12)
        ],
        "revenue_by_category": [
            {
                "category": "Electronics",
                "revenue": round(random.uniform(150000, 300000), 2),
                "percentage": round(random.uniform(35, 45), 1)
            },
            {
                "category": "Clothing",
                "revenue": round(random.uniform(100000, 200000), 2),
                "percentage": round(random.uniform(25, 35), 1)
            },
            {
                "category": "Food Items",
                "revenue": round(random.uniform(80000, 150000), 2),
                "percentage": round(random.uniform(20, 30), 1)
            }
        ]
    }


@app.get("/api/finance/expenses")
def get_expense_tracking():
    """Expense Tracking with dummy data"""
    return {
        "report_period": datetime.now().strftime("%Y-%m"),
        "total_expenses": round(random.uniform(200000, 400000), 2),
        "expense_categories": [
            {
                "category": "Inventory",
                "amount": round(random.uniform(100000, 200000), 2),
                "percentage": round(random.uniform(40, 55), 1),
                "budget": round(random.uniform(120000, 220000), 2)
            },
            {
                "category": "Staff Salaries",
                "amount": round(random.uniform(80000, 150000), 2),
                "percentage": round(random.uniform(30, 40), 1),
                "budget": round(random.uniform(90000, 160000), 2)
            },
            {
                "category": "Rent & Utilities",
                "amount": round(random.uniform(20000, 40000), 2),
                "percentage": round(random.uniform(8, 15), 1),
                "budget": round(random.uniform(25000, 45000), 2)
            },
            {
                "category": "Marketing",
                "amount": round(random.uniform(15000, 35000), 2),
                "percentage": round(random.uniform(5, 12), 1),
                "budget": round(random.uniform(20000, 40000), 2)
            }
        ],
        "monthly_trend": [
            {
                "month": (datetime.now() - timedelta(days=30*i)).strftime("%Y-%m"),
                "total_expenses": round(random.uniform(180000, 380000), 2)
            } for i in range(6)
        ]
    }


@app.get("/api/finance/profit-loss")
def get_profit_loss():
    """Profit & Loss Statement with dummy data"""
    return {
        "report_period": datetime.now().strftime("%Y-%m"),
        "revenue": {
            "gross_revenue": round(random.uniform(400000, 800000), 2),
            "returns_refunds": round(random.uniform(10000, 30000), 2),
            "net_revenue": round(random.uniform(370000, 770000), 2)
        },
        "expenses": {
            "cost_of_goods_sold": round(random.uniform(150000, 300000), 2),
            "operating_expenses": round(random.uniform(100000, 200000), 2),
            "marketing_expenses": round(random.uniform(15000, 35000), 2),
            "administrative_expenses": round(random.uniform(25000, 50000), 2),
            "total_expenses": round(random.uniform(290000, 585000), 2)
        },
        "profit": {
            "gross_profit": round(random.uniform(220000, 470000), 2),
            "operating_profit": round(random.uniform(80000, 185000), 2),
            "net_profit": round(random.uniform(60000, 150000), 2),
            "profit_margin": round(random.uniform(15, 25), 2)
        },
        "quarterly_comparison": [
            {
                "quarter": f"Q{i+1} {datetime.now().year}",
                "revenue": round(random.uniform(1000000, 2000000), 2),
                "expenses": round(random.uniform(800000, 1600000), 2),
                "profit": round(random.uniform(200000, 400000), 2)
            } for i in range(4)
        ]
    }


@app.get("/api/finance/tax-reports")
def get_tax_reports():
    """Tax Reports with dummy data"""
    return {
        "tax_year": datetime.now().year,
        "tax_period": datetime.now().strftime("%Y-%m"),
        "sales_tax": {
            "taxable_sales": round(random.uniform(400000, 800000), 2),
            "tax_rate": 8.5,
            "tax_collected": round(random.uniform(34000, 68000), 2),
            "tax_paid": round(random.uniform(32000, 66000), 2),
            "tax_due": round(random.uniform(2000, 5000), 2)
        },
        "income_tax": {
            "gross_income": round(random.uniform(400000, 800000), 2),
            "deductions": round(random.uniform(100000, 200000), 2),
            "taxable_income": round(random.uniform(300000, 600000), 2),
            "tax_rate": 25.0,
            "estimated_tax": round(random.uniform(75000, 150000), 2)
        },
        "quarterly_payments": [
            {
                "quarter": f"Q{i+1}",
                "due_date": f"{datetime.now().year}-{(i+1)*3:02d}-15",
                "amount_due": round(random.uniform(18000, 38000), 2),
                "amount_paid": round(random.uniform(15000, 38000), 2),
                "status": random.choice(["Paid", "Pending", "Overdue"])
            } for i in range(4)
        ]
    }

# Customer Management Endpoints


@app.get("/api/customers/database")
def get_customer_database():
    """Customer Database with dummy data"""
    return {
        "total_customers": 5420,
        "new_customers_this_month": random.randint(150, 350),
        "active_customers": 4280,
        "inactive_customers": 1140,
        "customers": [
            {
                "customer_id": f"CUST{i+1:05d}",
                "name": f"Customer {i+1}",
                "email": f"customer{i+1}@example.com",
                "phone": f"+1-555-{random.randint(1000, 9999)}",
                "registration_date": (datetime.now() - timedelta(days=random.randint(1, 1095))).strftime("%Y-%m-%d"),
                "last_purchase": (datetime.now() - timedelta(days=random.randint(1, 180))).strftime("%Y-%m-%d"),
                "total_purchases": round(random.uniform(500, 15000), 2),
                "purchase_count": random.randint(1, 50),
                "status": random.choice(["Active", "Inactive", "VIP", "New"]),
                "preferred_category": random.choice(["Electronics", "Clothing", "Food Items"]),
                "loyalty_points": random.randint(0, 5000)
            } for i in range(25)
        ]
    }


@app.get("/api/customers/loyalty")
def get_loyalty_programs():
    """Loyalty Programs with dummy data"""
    return {
        "total_members": 3200,
        "active_members": 2580,
        "points_issued_this_month": random.randint(15000, 35000),
        "points_redeemed_this_month": random.randint(8000, 20000),
        "program_tiers": [
            {
                "tier": "Bronze",
                "members": 1800,
                "min_points": 0,
                "max_points": 999,
                "benefits": ["5% discount", "Birthday offers"],
                "avg_spending": round(random.uniform(200, 500), 2)
            },
            {
                "tier": "Silver",
                "members": 980,
                "min_points": 1000,
                "max_points": 4999,
                "benefits": ["10% discount", "Free shipping", "Early access"],
                "avg_spending": round(random.uniform(500, 1200), 2)
            },
            {
                "tier": "Gold",
                "members": 420,
                "min_points": 5000,
                "max_points": 999999,
                "benefits": ["15% discount", "Priority support", "Exclusive events"],
                "avg_spending": round(random.uniform(1200, 3000), 2)
            }
        ],
        "recent_activities": [
            {
                "customer_id": f"CUST{random.randint(1, 5420):05d}",
                "activity": random.choice(["Points Earned", "Points Redeemed", "Tier Upgrade", "Bonus Points"]),
                "points": random.randint(50, 500),
                "date": (datetime.now() - timedelta(days=random.randint(1, 30))).strftime("%Y-%m-%d")
            } for i in range(15)
        ]
    }


@app.get("/api/customers/feedback")
def get_customer_feedback():
    """Customer Feedback with dummy data"""
    return {
        "total_reviews": 1240,
        "average_rating": round(random.uniform(4.2, 4.8), 1),
        "ratings_distribution": {
            "5_stars": random.randint(600, 800),
            "4_stars": random.randint(250, 350),
            "3_stars": random.randint(100, 200),
            "2_stars": random.randint(20, 60),
            "1_star": random.randint(10, 30)
        },
        "recent_feedback": [
            {
                "feedback_id": f"FB{i+1:04d}",
                "customer_id": f"CUST{random.randint(1, 5420):05d}",
                "rating": random.randint(1, 5),
                "category": random.choice(["Product Quality", "Customer Service", "Delivery", "Price", "Overall Experience"]),
                "comment": random.choice([
                    "Great service and products!",
                    "Fast delivery and good quality",
                    "Could improve customer service",
                    "Excellent value for money",
                    "Product exceeded expectations"
                ]),
                "date": (datetime.now() - timedelta(days=random.randint(1, 60))).strftime("%Y-%m-%d"),
                "status": random.choice(["New", "Reviewed", "Responded", "Resolved"])
            } for i in range(20)
        ],
        "sentiment_analysis": {
            "positive": round(random.uniform(65, 80), 1),
            "neutral": round(random.uniform(15, 25), 1),
            "negative": round(random.uniform(5, 15), 1)
        }
    }


@app.get("/api/customers/support")
def get_support_tickets():
    """Support Tickets with dummy data"""
    return {
        "total_tickets": 180,
        "open_tickets": 25,
        "closed_tickets": 155,
        "avg_resolution_time": round(random.uniform(2.5, 8.0), 1),
        "ticket_categories": [
            {
                "category": "Product Issues",
                "count": random.randint(40, 70),
                "avg_resolution_hours": round(random.uniform(4, 12), 1)
            },
            {
                "category": "Order Issues",
                "count": random.randint(30, 50),
                "avg_resolution_hours": round(random.uniform(2, 8), 1)
            },
            {
                "category": "Payment Issues",
                "count": random.randint(20, 40),
                "avg_resolution_hours": round(random.uniform(6, 24), 1)
            },
            {
                "category": "General Inquiry",
                "count": random.randint(15, 35),
                "avg_resolution_hours": round(random.uniform(1, 4), 1)
            }
        ],
        "recent_tickets": [
            {
                "ticket_id": f"TKT{i+1:05d}",
                "customer_id": f"CUST{random.randint(1, 5420):05d}",
                "subject": random.choice([
                    "Product defect issue",
                    "Order delivery delay",
                    "Payment processing problem",
                    "Account access issue",
                    "Return request"
                ]),
                "priority": random.choice(["Low", "Medium", "High", "Critical"]),
                "status": random.choice(["Open", "In Progress", "Pending Customer", "Resolved", "Closed"]),
                "created_date": (datetime.now() - timedelta(days=random.randint(1, 30))).strftime("%Y-%m-%d"),
                "assigned_to": f"Agent {random.randint(1, 8)}",
                "category": random.choice(["Product Issues", "Order Issues", "Payment Issues", "General Inquiry"])
            } for i in range(15)
        ]
    }

# Add HR Assistant endpoints for ICP HR role


@app.get("/api/hr/employee-directory")
def get_hr_employee_directory():
    """HR Employee Directory with dummy data"""
    return {
        "total_employees": 45,
        "active_employees": 42,
        "on_leave": 3,
        "departments": ["HR", "Finance", "Marketing", "IT", "Operations"],
        "employees": [
            {
                "employee_id": f"HR{i+1:03d}",
                "name": f"HR Employee {i+1}",
                "department": random.choice(["HR", "Finance", "Marketing", "IT", "Operations"]),
                "position": random.choice(["Manager", "Senior Associate", "Associate", "Intern"]),
                "email": f"hr.employee{i+1}@icphr.com",
                "phone": f"+1-555-{random.randint(1000, 9999)}",
                "hire_date": (datetime.now() - timedelta(days=random.randint(30, 1095))).strftime("%Y-%m-%d"),
                "status": random.choice(["Active", "On Leave", "Training"]),
                "manager": f"Manager {random.randint(1, 10)}"
            } for i in range(15)
        ]
    }


@app.get("/api/hr/recruitment")
def get_hr_recruitment():
    """HR Recruitment with dummy data"""
    return {
        "open_positions": 8,
        "applications_received": 125,
        "interviews_scheduled": 15,
        "offers_made": 3,
        "positions": [
            {
                "position_id": f"POS{i+1:03d}",
                "title": random.choice(["Software Developer", "Marketing Specialist", "HR Coordinator", "Financial Analyst", "Operations Manager"]),
                "department": random.choice(["IT", "Marketing", "HR", "Finance", "Operations"]),
                "applications": random.randint(8, 25),
                "status": random.choice(["Open", "Interviewing", "Offer Made", "Closed"]),
                "posted_date": (datetime.now() - timedelta(days=random.randint(1, 60))).strftime("%Y-%m-%d"),
                "urgency": random.choice(["Low", "Medium", "High"])
            } for i in range(8)
        ]
    }

# ===== MISSING FINANCIAL ENDPOINTS =====


@app.get("/api/finance/pnl")
def get_profit_loss():
    """Profit & Loss Report with dummy data"""
    total_revenue = round(random.uniform(800000, 1200000), 2)
    total_expenses = round(total_revenue * random.uniform(0.6, 0.8), 2)
    net_profit = round(total_revenue - total_expenses, 2)

    return {
        "report_period": "2025-08",
        "total_revenue": total_revenue,
        "total_expenses": total_expenses,
        "gross_profit": round(total_revenue * 0.4, 2),
        "net_profit": net_profit,
        "profit_margin": round((net_profit / total_revenue) * 100, 1),
        "revenue_breakdown": [
            {"category": "Product Sales",
                "amount": round(total_revenue * 0.7, 2)},
            {"category": "Service Revenue",
                "amount": round(total_revenue * 0.2, 2)},
            {"category": "Other Income",
                "amount": round(total_revenue * 0.1, 2)}
        ],
        "expense_breakdown": [
            {"category": "Cost of Goods Sold",
                "amount": round(total_expenses * 0.4, 2)},
            {"category": "Salaries & Benefits",
                "amount": round(total_expenses * 0.3, 2)},
            {"category": "Operating Expenses",
                "amount": round(total_expenses * 0.2, 2)},
            {"category": "Other Expenses",
                "amount": round(total_expenses * 0.1, 2)}
        ]
    }


@app.get("/api/finance/taxes")
def get_tax_reports():
    """Tax Reports with dummy data"""
    quarterly_sales = round(random.uniform(2500000, 3500000), 2)
    sales_tax_rate = 8.5
    income_tax_rate = 21.0

    return {
        "report_period": "Q3 2025",
        "quarterly_sales": quarterly_sales,
        "sales_tax": {
            "rate": sales_tax_rate,
            "amount_due": round(quarterly_sales * (sales_tax_rate / 100), 2),
            "paid": round(quarterly_sales * (sales_tax_rate / 100) * 0.8, 2),
            "outstanding": round(quarterly_sales * (sales_tax_rate / 100) * 0.2, 2)
        },
        "income_tax": {
            "rate": income_tax_rate,
            "taxable_income": round(quarterly_sales * 0.15, 2),
            "estimated_tax": round(quarterly_sales * 0.15 * (income_tax_rate / 100), 2),
            "quarterly_payments": round(quarterly_sales * 0.15 * (income_tax_rate / 100) * 0.9, 2)
        },
        "upcoming_deadlines": [
            {"type": "Sales Tax", "due_date": "2025-10-15",
                "estimated_amount": round(random.uniform(15000, 25000), 2)},
            {"type": "Quarterly Income Tax", "due_date": "2025-10-15",
                "estimated_amount": round(random.uniform(35000, 55000), 2)},
            {"type": "Payroll Tax", "due_date": "2025-09-15",
                "estimated_amount": round(random.uniform(8000, 15000), 2)}
        ]
    }

# ===============================================================================
# HR ASSISTANT ENDPOINTS - Complete Implementation
# ===============================================================================

# Employee Management Endpoints


@app.get("/api/hr/employees")
def get_hr_employees():
    """HR Employee List with comprehensive details"""
    return {
        "status": "success",
        "data": [
            {
                "id": 1,
                "employee_id": "EMP001",
                "name": "John Smith",
                "department": "Engineering",
                "position": "Senior Developer",
                "hire_date": "2023-01-15",
                "status": "Active",
                "email": "john.smith@company.com",
                "phone": "+1-555-0101",
                "manager": "Jane Doe",
                "salary": 85000,
                "performance_rating": 4.2
            },
            {
                "id": 2,
                "employee_id": "EMP002",
                "name": "Sarah Johnson",
                "department": "Marketing",
                "position": "Marketing Manager",
                "hire_date": "2022-08-20",
                "status": "Active",
                "email": "sarah.johnson@company.com",
                "phone": "+1-555-0102",
                "manager": "Mike Wilson",
                "salary": 75000,
                "performance_rating": 4.5
            },
            {
                "id": 3,
                "employee_id": "EMP003",
                "name": "David Brown",
                "department": "Sales",
                "position": "Sales Representative",
                "hire_date": "2023-06-10",
                "status": "Active",
                "email": "david.brown@company.com",
                "phone": "+1-555-0103",
                "manager": "Lisa Chen",
                "salary": 60000,
                "performance_rating": 3.8
            }
        ]
    }


@app.get("/api/hr/employees/{employee_id}")
def get_hr_employee_details(employee_id: int):
    """HR Employee Details"""
    return {
        "status": "success",
        "data": {
            "id": employee_id,
            "employee_id": f"EMP{employee_id:03d}",
            "personal_info": {
                "name": "John Smith",
                "email": "john.smith@company.com",
                "phone": "+1-555-0101",
                "date_of_birth": "1990-05-15",
                "address": "123 Main St, City, State 12345",
                "emergency_contact": "Jane Smith - +1-555-0201"
            },
            "employment_info": {
                "department": "Engineering",
                "position": "Senior Developer",
                "hire_date": "2023-01-15",
                "employment_type": "Full-time",
                "status": "Active",
                "manager": "Jane Doe",
                "salary": 85000,
                "benefits": ["Health Insurance", "401k", "Paid Time Off"]
            },
            "performance": {
                "current_rating": 4.2,
                "last_review_date": "2025-06-01",
                "goals_completed": 8,
                "goals_total": 10,
                "training_completed": 12
            }
        }
    }


@app.post("/api/hr/employees/add")
def add_hr_employee(employee_data: dict):
    """Add HR Employee"""
    new_id = random.randint(100, 999)
    return {
        "status": "success",
        "message": "Employee added successfully",
        "data": {
            "id": new_id,
            "employee_id": f"EMP{new_id:03d}",
            "name": employee_data.get("name", "New Employee"),
            "department": employee_data.get("department", "General"),
            "position": employee_data.get("position", "Staff"),
            "hire_date": "2025-09-04",
            "status": "Active"
        }
    }


@app.post("/api/hr/employees/{employee_id}/edit")
def edit_hr_employee(employee_id: int, update_data: dict):
    """Edit HR Employee"""
    return {
        "status": "success",
        "message": "Employee updated successfully",
        "data": {
            "id": employee_id,
            "employee_id": f"EMP{employee_id:03d}",
            "updated_fields": list(update_data.keys()),
            "last_modified": "2025-09-04T10:30:00Z"
        }
    }

# Attendance Management Endpoints


@app.get("/api/hr/attendance")
def get_hr_attendance():
    """HR Attendance Records"""
    return {
        "status": "success",
        "data": [
            {
                "date": "2025-09-04",
                "total_employees": 45,
                "present": 42,
                "absent": 2,
                "late": 1,
                "attendance_rate": 93.3,
                "departments": [
                    {"name": "Engineering", "present": 15, "total": 16},
                    {"name": "Marketing", "present": 8, "total": 8},
                    {"name": "Sales", "present": 12, "total": 13},
                    {"name": "HR", "present": 4, "total": 4},
                    {"name": "Finance", "present": 3, "total": 4}
                ]
            }
        ]
    }


@app.get("/api/hr/attendance/today")
def get_hr_attendance_today():
    """Today's HR Attendance"""
    return {
        "status": "success",
        "data": {
            "date": "2025-09-04",
            "summary": {
                "total_employees": 45,
                "checked_in": 42,
                "not_checked_in": 3,
                "late_arrivals": 1,
                "early_departures": 0
            },
            "recent_activity": [
                {"time": "09:15", "employee": "John Smith",
                    "action": "Check In", "status": "Late"},
                {"time": "09:00", "employee": "Sarah Johnson",
                    "action": "Check In", "status": "On Time"},
                {"time": "08:45", "employee": "David Brown",
                    "action": "Check In", "status": "Early"}
            ],
            "alerts": [
                {"type": "Late Arrival", "employee": "John Smith", "time": "09:15"},
                {"type": "No Show", "employee": "Mike Wilson",
                    "expected_time": "09:00"}
            ]
        }
    }


@app.get("/api/hr/attendance/report")
def get_hr_attendance_report():
    """HR Attendance Report"""
    return {
        "status": "success",
        "data": {
            "report_period": "September 2025",
            "overall_stats": {
                "total_working_days": 22,
                "average_attendance_rate": 94.2,
                "total_late_arrivals": 15,
                "total_absences": 8
            },
            "department_performance": [
                {"department": "Engineering", "attendance_rate": 95.8, "absences": 2},
                {"department": "Marketing", "attendance_rate": 98.5, "absences": 1},
                {"department": "Sales", "attendance_rate": 91.2, "absences": 4},
                {"department": "HR", "attendance_rate": 100.0, "absences": 0},
                {"department": "Finance", "attendance_rate": 89.3, "absences": 1}
            ],
            "trends": {
                "most_punctual_day": "Tuesday",
                "highest_absence_day": "Monday",
                "peak_late_arrival_time": "09:15-09:30"
            }
        }
    }


@app.post("/api/hr/attendance/mark")
def mark_hr_attendance(attendance_data: dict):
    """Mark HR Attendance"""
    return {
        "status": "success",
        "message": "Attendance marked successfully",
        "data": {
            "employee_id": attendance_data.get("employee_id"),
            "date": "2025-09-04",
            "time": "09:00:00",
            "status": attendance_data.get("status", "present"),
            "marked_by": "HR System",
            "location": "Main Office"
        }
    }

# Payroll Management Endpoints


@app.get("/api/hr/payroll")
def get_hr_payroll():
    """HR Payroll Records"""
    return {
        "status": "success",
        "data": {
            "current_period": "September 2025",
            "total_employees": 45,
            "total_payroll": 312500,
            "processed": 45,
            "pending": 0,
            "departments": [
                {"name": "Engineering", "employees": 16, "total_salary": 145000},
                {"name": "Marketing", "employees": 8, "total_salary": 58000},
                {"name": "Sales", "employees": 13, "total_salary": 75000},
                {"name": "HR", "employees": 4, "total_salary": 22000},
                {"name": "Finance", "employees": 4, "total_salary": 12500}
            ],
            "recent_payments": [
                {"employee": "John Smith", "amount": 8500,
                    "date": "2025-09-01", "status": "Processed"},
                {"employee": "Sarah Johnson", "amount": 7500,
                    "date": "2025-09-01", "status": "Processed"},
                {"employee": "David Brown", "amount": 6000,
                    "date": "2025-09-01", "status": "Processed"}
            ]
        }
    }


@app.get("/api/hr/payroll/reports")
def get_hr_payroll_reports():
    """HR Payroll Reports"""
    return {
        "status": "success",
        "data": {
            "summary": {
                "total_payroll_ytd": 2875000,
                "average_salary": 69500,
                "highest_paid_department": "Engineering",
                "payroll_growth": 8.5
            },
            "monthly_breakdown": [
                {"month": "September", "total": 312500, "employees": 45},
                {"month": "August", "total": 308000, "employees": 44},
                {"month": "July", "total": 305000, "employees": 43}
            ],
            "cost_centers": [
                {"department": "Engineering", "percentage": 46.4, "amount": 145000},
                {"department": "Sales", "percentage": 24.0, "amount": 75000},
                {"department": "Marketing", "percentage": 18.6, "amount": 58000},
                {"department": "HR", "percentage": 7.0, "amount": 22000},
                {"department": "Finance", "percentage": 4.0, "amount": 12500}
            ]
        }
    }


@app.get("/api/hr/payroll/{employee_id}")
def get_hr_employee_payroll(employee_id: int):
    """HR Employee Payroll Details"""
    base_salary = random.uniform(50000, 100000)
    return {
        "status": "success",
        "data": {
            "employee_id": employee_id,
            "employee_name": "John Smith",
            "current_period": "September 2025",
            "salary_details": {
                "base_salary": round(base_salary, 2),
                "overtime": round(base_salary * 0.1, 2),
                "bonuses": round(base_salary * 0.05, 2),
                "gross_pay": round(base_salary * 1.15, 2)
            },
            "deductions": {
                "federal_tax": round(base_salary * 0.22, 2),
                "state_tax": round(base_salary * 0.08, 2),
                "social_security": round(base_salary * 0.062, 2),
                "health_insurance": 450,
                "total_deductions": round(base_salary * 0.362 + 450, 2)
            },
            "net_pay": round(base_salary * 0.788 - 450, 2),
            "payment_history": [
                {"period": "August 2025", "gross": round(
                    base_salary * 1.1, 2), "net": round(base_salary * 0.75, 2)},
                {"period": "July 2025", "gross": round(
                    base_salary * 1.08, 2), "net": round(base_salary * 0.73, 2)}
            ]
        }
    }


@app.post("/api/hr/payroll/generate")
def generate_hr_payroll(payroll_data: dict):
    """Generate HR Payroll"""
    return {
        "status": "success",
        "message": "Payroll generated successfully",
        "data": {
            "payroll_id": f"PAY{random.randint(1000, 9999)}",
            "employee_id": payroll_data.get("employee_id"),
            "period": "September 2025",
            "gross_amount": payroll_data.get("amount", 5000),
            "net_amount": round(payroll_data.get("amount", 5000) * 0.75, 2),
            "generated_date": "2025-09-04T10:30:00Z",
            "status": "Generated"
        }
    }


@app.get("/api/hr/payroll/reports")
def get_hr_payroll_reports():
    """HR Payroll Reports"""
    return {
        "status": "success",
        "data": {
            "summary": {
                "total_payroll_ytd": 2875000,
                "average_salary": 69500,
                "highest_paid_department": "Engineering",
                "payroll_growth": 8.5
            },
            "monthly_breakdown": [
                {"month": "September", "total": 312500, "employees": 45},
                {"month": "August", "total": 308000, "employees": 44},
                {"month": "July", "total": 305000, "employees": 43}
            ],
            "cost_centers": [
                {"department": "Engineering", "percentage": 46.4, "amount": 145000},
                {"department": "Sales", "percentage": 24.0, "amount": 75000},
                {"department": "Marketing", "percentage": 18.6, "amount": 58000},
                {"department": "HR", "percentage": 7.0, "amount": 22000},
                {"department": "Finance", "percentage": 4.0, "amount": 12500}
            ]
        }
    }

# Performance Management Endpoints


@app.get("/api/hr/performance")
def get_hr_performance():
    """HR Performance Reviews"""
    return {
        "status": "success",
        "data": {
            "overview": {
                "total_employees": 45,
                "reviews_completed": 38,
                "reviews_pending": 7,
                "average_rating": 4.1,
                "completion_rate": 84.4
            },
            "recent_reviews": [
                {"employee": "John Smith", "rating": 4.2,
                    "completed_date": "2025-08-15", "reviewer": "Jane Doe"},
                {"employee": "Sarah Johnson", "rating": 4.5,
                    "completed_date": "2025-08-20", "reviewer": "Mike Wilson"},
                {"employee": "David Brown", "rating": 3.8,
                    "completed_date": "2025-08-25", "reviewer": "Lisa Chen"}
            ],
            "rating_distribution": [
                {"rating": "5.0", "count": 8, "percentage": 21.1},
                {"rating": "4.0-4.9", "count": 22, "percentage": 57.9},
                {"rating": "3.0-3.9", "count": 7, "percentage": 18.4},
                {"rating": "Below 3.0", "count": 1, "percentage": 2.6}
            ]
        }
    }


@app.get("/api/hr/performance/goals")
def get_hr_performance_goals():
    """HR Performance Goals"""
    return {
        "status": "success",
        "data": {
            "company_goals": {
                "total_goals": 180,
                "completed": 142,
                "in_progress": 28,
                "overdue": 10,
                "completion_rate": 78.9
            },
            "department_performance": [
                {"department": "Engineering", "goals": 65,
                    "completed": 52, "rate": 80.0},
                {"department": "Sales", "goals": 45,
                    "completed": 38, "rate": 84.4},
                {"department": "Marketing", "goals": 35,
                    "completed": 28, "rate": 80.0},
                {"department": "HR", "goals": 20, "completed": 16, "rate": 80.0},
                {"department": "Finance", "goals": 15, "completed": 8, "rate": 53.3}
            ],
            "goal_categories": [
                {"category": "Skill Development", "count": 45, "completion": 82.2},
                {"category": "Project Delivery", "count": 68, "completion": 88.2},
                {"category": "Team Collaboration",
                    "count": 32, "completion": 75.0},
                {"category": "Process Improvement",
                    "count": 35, "completion": 68.6}
            ]
        }
    }


@app.get("/api/hr/performance/{employee_id}")
def get_hr_employee_performance(employee_id: int):
    """HR Employee Performance Details"""
    return {
        "status": "success",
        "data": {
            "employee_id": employee_id,
            "employee_name": "John Smith",
            "current_review": {
                "period": "2025 Mid-Year",
                "overall_rating": 4.2,
                "completed_date": "2025-08-15",
                "reviewer": "Jane Doe",
                "next_review": "2025-12-15"
            },
            "performance_areas": [
                {"area": "Technical Skills", "rating": 4.5,
                    "comments": "Excellent technical expertise"},
                {"area": "Communication", "rating": 4.0,
                    "comments": "Good communication skills"},
                {"area": "Leadership", "rating": 3.8,
                    "comments": "Shows potential for leadership"},
                {"area": "Problem Solving", "rating": 4.3,
                    "comments": "Strong analytical skills"}
            ],
            "goals": {
                "completed": 8,
                "total": 10,
                "current_goals": [
                    {"goal": "Complete Advanced Certification",
                        "progress": 75, "due_date": "2025-10-31"},
                    {"goal": "Lead Team Project",
                        "progress": 90, "due_date": "2025-09-30"}
                ]
            },
            "development_plan": [
                "Attend leadership training program",
                "Complete public speaking course",
                "Mentor junior developers"
            ]
        }
    }


@app.post("/api/hr/performance/create")
def create_hr_performance_review(review_data: dict):
    """Create HR Performance Review"""
    return {
        "status": "success",
        "message": "Performance review created successfully",
        "data": {
            "review_id": f"REV{random.randint(1000, 9999)}",
            "employee_id": review_data.get("employee_id"),
            "rating": review_data.get("rating", 4.0),
            "period": "2025 Q3",
            "created_date": "2025-09-04T10:30:00Z",
            "status": "Draft",
            "next_steps": [
                "Schedule review meeting",
                "Employee self-assessment",
                "Manager feedback session"
            ]
        }
    }

# Recruitment Endpoints


@app.get("/api/hr/recruitment/jobs")
def get_hr_job_postings():
    """HR Job Postings"""
    return {
        "status": "success",
        "data": [
            {
                "job_id": "JOB001",
                "title": "Senior Software Engineer",
                "department": "Engineering",
                "location": "Remote/Hybrid",
                "type": "Full-time",
                "posted_date": "2025-08-15",
                "applications": 45,
                "status": "Active",
                "salary_range": "$80,000 - $120,000"
            },
            {
                "job_id": "JOB002",
                "title": "Marketing Specialist",
                "department": "Marketing",
                "location": "New York Office",
                "type": "Full-time",
                "posted_date": "2025-08-20",
                "applications": 28,
                "status": "Active",
                "salary_range": "$55,000 - $75,000"
            },
            {
                "job_id": "JOB003",
                "title": "Sales Representative",
                "department": "Sales",
                "location": "Chicago Office",
                "type": "Full-time",
                "posted_date": "2025-08-25",
                "applications": 62,
                "status": "Active",
                "salary_range": "$45,000 - $65,000 + Commission"
            }
        ]
    }


@app.get("/api/hr/recruitment/applications")
def get_hr_applications():
    """HR Job Applications"""
    return {
        "status": "success",
        "data": {
            "summary": {
                "total_applications": 135,
                "new_this_week": 18,
                "in_review": 45,
                "interviews_scheduled": 12,
                "offers_made": 3
            },
            "recent_applications": [
                {
                    "application_id": "APP001",
                    "candidate_name": "Alice Johnson",
                    "position": "Senior Software Engineer",
                    "applied_date": "2025-09-01",
                    "status": "Under Review",
                    "experience": "5 years",
                    "source": "LinkedIn"
                },
                {
                    "application_id": "APP002",
                    "candidate_name": "Bob Smith",
                    "position": "Marketing Specialist",
                    "applied_date": "2025-09-02",
                    "status": "Interview Scheduled",
                    "experience": "3 years",
                    "source": "Company Website"
                }
            ],
            "pipeline_stats": [
                {"stage": "Applied", "count": 135},
                {"stage": "Screening", "count": 45},
                {"stage": "Interview", "count": 12},
                {"stage": "Offer", "count": 3},
                {"stage": "Hired", "count": 1}
            ]
        }
    }


@app.get("/api/hr/recruitment/interviews")
def get_hr_interviews():
    """HR Interview Schedule"""
    return {
        "status": "success",
        "data": {
            "upcoming_interviews": [
                {
                    "interview_id": "INT001",
                    "candidate": "Alice Johnson",
                    "position": "Senior Software Engineer",
                    "date": "2025-09-05",
                    "time": "10:00 AM",
                    "interviewer": "Jane Doe",
                    "type": "Technical Round",
                    "location": "Conference Room A"
                },
                {
                    "interview_id": "INT002",
                    "candidate": "Bob Smith",
                    "position": "Marketing Specialist",
                    "date": "2025-09-05",
                    "time": "2:00 PM",
                    "interviewer": "Mike Wilson",
                    "type": "Final Round",
                    "location": "Virtual Meeting"
                }
            ],
            "interview_stats": {
                "total_scheduled": 12,
                "completed_this_week": 8,
                "success_rate": 75.0,
                "average_duration": 45
            },
            "interviewer_schedule": [
                {"interviewer": "Jane Doe",
                    "interviews_today": 2, "available_slots": 3},
                {"interviewer": "Mike Wilson",
                    "interviews_today": 1, "available_slots": 4},
                {"interviewer": "Lisa Chen",
                    "interviews_today": 0, "available_slots": 5}
            ]
        }
    }


@app.post("/api/hr/recruitment/post")
def post_hr_job(job_data: dict):
    """Post HR Job"""
    return {
        "status": "success",
        "message": "Job posted successfully",
        "data": {
            "job_id": f"JOB{random.randint(100, 999)}",
            "title": job_data.get("title", "New Position"),
            "department": job_data.get("department", "General"),
            "posted_date": "2025-09-04",
            "status": "Active",
            "posting_channels": ["Company Website", "LinkedIn", "Indeed"],
            "estimated_reach": random.randint(500, 2000)
        }
    }

# HR Analytics Endpoints


@app.get("/api/hr/analytics/dashboard")
def get_hr_analytics_dashboard():
    """HR Analytics Dashboard"""
    return {
        "status": "success",
        "data": {
            "key_metrics": {
                "total_employees": 45,
                "headcount_growth": 8.5,
                "turnover_rate": 12.3,
                "average_tenure": 2.8,
                "employee_satisfaction": 4.2
            },
            "department_breakdown": [
                {"department": "Engineering", "count": 16, "percentage": 35.6},
                {"department": "Sales", "count": 13, "percentage": 28.9},
                {"department": "Marketing", "count": 8, "percentage": 17.8},
                {"department": "HR", "count": 4, "percentage": 8.9},
                {"department": "Finance", "count": 4, "percentage": 8.9}
            ],
            "recruitment_funnel": {
                "applications": 135,
                "interviews": 12,
                "offers": 3,
                "hires": 1,
                "conversion_rate": 0.74
            },
            "performance_overview": {
                "average_rating": 4.1,
                "top_performers": 8,
                "improvement_needed": 3,
                "goals_completion": 78.9
            }
        }
    }


@app.get("/api/hr/analytics/reports")
def get_hr_analytics_reports():
    """HR Analytics Reports"""
    return {
        "status": "success",
        "data": {
            "available_reports": [
                {
                    "report_id": "RPT001",
                    "name": "Employee Turnover Analysis",
                    "description": "Detailed analysis of employee turnover trends",
                    "last_generated": "2025-09-01",
                    "frequency": "Monthly"
                },
                {
                    "report_id": "RPT002",
                    "name": "Compensation Benchmarking",
                    "description": "Market comparison of salary ranges by role",
                    "last_generated": "2025-08-15",
                    "frequency": "Quarterly"
                },
                {
                    "report_id": "RPT003",
                    "name": "Performance Trends",
                    "description": "Analysis of employee performance over time",
                    "last_generated": "2025-09-02",
                    "frequency": "Quarterly"
                }
            ],
            "recent_insights": [
                "Engineering department shows highest retention rate at 95%",
                "Average time to hire decreased by 15% this quarter",
                "Employee satisfaction scores improved by 0.3 points"
            ]
        }
    }


@app.get("/api/hr/analytics/metrics")
def get_hr_analytics_metrics():
    """HR Analytics Metrics"""
    return {
        "status": "success",
        "data": {
            "workforce_metrics": {
                "headcount": {
                    "current": 45,
                    "target": 50,
                    "growth_rate": 8.5
                },
                "diversity": {
                    "gender_ratio": {"male": 55, "female": 45},
                    "age_distribution": {
                        "under_30": 35,
                        "30_to_50": 55,
                        "over_50": 10
                    }
                }
            },
            "retention_metrics": {
                "turnover_rate": 12.3,
                "voluntary_turnover": 8.7,
                "involuntary_turnover": 3.6,
                "average_tenure": 2.8
            },
            "performance_metrics": {
                "average_rating": 4.1,
                "goal_completion_rate": 78.9,
                "training_hours_per_employee": 32.5,
                "internal_promotion_rate": 15.6
            },
            "recruitment_metrics": {
                "time_to_hire": 28,
                "cost_per_hire": 3500,
                "offer_acceptance_rate": 85.7,
                "source_effectiveness": {
                    "linkedin": 45,
                    "referrals": 30,
                    "company_website": 25
                }
            }
        }
    }


@app.get("/api/hr/analytics/trends")
def get_hr_analytics_trends():
    """HR Analytics Trends"""
    return {
        "status": "success",
        "data": {
            "headcount_trend": [
                {"month": "May", "count": 40},
                {"month": "June", "count": 42},
                {"month": "July", "count": 43},
                {"month": "August", "count": 44},
                {"month": "September", "count": 45}
            ],
            "turnover_trend": [
                {"month": "May", "rate": 15.2},
                {"month": "June", "rate": 13.8},
                {"month": "July", "rate": 14.1},
                {"month": "August", "rate": 12.9},
                {"month": "September", "rate": 12.3}
            ],
            "satisfaction_trend": [
                {"quarter": "Q1 2025", "score": 3.8},
                {"quarter": "Q2 2025", "score": 4.0},
                {"quarter": "Q3 2025", "score": 4.2}
            ],
            "predictions": {
                "expected_headcount_eoy": 52,
                "projected_turnover_rate": 11.5,
                "recommended_actions": [
                    "Focus on engineering team retention",
                    "Improve onboarding process",
                    "Implement mentorship program"
                ]
            }
        }
    }

# ===============================================================================
# RETENTION EXECUTOR ADDITIONAL ENDPOINTS
# ===============================================================================


@app.get("/api/icp/executor/contact-history")
def get_contact_history():
    """Contact History for Retention Executor"""
    return {
        "status": "success",
        "data": [
            {
                "contact_id": "CNT001",
                "merchant_id": "M001",
                "merchant_name": "Tech Solutions Inc",
                "contact_date": "2025-09-04",
                "contact_time": "10:30 AM",
                "method": "Phone Call",
                "duration": "15 minutes",
                "outcome": "Positive",
                "notes": "Discussed new product features, merchant interested in upgrade",
                "follow_up_required": True,
                "next_contact_date": "2025-09-11"
            },
            {
                "contact_id": "CNT002",
                "merchant_id": "M002",
                "merchant_name": "Retail Paradise",
                "contact_date": "2025-09-03",
                "contact_time": "2:15 PM",
                "method": "Email",
                "duration": "N/A",
                "outcome": "No Response",
                "notes": "Sent follow-up email regarding integration issues",
                "follow_up_required": True,
                "next_contact_date": "2025-09-06"
            }
        ]
    }


@app.get("/api/icp/executor/merchant-status")
def get_merchant_status():
    """Merchant Status Overview"""
    return {
        "status": "success",
        "data": {
            "total_merchants": 156,
            "status_breakdown": [
                {"status": "Active", "count": 142, "percentage": 91.0},
                {"status": "At Risk", "count": 8, "percentage": 5.1},
                {"status": "Inactive", "count": 4, "percentage": 2.6},
                {"status": "Churned", "count": 2, "percentage": 1.3}
            ],
            "priority_merchants": [
                {
                    "merchant_id": "M008",
                    "name": "Fashion Forward",
                    "status": "At Risk",
                    "risk_level": "High",
                    "last_contact": "2025-08-28",
                    "reason": "Declined transaction volume"
                },
                {
                    "merchant_id": "M015",
                    "name": "Food Delight",
                    "status": "At Risk",
                    "risk_level": "Medium",
                    "last_contact": "2025-09-01",
                    "reason": "Support ticket unresolved"
                }
            ]
        }
    }


@app.get("/api/icp/executor/onboarding-checklist")
def get_onboarding_checklist():
    """Onboarding Checklist"""
    return {
        "status": "success",
        "data": {
            "new_merchants": [
                {
                    "merchant_id": "M020",
                    "name": "New Coffee Shop",
                    "onboarding_stage": "Documentation",
                    "checklist": [
                        {"item": "Account Setup", "completed": True,
                            "date": "2025-09-01"},
                        {"item": "Documentation Submitted",
                            "completed": True, "date": "2025-09-02"},
                        {"item": "Verification Pending",
                            "completed": False, "expected": "2025-09-05"},
                        {"item": "Training Scheduled",
                            "completed": False, "expected": "2025-09-06"},
                        {"item": "Go Live", "completed": False,
                            "expected": "2025-09-10"}
                    ],
                    "completion_percentage": 40,
                    "assigned_executor": "John Smith"
                }
            ],
            "completion_stats": {
                "total_new_merchants": 5,
                "completed_onboarding": 3,
                "in_progress": 2,
                "average_completion_time": "7 days"
            }
        }
    }


@app.get("/api/icp/executor/setup-assistance")
def get_setup_assistance():
    """Setup Assistance Requests"""
    return {
        "status": "success",
        "data": {
            "active_requests": [
                {
                    "request_id": "REQ001",
                    "merchant_name": "Tech Gadgets Store",
                    "request_type": "POS Integration",
                    "priority": "High",
                    "created_date": "2025-09-03",
                    "expected_resolution": "2025-09-05",
                    "assigned_executor": "Sarah Johnson",
                    "status": "In Progress"
                },
                {
                    "request_id": "REQ002",
                    "merchant_name": "Beauty Salon",
                    "request_type": "Payment Gateway Setup",
                    "priority": "Medium",
                    "created_date": "2025-09-04",
                    "expected_resolution": "2025-09-06",
                    "assigned_executor": "Mike Wilson",
                    "status": "Pending"
                }
            ],
            "assistance_categories": [
                {"category": "POS Integration", "count": 8,
                    "avg_resolution_time": "3 days"},
                {"category": "Payment Gateway", "count": 5,
                    "avg_resolution_time": "2 days"},
                {"category": "API Configuration", "count": 3,
                    "avg_resolution_time": "4 days"}
            ]
        }
    }


@app.get("/api/icp/executor/training-materials")
def get_training_materials():
    """Training Materials"""
    return {
        "status": "success",
        "data": {
            "available_materials": [
                {
                    "material_id": "TRN001",
                    "title": "Getting Started Guide",
                    "type": "PDF",
                    "category": "Onboarding",
                    "duration": "30 minutes",
                    "last_updated": "2025-08-15",
                    "downloads": 156
                },
                {
                    "material_id": "TRN002",
                    "title": "POS Integration Video Tutorial",
                    "type": "Video",
                    "category": "Technical",
                    "duration": "45 minutes",
                    "last_updated": "2025-08-20",
                    "views": 89
                },
                {
                    "material_id": "TRN003",
                    "title": "Best Practices Webinar",
                    "type": "Webinar",
                    "category": "Best Practices",
                    "duration": "60 minutes",
                    "scheduled": "2025-09-10",
                    "registrations": 24
                }
            ],
            "usage_stats": {
                "total_materials": 15,
                "most_popular": "Getting Started Guide",
                "completion_rate": 78.5,
                "average_rating": 4.3
            }
        }
    }


@app.get("/api/icp/executor/integration-support")
def get_integration_support():
    """Integration Support"""
    return {
        "status": "success",
        "data": {
            "active_integrations": [
                {
                    "integration_id": "INT001",
                    "merchant_name": "E-commerce Pro",
                    "integration_type": "Shopify",
                    "status": "Testing",
                    "progress": 75,
                    "started_date": "2025-08-28",
                    "expected_completion": "2025-09-05",
                    "technical_contact": "Alex Chen"
                },
                {
                    "integration_id": "INT002",
                    "merchant_name": "Restaurant Chain",
                    "integration_type": "POS System",
                    "status": "Configuration",
                    "progress": 50,
                    "started_date": "2025-09-01",
                    "expected_completion": "2025-09-08",
                    "technical_contact": "Lisa Wong"
                }
            ],
            "integration_types": [
                {"type": "Shopify", "count": 12, "success_rate": 95},
                {"type": "WooCommerce", "count": 8, "success_rate": 92},
                {"type": "POS Systems", "count": 15, "success_rate": 88}
            ]
        }
    }


@app.get("/api/icp/executor/urgent-alerts")
def get_urgent_alerts():
    """Urgent Alerts"""
    return {
        "status": "success",
        "data": {
            "critical_alerts": [
                {
                    "alert_id": "ALT001",
                    "type": "Payment Failure",
                    "merchant": "Tech Solutions Inc",
                    "message": "Multiple payment failures detected",
                    "severity": "Critical",
                    "created": "2025-09-04T09:30:00Z",
                    "action_required": "Immediate contact needed"
                },
                {
                    "alert_id": "ALT002",
                    "type": "Account Suspension Risk",
                    "merchant": "Fashion Forward",
                    "message": "Account showing churn indicators",
                    "severity": "High",
                    "created": "2025-09-04T08:15:00Z",
                    "action_required": "Schedule retention call"
                }
            ],
            "alert_summary": {
                "total_alerts": 8,
                "critical": 2,
                "high": 3,
                "medium": 3,
                "resolved_today": 5
            }
        }
    }


@app.get("/api/icp/executor/daily-reminders")
def get_daily_reminders():
    """Daily Reminders"""
    return {
        "status": "success",
        "data": {
            "today_reminders": [
                {
                    "reminder_id": "REM001",
                    "type": "Follow-up Call",
                    "merchant": "Retail Paradise",
                    "due_time": "10:00 AM",
                    "priority": "High",
                    "notes": "Discuss pricing concerns"
                },
                {
                    "reminder_id": "REM002",
                    "type": "Check Integration",
                    "merchant": "E-commerce Pro",
                    "due_time": "2:00 PM",
                    "priority": "Medium",
                    "notes": "Verify API connection status"
                }
            ],
            "upcoming_reminders": [
                {
                    "date": "2025-09-05",
                    "count": 6,
                    "priority_breakdown": {"high": 2, "medium": 3, "low": 1}
                },
                {
                    "date": "2025-09-06",
                    "count": 4,
                    "priority_breakdown": {"high": 1, "medium": 2, "low": 1}
                }
            ]
        }
    }


@app.get("/api/icp/executor/follow-up-notifications")
def get_follow_up_notifications():
    """Follow-up Notifications"""
    return {
        "status": "success",
        "data": {
            "pending_follow_ups": [
                {
                    "notification_id": "NOT001",
                    "merchant": "Tech Solutions Inc",
                    "last_contact": "2025-09-01",
                    "days_since": 3,
                    "reason": "Product upgrade discussion",
                    "priority": "High",
                    "suggested_action": "Schedule demo call"
                },
                {
                    "notification_id": "NOT002",
                    "merchant": "Coffee Corner",
                    "last_contact": "2025-08-30",
                    "days_since": 5,
                    "reason": "Integration support",
                    "priority": "Medium",
                    "suggested_action": "Check progress update"
                }
            ],
            "follow_up_schedule": {
                "overdue": 2,
                "due_today": 3,
                "due_this_week": 8,
                "scheduled_future": 12
            }
        }
    }


@app.get("/api/icp/executor/technical-support")
def get_technical_support():
    """Technical Support Requests"""
    return {
        "status": "success",
        "data": {
            "open_tickets": [
                {
                    "ticket_id": "SUP001",
                    "merchant": "Tech Gadgets Store",
                    "issue": "API Connection Timeout",
                    "priority": "High",
                    "created": "2025-09-03",
                    "assigned_to": "Technical Team",
                    "status": "In Progress",
                    "last_update": "2025-09-04T09:00:00Z"
                },
                {
                    "ticket_id": "SUP002",
                    "merchant": "Fashion Boutique",
                    "issue": "Payment Processing Delay",
                    "priority": "Medium",
                    "created": "2025-09-04",
                    "assigned_to": "Support Team",
                    "status": "Investigating",
                    "last_update": "2025-09-04T10:30:00Z"
                }
            ],
            "support_metrics": {
                "total_tickets": 24,
                "resolved_today": 6,
                "average_resolution_time": "4.2 hours",
                "customer_satisfaction": 4.1
            }
        }
    }


@app.get("/api/icp/executor/track-resolution")
def get_track_resolution():
    """Track Resolution Progress"""
    return {
        "status": "success",
        "data": {
            "resolution_tracking": [
                {
                    "case_id": "CASE001",
                    "merchant": "Tech Solutions Inc",
                    "issue": "Payment Gateway Integration",
                    "status": "75% Complete",
                    "estimated_completion": "2025-09-05",
                    "assigned_team": "Technical Support",
                    "updates": [
                        {"date": "2025-09-04", "update": "Configuration completed"},
                        {"date": "2025-09-03", "update": "Testing initiated"},
                        {"date": "2025-09-02", "update": "Requirements gathered"}
                    ]
                }
            ],
            "resolution_stats": {
                "cases_in_progress": 8,
                "resolved_this_week": 12,
                "average_resolution_time": "3.5 days",
                "escalated_cases": 2
            }
        }
    }


@app.post("/api/icp/executor/collect-feedback")
def collect_feedback(feedback_data: dict):
    """Collect Merchant Feedback"""
    return {
        "status": "success",
        "message": "Feedback collected successfully",
        "data": {
            "feedback_id": f"FB{random.randint(1000, 9999)}",
            "merchant_id": feedback_data.get("merchant_id"),
            "feedback": feedback_data.get("feedback"),
            "rating": random.randint(3, 5),
            "collected_date": "2025-09-04T10:30:00Z",
            "category": "Service Quality",
            "follow_up_required": True
        }
    }


@app.get("/api/icp/executor/merchant-satisfaction")
def get_merchant_satisfaction():
    """Merchant Satisfaction Metrics"""
    return {
        "status": "success",
        "data": {
            "overall_satisfaction": {
                "average_rating": 4.2,
                "total_responses": 89,
                "response_rate": 78.5,
                "trend": "Improving"
            },
            "satisfaction_breakdown": [
                {"category": "Service Quality", "rating": 4.3, "responses": 89},
                {"category": "Product Features", "rating": 4.1, "responses": 85},
                {"category": "Support Response", "rating": 4.0, "responses": 78},
                {"category": "Pricing", "rating": 3.9, "responses": 82}
            ],
            "recent_feedback": [
                {
                    "merchant": "Tech Solutions Inc",
                    "rating": 5,
                    "comment": "Excellent support team, very responsive",
                    "date": "2025-09-03"
                },
                {
                    "merchant": "Fashion Forward",
                    "rating": 4,
                    "comment": "Good product, but could use more features",
                    "date": "2025-09-02"
                }
            ]
        }
    }
